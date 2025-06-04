import streamlit as st
import pickle
import os
from tkinter import Tk, filedialog as fd
import datetime
import pandas as pd
from copy import copy
import win32com.client
from openpyxl import load_workbook 

# =============================================================================
# File/Directory & System Utilities
# =============================================================================

def open_file_selection(initialdir='', filter_name='*.*'):
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    filetypes = (
        (filter_name, f"*{filter_name.split(' ')[0]}*.*"),
        ('All files', '*.*'),
        ('Excel', '*.xlsx'),
        ('CSV', '*.doc')
    )
    files = fd.askopenfilenames(filetypes=filetypes, initialdir=initialdir)
    root.destroy()
    return files

def select_directory(initialdir):
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    directory = fd.askdirectory(initialdir=initialdir)
    root.destroy()
    return directory

def set_paths(path):
    output_paths = {}
    output_paths['path_xl_format'] = os.path.join(path, 'columns and formatting.xlsx')
    output_paths['path_plan'] = os.path.join(path, 'manufacturing plan.xlsx')
    output_paths['path_assigned'] = os.path.join(path, 'status de ordenes.xlsx')
    output_paths['path_report'] = os.path.join(path, 'reporte de manufactura.xlsx')
    return output_paths

def set_col_rel(output_paths):
    df_columns = read_excel(output_paths['path_xl_format'], sheet_name='column_format')
    df_col_rel = df_columns[~df_columns['std_name'].isnull()].copy()
    return {'col_rel': df_col_rel, 'columns': df_columns}

# =============================================================================
# Excel Management Functions
# =============================================================================

def read_excel(path=None, sheet_name=0, header=0, keep_default_na=True, dtype=None):
    with pd.ExcelFile(path) as xls:
        df = pd.read_excel(path, sheet_name=sheet_name, header=header, keep_default_na=keep_default_na, dtype=dtype)
    return df

def load_excel_with_header_key(file_path, sheet_name=0, key_text='', dtype=None, **kwargs):
    df = read_excel(file_path, sheet_name=sheet_name, keep_default_na=False, dtype=dtype)
    header_row = None
    if key_text in df.columns:
        header_row = 0
    else:
        for col in df.columns:
            for i, cell in df[col].items():
                if pd.notna(cell) and key_text in str(cell):
                    header_row = i
                    break
            if header_row is not None:
                break
        if header_row is None:
            raise ValueError(f"Key text '{key_text}' not found in the sheet.")
        df.columns = df.iloc[header_row]
        df = df.iloc[header_row + 1:].reset_index(drop=True)
    return df

def find_cell_by_text(ws, text):
    """
    Find the cell containing the specified date in the worksheet.
    
    :param ws: The worksheet object from openpyxl
    :param date_value: The date value to search for (datetime object or string)
    :return: The cell address (e.g., 'B2') or None if not found
    """
    # for row in ws.iter_rows():
    for row in ws[ws.calculate_dimension()]:
        for cell in row:
            if cell.value == text:
                return cell.coordinate  # Return cell address (e.g., 'B2')
    return None 

def get_column_info(ws, col_name, raise_error=True):
    """
    Returns a dictionary with the column cell and the data range for the column.
    """
    col_cell=find_cell_by_text(ws, col_name)
    if not col_cell:
        if not raise_error:
            return False
        st.error(f"Column '{col_name}' not found in the sheet.")
        st.stop()
    col_cell=ws[col_cell]
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    data_range=ws[col_cell.offset(1,0).coordinate:ws.cell(ws[last_cell].row,col_cell.offset(1,0).column).coordinate]
    data_range_list=[cell[0] for cell in data_range]
    return {'data_range':data_range_list,
            'col_cell':col_cell}

def get_col_sizes(wb):
    """
    Obtains a dictionary with the column sizes of each sheet in a workbook
    """
    col_sizes={}
    for sheet in wb.sheetnames:
        col_sizes[sheet]=wb[sheet].column_dimensions
    return col_sizes

def apply_col_sizes(wb,col_sizes):
    """
    Applies the column sizes obtained by function get_col_sizes
    """
    for sheet in col_sizes.keys():
        if sheet not in wb.sheetnames:
            continue
        ws=wb[sheet]
        for column_letter in col_sizes[sheet].keys():
            if column_letter==0:
                continue
            ws.column_dimensions[column_letter].width=col_sizes[sheet][column_letter].width
    return wb

def get_cell_properties(cell):
    properties = {}
    fill = cell.fill
    properties["background_color"] = fill
    properties["font"]=cell.font
    properties["alignment"] = cell.alignment
    return properties

def format_cell(cell,properties):
    "Apply properties based on the dict of prooperties"
    cell.fill=copy(properties['background_color'])
    cell.font = copy(properties["font"])
    cell.alignment = copy(properties["alignment"])

def get_xl_formatting(table_name=None):
    """
    Returns the formatting for the excel file defined in columns and formatting.xlsx
    return example
    {'oor':
    {
        'columna':{
            'head':{properties},
            'data':{properties}
        }
    }
    }
    """
    wb = load_workbook(st.session_state.output_paths['path_xl_format'])
    ws = wb['column_format']
    col_info=get_column_info(ws,'column_name')
    cell_properties={}
    for cell in col_info['data_range']:
        head_properties=get_cell_properties(cell)
        data_properties=get_cell_properties(cell.offset(0, 1))
        if cell.offset(0, -1).value not in cell_properties:
            cell_properties[cell.offset(0, -1).value]={} #Table name
        if cell.value not in cell_properties[cell.offset(0, -1).value]:
            cell_properties[cell.offset(0, -1).value][cell.value]= {} #Column name
        cell_properties[cell.offset(0, -1).value][cell.value]['head_properties']=head_properties
        cell_properties[cell.offset(0, -1).value][cell.value]['data_properties']=data_properties
    ws = wb['special_format']
    col_info=get_column_info(ws,'format_name')
    cell_properties['special_format']={}
    for cell in col_info['data_range']:
        cell_properties['special_format'][cell.value]=get_cell_properties(cell)

    if table_name:
        cell_properties=cell_properties[table_name]
    return cell_properties  
# =============================================================================
# Persistence / State Management
# =============================================================================

def save_state_pickle(state, filename='folder_state.pkl'):
    with open(filename, 'wb') as f:
        pickle.dump(state, f)

def load_state_pickle(filename='folder_state.pkl'):
    try:
        with open(filename, 'rb') as f:
            return pickle.load(f)
    except FileNotFoundError:
        return {"folder_output": None, "selections": {}, "initial_date": datetime.date.today()}

def is_file_open(filepath):
    # Check if the file exists
    if not os.path.exists(filepath):
        return False  # File does not exist, so treat it as "open" for your logic

    try:
        # Try to open the file in write mode
        with open(filepath, 'a'):
            pass
        return False  # File is not open (no exception raised)
    except PermissionError:
        return True 
    
def close_xl_if_open(path):
    if is_file_open(path):
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            workbook = excel.Workbooks(path)
            workbook.Save()
            workbook.Close()
        except:
            st.error(f"Cerrar el archivo: {path}")
            st.stop()   


# =============================================================================
# DataFrame Management & Data Processing
# =============================================================================

def rename_columns(df, df_col_rel, table_from='std_name', sheet_from='only', table_to='std_name', sheet_to='only'):
    required_cols = {'table', 'sheet', 'column_name', 'std_name'}
    missing_cols = required_cols - set(df_col_rel.columns)
    if missing_cols:
        raise ValueError(f"df_col_rel is missing required columns: {missing_cols}")
    mapping = {}
    if table_to == 'std_name':
        filtered = df_col_rel[(df_col_rel['table'] == table_from) & (df_col_rel['sheet'] == sheet_from)]
        if filtered.empty:
            raise ValueError("No matching mapping found for given table_from and sheet_from.")
        mapping = filtered.set_index('column_name')['std_name'].to_dict()
    elif table_from == 'std_name':
        filtered = df_col_rel[(df_col_rel['table'] == table_to) & (df_col_rel['sheet'] == sheet_to)]
        if filtered.empty:
            raise ValueError("No matching mapping found for given table_to and sheet_to.")
        mapping = filtered.set_index('std_name')['column_name'].to_dict()
    else:
        filtered_from = df_col_rel[(df_col_rel['table'] == table_from) & (df_col_rel['sheet'] == sheet_from)]
        if filtered_from.empty:
            raise ValueError("No matching mapping found for given table_from and sheet_from.")
        mapping_from = filtered_from.set_index('column_name')['std_name'].to_dict()
        df = df.rename(columns=mapping_from)
        filtered_to = df_col_rel[(df_col_rel['table'] == table_to) & (df_col_rel['sheet'] == sheet_to)]
        if filtered_to.empty:
            raise ValueError("No matching mapping found for given table_to and sheet_to.")
        mapping = filtered_to.set_index('std_name')['column_name'].to_dict()
    return df.rename(columns=mapping)

def format_dates(df, date_cols=[],type='iso'):
    for col in date_cols:
        if not col in df.columns:
            continue
        if type=='iso':
            df[col]=pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
        else:
            df[col]=pd.to_datetime(df[col], errors='coerce')
    return df

def read_predefined_excel(path,df_columns=pd.DataFrame(),table='',sheet='only',check_mandatory=False):
    """
    If the file exists, it reads it and returns the dataframe, 
    otherwise it returns an empty dataframe but with the columns of the file
    it checks mandatory columns according to columns and formatting file but rises it only if indicated
    """
    if not os.path.exists(path):
        df=get_predefined_df(df_columns=df_columns,
                             table=table,
                             sheet=sheet)
    else:
        df=read_excel(path)    
        if check_mandatory:
            check_mandatory_columns_df(df.columns,df_columns=df_columns,table=table,sheet=sheet)
    return df

def get_predefined_df(df_columns=pd.DataFrame(),table='',sheet='only'):
    cols=df_columns[(df_columns['table']==table)&
                                        (df_columns['sheet']==sheet)]['column_name'].to_list()
    if len(cols)==0:
        st.error(f"Tabla: {table} no definida en archivo columns and formatting")
        st.stop()
    df=pd.DataFrame(columns=cols)
    return df

def check_mandatory_columns_df(cols=[],df_columns=pd.DataFrame(),table='',sheet='only'):
    """
    Check mandatory columns against the columns marked in
    columns and formatting file
    """
    mandatory_cols=df_columns[(df_columns['table']==table)&
                                        (df_columns['sheet']==sheet)&
                                        (~df_columns['mandatory_column'].isna())]['column_name'].to_list()
    missing_columns = [col for col in mandatory_cols if col not in cols]
    if len(missing_columns)>0:
        st.error(f"No se encontraron las siguientes columnas en el archivo {table}: {missing_columns}")
        st.stop()

def append_df_to_df(df_new=pd.DataFrame(),df_old=pd.DataFrame(),table='',keys=[],date_cols=[],allow_duplicates=False):
    """
    Appends a dataframe to an existing Dataframe
    keys: Fields that should not be duplicated, the old records are kept
    table: Table which is source of the dataframe, just to show it in the error
    date_cols: Columns transformed to date
    """
    if df_new.empty:
        return
    if not keys:
        keys=df_new.columns
    df_new=df_new.copy()
    df_new=format_dates(df_new,date_cols)
    df_old=format_dates(df_old,date_cols)
    df_new=get_common_records(df_new,df_old,keys=keys,how='uncommon')
    df_old=pd.concat([df_old,df_new])  
    df_old.reset_index(inplace=True,drop=True)
    df_grp=df_old.groupby(keys).count()
    if allow_duplicates==True:
        return df_old
    df_grp=df_grp[df_grp[df_grp.columns[0]]>1]
    if len(df_grp)>0:
        df_grp.reset_index(inplace=True)
        st.error(f"Hay duplicados en el archivo {table} para la llave {keys}:")
        st.error(df_grp[keys].sort_values(keys))
        st.stop()
    return df_old

def get_common_records(df_new=pd.DataFrame(),df_old=pd.DataFrame(),keys=[],how='common'):
    """
    Gets common records in two dataframes based on a key of columns
    when getting uncommon only records the records of the new are returned
    """
    df_old=df_old.copy()
    df_new=df_new.copy()
    df_old['composite_key'] = list(zip(*(df_old[col] for col in keys)))
    df_new['composite_key'] = list(zip(*(df_new[col] for col in keys)))
    if how=='common':
        df_new=df_new[df_new['composite_key'].isin(df_old['composite_key'])]
    else:
        df_new=df_new[~df_new['composite_key'].isin(df_old['composite_key'])]
    df_new.drop(columns=['composite_key'],inplace=True)
    return df_new
