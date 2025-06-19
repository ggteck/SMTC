"""
# Carga de demanda
## Toma la demanda del cliente, el reporte de POs firmes del dia y descuenta el BOM para generar la carga de la demanda
- V8. 2025-06-18
    - Migración a streamlit
- V7. 2024-11-09
    - Los ensambles sin BOM no se descuentan del forecast, los accesorios se agregaran al bom con nivel 0
    - Se genera un archivo KorrusFile_Pendientes.xlsx con lo pendiente de considerar debido a falta de BOM
    - Se permite seleccionar mas de un KorrusFile para integrarlos al mismo tiempo
- V6. 2024-11-07
    - Solo se hacen cambios en registros del Write Forcast con FC
- V5. 2024-11-06
    - Agregar selección de fecha para arcivo Korrus, las POs se agregaran al mes de la fecha seleccionada
    - Si no hay forecast para cierto mes, no aparecen las POs a descontar ese mes en el Forecast Apollo
    - Los ensambles sin BOM se toman como componentes y se intentan descontar del Forecast
    - El archivo Shipments no es mandatorio
    - Corrección de errores al abrir Reportes.xlsx
- V4. 2024-04-11
    - Permitir carga de Korrus file separado por comas
- V3. 2024-10-31
    - Descuento de POs embarcadas del write forecast
    - Reporte de POs no encontradas en write forecast
- V2. 2024-10-28.
    - Se genera reporte de missing boms y Alerta en el punto 2
    - Se elimina el filtro de Trov forecast para incluir Rise Forecast
    - Se agregan las POs en el write forecast
- V1. 2024-10-23. Version inicial 
"""


# Seleccion de archivos de trabajo
import pandas as pd
import os
import ipywidgets as widgets
import pickle
import streamlit as st
import warnings
import gc
from ipyfilechooser import FileChooser
from IPython.display import display
import datetime
from openpyxl.cell import MergedCell
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.formatting.rule import CellIsRule

from tkinter import Tk, filedialog as fd
from pandas.tseries.offsets import BDay
from pathlib import Path
from streamlit.errors import StreamlitAPIException 
from copy import copy 
_THIS_PAGE = Path(__file__).stem  

prev = st.session_state.get("_active_page")

if prev is None or prev != _THIS_PAGE:
    # First load OR coming from a different page → purge everything
    for k in list(st.session_state.keys()):
        del st.session_state[k]

# Record that we're now on this page
st.session_state["_active_page"] = _THIS_PAGE

warnings.filterwarnings("ignore")

# -------------------------------------------------------------------
# =============================================================================
# File/Directory & System Utilities
# =============================================================================

def open_file_selection(initialdir='', filter_name='*.*'):
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    filetypes = (
        (filter_name, f"*{filter_name.split(' ')[0]}*.*"),
        ('All files', '*.*'),
        ('Excel', '*.xlsx'),
        ('CSV', '*.doc')
    )
    files = fd.askopenfilenames(filetypes=filetypes, initialdir=initialdir)
    root.destroy()
    return files

def set_paths(path):
    output_paths = {}
    output_paths['path_xl_format'] = os.path.join(path, 'columns and formatting.xlsx')
    return output_paths


def select_directory(initialdir):
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    directory = fd.askdirectory(initialdir=initialdir)
    root.destroy()
    return directory
   
def manage_file_selector(selector_key, display_label, state):
    button_ph = st.empty()
    msg_ph = st.empty()
    if not state["selections"].get(selector_key):
        if button_ph.button(f"{display_label} File", key=f"select_{selector_key}"):
            files = open_file_selection(initialdir=state["folder_output"] or os.getcwd())
            if files:
                state["selections"][selector_key] = files[0]
                save_state_pickle(state,filename=path_pickle)
                st.rerun()
        msg_ph.info(f"{display_label} no seleccionado.")
    else:
        st.success(f"Selected {display_label} File: {state['selections'][selector_key]}")
        if button_ph.button(f"Change {display_label} File", key=f"change_{selector_key}"):
            state["selections"][selector_key] = ""
            save_state_pickle(state,filename=path_pickle)
            st.rerun()

# =============================================================================
# Persistence / State Management
# =============================================================================

def save_state_pickle(state, filename='folder_state.pkl'):
    with open(filename, 'wb') as f:
        pickle.dump(state, f)

def load_state_pickle(filename='folder_state.pkl'):
    try:
        with open(filename, 'rb') as f:
            return pickle.load(f)
    except FileNotFoundError:
        return {
            "folder_output": None,
            "selections": {}
        }

# Decorator to handle the permission error
def handle_permission_error_with_popup(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except PermissionError as e:
            if e.errno == 13:  # Permission denied error
                st.error(f"Error: {e}\nFavor de cerrar el archivo.")
    return wrapper

# Decorate the function where the file is being saved
@handle_permission_error_with_popup
def save_df(df, filepath,sheet_name,index):
    df.to_excel(filepath,sheet_name=sheet_name,index=index)

@handle_permission_error_with_popup
def save_wb(wb, filepath):
    wb.save(filepath)

@handle_permission_error_with_popup
def append_sheet(df,path,sheet_name,index):
    with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index) 

##### Excel management functions
def find_cell_by_text(ws, text):
    """
    Find the cell containing the specified date in the worksheet.
    
    :param ws: The worksheet object from openpyxl
    :param date_value: The date value to search for (datetime object or string)
    :return: The cell address (e.g., 'B2') or None if not found
    """
    # for row in ws.iter_rows():
    for row in ws[ws.calculate_dimension()]:
        for cell in row:
            if cell.value == text:
                return cell.coordinate  # Return cell address (e.g., 'B2')
    return None 

def find_all_cells_by_text(ws, text):
    """
    Find the cell containing the specified date in the worksheet.
    
    :param ws: The worksheet object from openpyxl
    :param date_value: The date value to search for (datetime object or string)
    :return: The cell address (e.g., 'B2') or None if not found
    """
    # for row in ws.iter_rows():
    cells=[]
    for row in ws[ws.calculate_dimension()]:
        for cell in row:
            if text in str(cell.value):
                cells.append(cell.coordinate)
    return cells 

def set_number_format(ws,col_name,format):
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    col_head_address=find_cell_by_text(ws,col_name)
    for row in ws[f"{col_head_address}:{last_cell}"]:
        row[0].number_format = format

def fill_column(ws,col_name,fill):
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    col_head_address=ws[find_cell_by_text(ws,col_name)].offset(1,0).coordinate
    for row in ws[f"{col_head_address}:{last_cell}"]:
        row[0].fill=fill

def font_column(ws,col_name,font):
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    col_head_address=ws[find_cell_by_text(ws,col_name)].offset(1,0).coordinate
    for row in ws[f"{col_head_address}:{last_cell}"]:
        row[0].font=font    



def fill_formula(ws,col_name,formula):
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    col_head_address=ws[find_cell_by_text(ws,col_name)].offset(1,0).coordinate
    for row in ws[f"{col_head_address}:{last_cell}"]:
        row[0].value=formula

def is_file_open(filepath):
    # Check if the file exists
    if not os.path.exists(filepath):
        return False  # File does not exist, so treat it as "open" for your logic

    try:
        # Try to open the file in write mode
        with open(filepath, 'a'):
            pass
        return False  # File is not open (no exception raised)
    except PermissionError:
        return True 

     
# Agregar las columnas criticas por archivo
mandatory_cols={'BOM':[
                            'BOM',
                            'Component',
                            'Component Description',
                            'Qty Per'
                            ],
                'KorrusFile':[ 
                       'PurchaseOrder', 
                       'PODate', 
                       'Quantity',
                       'ProductServiceID'
                       ],
                'Apollo':['Part Number',
                          'Description'],
                'Write Forecast':[
                        'Part',
                        'Part Site',
                        'Order Id',
                        'O/T',
                        'Customer',
                        'LN/DEL',
                        'Due Date',
                        'Promised Date',
                        'Quantity',
                        'Shipped Qty',
                        'Unit Selling Price',
                        'Del#',
                        'Status',
                        'Org. Duedate',
                        'Model',
                        'Order Site',
                        'U_REFRNC'
                ],
                'Shipment':[
                    'Part No.',
                    'Customer PO#',
                    'Qty. Shipped'
                    ]
                }


def check_mandatory_cols(cols,selector_name):
    missing_columns = [col for col in mandatory_cols[selector_name] if col not in cols]
    if len(missing_columns)>0:
        st.error(f"No se encontraron las siguientes columnas en el archivo {selector_name}: {missing_columns}")
        st.stop()
    return


# Cell colors
light_green='99FF99'
avocato_green='E2EFDA'
dark_blue='002060'
light_blue='DDEBF7'
grey='595959'
dark_red='9C0006'
melon='FCE4D6'
white='FFFFFF'
light_pink='FFC7CE'
light_yellow='FFF2CC'
light_grey='E8E8E8'

# =============================================================================
# Main Functions
# =============================================================================

def generate_demand():
    path_pos_bom_raw=os.path.join(state['folder_output'],'Pos_bom_raw.xlsx')
    path_apollo_updated=os.path.join(state['folder_output'],'Forecast de Apollo actualizado.xlsx')
    path_reportes=os.path.join(state['folder_output'],'Reportes.xlsx')
    open_files = [file for file in [path_pos_bom_raw,path_apollo_updated,path_reportes] if is_file_open(file)]
    if open_files:
        st.info(f"Favor de cerrar los siguientes archivos: {', '.join(open_files)}")
        st.stop()


    path_pos=state.get('korrus_folder')
    if not path_pos:
        st.info('Favor de seleccionar al menos un KorrusFile')
        st.stop()
    list_pos=os.listdir(path_pos)
    df_pos=pd.DataFrame()
    for po_file in list_pos:
        if not '850KorrusFile' in po_file:
            continue
        if os.path.splitext(os.path.basename(po_file))[1].lower()==".xlsx":
            df = pd.read_excel(os.path.join(path_pos,po_file))
        else:
            try:
                df = pd.read_csv(os.path.join(path_pos,po_file), sep='\t', encoding='utf-16')
            except:
                df = pd.read_csv(os.path.join(path_pos,po_file), sep=',')
        check_mandatory_cols(df.columns,'KorrusFile')
        df_pos=pd.concat([df,df_pos])

    df_pos.reset_index(drop=True,inplace=True)


    df_korrus=df_pos.copy()
    df_pos=df_pos[mandatory_cols['KorrusFile']]
    df_pos = df_pos[pd.to_numeric(df_pos['Quantity'], errors='coerce').notnull()]
    df_pos['Quantity']=df_pos['Quantity'].astype(float)
    df_pos['ProductServiceID']=df_pos['ProductServiceID'].str.upper()

    df_pos_bom=df_pos.groupby(['ProductServiceID']).sum('Quantity')
    df_pos_bom.reset_index(inplace=True)

    # Cambiar por selector si es necesario
    df_pos_bom['extract_date']=pd.to_datetime(state['korrus_date']).strftime('%Y-%m-%d')
    df_pos_bom['year_month']=pd.to_datetime(state['korrus_date']).strftime('%Y-%m')

    if df_pos['PODate'].min() == df_pos['PODate'].max():
        df_pos_bom['head_date'] = f"POs {df_pos['PODate'].min()}"
    else:
        df_pos_bom['head_date'] = f"POs {df_pos['PODate'].min()} - {df_pos['PODate'].max()}"

    path_bom=st.session_state.selected_paths['bom_file']
    df_bom=pd.read_excel(path_bom)
    check_mandatory_cols(df_bom.columns,'BOM')

    df_pos_bom=df_pos_bom.merge(df_bom,how='left',left_on=['ProductServiceID'],right_on=['BOM'])


    # Reporte de missing boms
    df_missing_bom=df_pos_bom[df_pos_bom['BOM'].isnull()]
    if df_missing_bom.shape[0]>0:
        df_missing_bom=df_missing_bom[['ProductServiceID','Quantity','extract_date','year_month','head_date']]
        # Eliminar POs de productos sin BOM, df_po se agregara al Write Forecast
        df_pos=df_pos[~df_pos['ProductServiceID'].isin(df_missing_bom['ProductServiceID'])]
        save_df(df_missing_bom,filepath=path_reportes,sheet_name='Missing Boms',index=False)
        
        st.info(f"Advertencia: Se encontraron Boms faltantes, revisar el archivo: {path_reportes}")


    df_korrus=df_korrus[df_korrus['ProductServiceID'].isin(df_missing_bom['ProductServiceID'])]
    path_korrus_peding=os.path.join(state["folder_output"],f'KorrusFile_Pendientes_{datetime.date.today().strftime('%Y-%m-%d')}.xlsx')
    save_df(df_korrus,filepath=path_korrus_peding,sheet_name='KorrusFile',index=False)

    # Se eliminan los ensambles sin BOMS, los accesorios se agregaran al bom como nivel 0
    df_pos_bom=df_pos_bom[~df_pos_bom['BOM'].isnull()]


    df_pos_bom['Extended']=df_pos_bom['Qty Per']*df_pos_bom['Quantity']

    df_pos_bom=df_pos_bom.groupby(['extract_date','year_month','head_date','Component']).sum('Extended')
    df_pos_bom.reset_index(inplace=True)
    df_pos_bom=df_pos_bom[['extract_date','year_month','head_date','Component','Extended']]
    df_pos_bom.rename({'Component':'Part Number'},inplace=True,axis=1)
    #Se guarda el archivo sin cambios y sin consolidar para descontar luego del write forecast
    df_pos_bom_selected=df_pos_bom.copy()
    # Se consolida con el archivo existente, si es del mismo dia se reemplaza
    if os.path.exists(path_pos_bom_raw):
        df=pd.read_excel(path_pos_bom_raw)
        df=df[~df['extract_date'].isin(df_pos_bom['extract_date'])]
        if df.shape[0]>0:
            df_pos_bom=pd.concat([df,df_pos_bom])
    save_df(df=df_pos_bom,filepath=path_pos_bom_raw,sheet_name='forecast_apollo',index=False)

    # ## 2.2 Explosionar archivo de Apollo


    # Archivo de apollo
    path_forecast=st.session_state.selected_paths['apollo_file']
    open_files = [file for file in [path_forecast,path_reportes] if is_file_open(file)]
    if open_files:
        st.info(f"Favor de cerrar los siguientes archivos: {', '.join(open_files)}")
        st.stop()

    forecast_wb=pd.read_excel(path_forecast,sheet_name=None)

    df_forecast=pd.DataFrame()
    for key in forecast_wb.keys():
        if 'Forecast' in key:
            df=forecast_wb[key]
            df['source']=key
            df_forecast=pd.concat([df,df_forecast])
    check_mandatory_cols(df_forecast.columns,'Apollo')
    # # Por ahora filtro solo Trov (Validar)
    date_cols=[col for col in df_forecast.columns if type(col)==datetime.datetime]
    df_forecast_raw=pd.DataFrame()
    for col in date_cols:
        df=df_forecast[['source','Type','Part Number','Description']+[col]]
        df.rename({col:'Qty'},axis=1,inplace=True)
        df['year_month']=col.strftime('%Y-%m')
        df_forecast_raw=pd.concat([df,df_forecast_raw])

    df_pos_bom=df_pos_bom[df_pos_bom['year_month'].isin(df_forecast_raw['year_month'])]

    if df_forecast_raw.shape[0]==0:
        st.info(f"No se encontraron datos en el archivo: {path_forecast}")
        st.stop()    
    pivot_forecast = pd.pivot_table(df_forecast_raw, 
                        values=['Qty'],  
                        index=['Type','Part Number','Description'],         
                        columns=['year_month'],       
                        aggfunc='sum',
                        fill_value=0
                            ) 
    pivot_po_bom = pd.pivot_table(df_pos_bom, 
                        values=['Extended'],  
                        index=['Part Number'],         
                        columns=['head_date'],       
                        aggfunc='sum',
                        fill_value=0
                            ) 


    pivot_forecast.reset_index(inplace=True)
    pivot_po_bom.reset_index(inplace=True)

    pivot_forecast=pivot_forecast.merge(pivot_po_bom,how='left',on=['Part Number']).fillna(0)


    pivot_forecast.columns=pivot_forecast.columns.droplevel(0)

    pivot_forecast.columns.values[0:3] = ['Type', 'Part Number', 'Description']

    #Ordenar columnas deacuerdo al mes
    df_pos_bom_cols=df_pos_bom[['year_month','head_date']].drop_duplicates()
    df_forecast_raw_cols=df_forecast_raw[['year_month']].drop_duplicates()
    df_forecast_raw_cols['head_date']=df_forecast_raw_cols['year_month']

    df_new_demand_cols=df_pos_bom_cols.copy()
    df_new_demand_cols['head_date']='xNew '+df_new_demand_cols['year_month']
    df_new_demand_cols.reset_index(inplace=True,drop=True)
    df_pos_bom_cols.reset_index(inplace=True,drop=True)

    df_pos_bom_cols=pd.concat([df_pos_bom_cols,df_new_demand_cols]).sort_index()

    df_cols=pd.concat([df_forecast_raw_cols,df_pos_bom_cols])
    df_cols.reset_index(inplace=True)
    df_cols.sort_values(['year_month','index','head_date'],inplace=True)


    ordered_cols=df_cols['head_date'].to_list()
    pivot_forecast=pivot_forecast.reindex(['Type','Part Number','Description']+ordered_cols,axis=1)

    save_df(pivot_forecast,filepath=path_apollo_updated,sheet_name='ForecastApollo',index=False)


    # ### Formato para forecast de apollo

    if is_file_open(path_apollo_updated):
        st.info(f"Favor de cerrar un archivo {path_apollo_updated}")
        st.stop()  

    wb = load_workbook(path_apollo_updated)
    ws = wb['ForecastApollo']
    fill = PatternFill(start_color=light_grey, end_color=light_grey, fill_type="solid")
    cell=ws[find_cell_by_text(ws,'Description')]
    cell.fill=fill
    cell=ws[find_cell_by_text(ws,'Part Number')]
    cell.fill=fill

    fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
    font = Font(size=11,name='Calibri',color=dark_red) 
    ws.conditional_formatting.add(ws.calculate_dimension(),  
        CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=fill, font=font))


    for column in ws.columns:
        if column[0].column<=3:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) 
            ws.column_dimensions[column_letter].width = adjusted_width
            

    for cel in ws[ws.calculate_dimension()][0]:
        cel.alignment=Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Formulas
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    new_month_cells=find_all_cells_by_text(ws,'xNew')

    for cel in new_month_cells:
        col_head_address=ws[cel].offset(1,0).coordinate
        cell_a=ws[col_head_address].offset(0,-2).coordinate
        cell_b=ws[col_head_address].offset(0,-1).coordinate
        for row in ws[f"{col_head_address}:{last_cell}"]:
            row[0].value=f"={ws.cell(row[0].row,ws[cell_a].column).coordinate}-{ws.cell(row[0].row,ws[cell_b].column).coordinate}"

    save_wb(wb=wb,filepath=path_apollo_updated)
    try:
        os.startfile(path_apollo_updated)
    except:
        st.info(f"El archivo Forecast de apollo esta listo.")


    # ## Actualizar Write Forecast
    # - Resta el bom de las POs del archivo seleccionado al write forecast
    # - Las lineas menores o iguales a cero se eliminan


    # Write forecast
    path_write_forecast_new=os.path.join(state["folder_output"],'Write_forecast_updated.xlsx')
    path_write_forecast_changes=os.path.join(state["folder_output"],'Write_forecast_changes.xlsx')
    path_write_forecast=st.session_state.selected_paths['write_forecast']

    open_files = [file for file in [path_write_forecast_new, path_write_forecast_changes,path_write_forecast,path_reportes] if is_file_open(file)]
    if open_files:
        st.info(f"Favor de cerrar los siguientes archivos: {', '.join(open_files)}")
        st.stop()

    df_write_forecast=pd.read_excel(path_write_forecast)
    check_mandatory_cols(df_write_forecast.columns,'Write Forecast')
    df_write_forecast['year_month']= pd.to_datetime(df_write_forecast['Due Date']).dt.strftime('%Y-%m')
    df_pos_bom_write=df_pos_bom_selected[['year_month','Part Number','Extended']]
    df_pos_bom_write.rename({'Part Number':'Part'},axis=1,inplace=True)
    # Componentes de PO que ya no estan en el write forecast
    df_not_in_forecast=df_pos_bom_write.merge(df_write_forecast,how='left',on=['year_month','Part'])
    df_not_in_forecast=df_not_in_forecast[df_not_in_forecast['Part Site'].isnull()][['year_month','Part','Extended']]
    df_write_forecast=df_write_forecast.merge(df_pos_bom_write,how='left',on=['year_month','Part'])

    if df_not_in_forecast.shape[0]>0:
        if os.path.exists(path_reportes):
            append_sheet(df_not_in_forecast,path=path_reportes,sheet_name='Componentes sin Forecast',index=False)
        else:
            save_df(df_not_in_forecast,filepath=path_reportes,sheet_name='No en WriteForecast',index=False)
        st.info(f"Advertencia: No se encontraron algunas POs en write forecast, revisar el archivo: {path_reportes}")

    df_write_forecast['Extended'].fillna(0,inplace=True)

    # Solo se hacen cambios en registros con FC (ForeCast)
    df_write_forecast.loc[~df_write_forecast['U_REFRNC'].str.contains('FC'),'Extended']=0
    df_write_forecast_changed=df_write_forecast[df_write_forecast['Extended']>0]

    df_write_forecast_changed['New Quantity']=df_write_forecast_changed['Quantity']-df_write_forecast_changed['Extended']
    df_write_forecast['Quantity']=df_write_forecast['Quantity']-df_write_forecast['Extended']

    df_write_forecast['Model']=df_write_forecast['Model'].fillna('None')
    df_write_forecast=df_write_forecast[df_write_forecast['Quantity']>0]
    df_write_forecast.drop(['Extended','year_month'],axis=1, inplace=True)

    # ### Agregar POs al write forecast
    # - Agrega las POs como demanda al write forecast
    # - Revisa que po/modelo no existan
    # - Solo agrega PO si tienen BOM



    # Agregar POs
    df_pos_write_add=df_pos.copy()
    df_pos_write_add.rename(axis=1,inplace=True,mapper={
        'ProductServiceID':'Part',
        'PODate':'Due Date',
        'PurchaseOrder':'U_REFRNC'
    })
    # Revisar que no se inserten pos que ya existan (Validar campos de PO y Part correctos)
    df_pos_write_add=df_pos_write_add[~(df_pos_write_add['U_REFRNC']+df_pos_write_add['Part']).isin(df_write_forecast['U_REFRNC']+df_write_forecast['Part'])]
    df_pos_write_add['Due Date']=pd.to_datetime(df_pos_write_add['Due Date'])
    df_pos_write_add['Promised Date']=df_pos_write_add['Due Date']
    df_pos_write_add['Quantity']=df_pos_write_add['Quantity'].astype(int)
    df_pos_write_add['Part Site']='CHI_KRS'
    df_pos_write_add['Customer']='ECO01E'
    df_pos_write_add['Order Id']=1
    df_pos_write_add['O/T']='FC'
    df_pos_write_add['LN/DEL']=1
    df_pos_write_add['Shipped Qty']=0
    df_pos_write_add['Unit Selling Price']=0
    df_pos_write_add['Del#']=''
    df_pos_write_add['Status']=3
    df_pos_write_add['Org. Duedate']=''
    df_pos_write_add['Model']='None'
    df_pos_write_add['Order Site']='CHI_KRS'


    df_write_forecast=pd.concat([df_pos_write_add,df_write_forecast])
    df_write_forecast_changed=pd.concat([df_pos_write_add,df_write_forecast_changed])
    df_write_forecast=df_write_forecast[mandatory_cols['Write Forecast']]
    #-----------------

    # ### Descontar POs embarcadas del write forecast

    #Shipments

    path_shipments=st.session_state.selected_paths.get('shipment_file')
    if path_shipments:
        df_shipments=pd.read_excel(path_shipments)

        df_shipments.dropna(subset=['Site'],inplace=True)
        check_mandatory_cols(df_shipments.columns,'Shipment')
        df_shipments=df_shipments[mandatory_cols['Shipment']]
        df_shipments.rename({'Customer PO#':'U_REFRNC','Part No.':'Part'},axis=1,inplace=True)
        #Sumarizar embarques
        df_shipments=df_shipments.groupby(['U_REFRNC','Part']).sum('Qty. Shipped')
        df_shipments.reset_index(inplace=True)
        df_write_forecast=df_write_forecast.merge(df_shipments,how='left',on=['U_REFRNC','Part'])
        df_write_forecast['Qty. Shipped'].fillna(0,inplace=True)
        #Agregar al reporte de cambios
        df=df_write_forecast[df_write_forecast['Qty. Shipped']>0]
        df['New Quantity']=df_write_forecast['Quantity']-df_write_forecast['Qty. Shipped']
        df_write_forecast_changed=pd.concat([df,df_write_forecast_changed])
        df_write_forecast['Quantity']=df_write_forecast['Quantity']-df_write_forecast['Qty. Shipped']
        df_write_forecast.drop(['Qty. Shipped'],axis=1,inplace=True)
    df_write_forecast=df_write_forecast[df_write_forecast['Quantity']>0]

    save_df(df_write_forecast,filepath=path_write_forecast_new,sheet_name='Write Forecast',index=False)
    save_df(df_write_forecast_changed,filepath=path_write_forecast_changes,sheet_name='Write Forecast Changes',index=False)

    #Format forecast
    if is_file_open(path_write_forecast_new):
        st.info(f"Favor de cerrar un archivo {path_write_forecast_new}")
        st.stop()  
    wb = load_workbook(path_write_forecast_new)
    ws = wb['Write Forecast']
    set_number_format(ws,col_name='Due Date',format='YYYYMMDD')
    set_number_format(ws,col_name='Promised Date',format='YYYYMMDD')
    save_wb(wb=wb,filepath=path_write_forecast_new)
    try:
        os.startfile(path_write_forecast_new)
        os.startfile(path_write_forecast_changes)
        os.startfile(path_reportes)
    except:
        st.info(f"El archivo Write Forecast esta listo.")


# =============================================================================
# Main Script: Streamlit App UI
# =============================================================================

path_pickle=os.path.join(Path(__file__).parent,'state_carga_demanda.pkl')
state = load_state_pickle(path_pickle)
st.session_state.folder_output = state['folder_output']
st.session_state.selected_paths = state['selections']
try:
    st.session_state.output_paths = set_paths(st.session_state.folder_output)
except:
    st.warning("Seleccione la carpeta de trabajo y reinicie")

try:
    st.set_page_config(page_title="Carga de demanda", page_icon=":scroll:")
except StreamlitAPIException:
    pass
st.markdown("<div style='position: absolute; top: 10px; left: 10px; font-size: 14px; color: gray;'>V8. 2025-06-18</div>", unsafe_allow_html=True)
st.markdown(
    r"""
    <style>
    .stAppDeployButton {
            visibility: hidden;
        }
    </style>
    """, unsafe_allow_html=True
)
st.title("Carga de demanda")


st.header("Seleccionar carpeta de trabajo")
if st.button("Seleccionar carpeta", key="select_folder"):
    folder = select_directory(initialdir=state["folder_output"] or os.getcwd())
    if folder:
        state["folder_output"] = folder
        save_state_pickle(state,filename=path_pickle)
        st.rerun()

if state["folder_output"]:
    st.success(f"Carpeta de trabajo: {state['folder_output']}")
else:
    st.info("No seleccionado.")


if st.button("Carpeta de archivos Korrus", key="select_korrus"):
    folder = select_directory(initialdir=state["folder_output"] or os.getcwd())
    if folder:
        state["korrus_folder"] = folder
        save_state_pickle(state,filename=path_pickle)
        st.rerun()

if "korrus_folder" in state:
    st.success(f"Carpeta de trabajo: {state['korrus_folder']}")
else:
    st.info("No seleccionado.")


st.header("Seleccion de archivos")
file_selectors = [
    ("korrus_file", "Korrus File"),
    ("bom_file", "BOM"),
    ("write_forecast", "Write Forecast"),
    ("apollo_file", "Apollo"),
    ("shipment_file", "Shipment")
]
for selector_key, display_label in file_selectors:
    manage_file_selector(selector_key, display_label, state)

selected_date = st.date_input("Fecha Korrus", value=state.get("korrus_date", datetime.date.today()))
if selected_date != state.get("korrus_date", datetime.date.today()):
    state["korrus_date"] = selected_date
    save_state_pickle(state,filename=path_pickle)
    st.rerun()

st.header("Generar demanda")

if st.button("Iniciar"):
    generate_demand()