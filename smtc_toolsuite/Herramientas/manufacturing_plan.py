"""
Manufacturing plan
- V15. 2025-08-05
    - Planeacion parte 2, se agrega manejo de diferentes routers y operaciones por maquina
- V14. 2025-06-07
    - Correccion para la primera vez que se ejecuta el script
    - Las ordenes con Status Planeada no se toman en cuenta para la programación
- V13. 2025-05-28
    - Se filtra SVT del 16 WK
- V12. 2025-05-25
    - Se corrige la rutina de asignación
    - Se agrega sugerencia de prioridades
- V11. 2025-05-12
    - Cambio mayor en la rutina de asignación, se realiza por slots para poder programar ordenes con diferentes doblados dependientes uno del otro
- V10. 2025-05-12
    - Se agrega fecha a nombre de reportes
    - Definición de precios tomando en cuenta sustitutos
    - El proceso continua aunque falte definir routers o maquinas
- V9. 2025-05-07
    - Se asignan números de parte solo a una maquina
- V8. 2025-05-05
    - Se corrige orden de los reportes, setup 
    - se simplifica el proceso de asignación
    - Se busca en routers y maquinas por numero de parte normal y sustituto
- V7. 2025-05-04
    - Programación en un periodo de tiempo, validación de plan, equivalencias
- V7. 2025-04-29
    - Se da formato a primer y segundo turno, se guarda el ancho de las columnas del primer archivo de reporte y se aplica a todos los reportes
- V6. 2025-04-27
    - Se agrega precio, horas por fecha, 
- V5. 2025-04-27
    - Mejoras, horas por maquina, orden de columnas depende del archivo de formatos
- V4. 2025-04-21
    - Corrección a extracción de routers, se agregan variables routers, turnos
- V3. 2025-04-21
    - Reportes en excel
- V2. 2025-04-19
    - Verificación de ordenes ya programadas, manejo de status, integración con plan existente
- V1. 2025-04-10
    - Version inicial, calculo de plan de producción
"""

import streamlit as st
import pickle
import os
from tkinter import Tk, filedialog as fd
import datetime
import pandas as pd
from copy import copy, deepcopy
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import win32com.client
from pandas.tseries.offsets import BDay
from openpyxl import load_workbook 
from pathlib import Path
from streamlit.errors import StreamlitAPIException
import uuid
try: 
    from Herramientas.excel_normalizer import ExcelNormalizer
except:
    pass
from difflib import SequenceMatcher

_THIS_PAGE = Path(__file__).stem        # e.g. "manufacturing_plan"

prev = st.session_state.get("_active_page")

if prev is None or prev != _THIS_PAGE:
    # First load OR coming from a different page → purge everything
    for k in list(st.session_state.keys()):
        del st.session_state[k]

# Record that we're now on this page
st.session_state["_active_page"] = _THIS_PAGE
# -------------------------------------------------------------------
# =============================================================================
# File/Directory & System Utilities
# =============================================================================

def open_file_selection(initialdir='', filter_name='*.*'):
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    filetypes = (
        (filter_name, f"*{filter_name.split(' ')[0]}*.*"),
        ('All files', '*.*'),
        ('Excel', '*.xlsx'),
        ('CSV', '*.doc')
    )
    files = fd.askopenfilenames(filetypes=filetypes, initialdir=initialdir)
    root.destroy()
    return files

def select_directory(initialdir):
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    directory = fd.askdirectory(initialdir=initialdir)
    root.destroy()
    return directory

def set_paths(path):
    output_paths = {}
    output_paths['path_xl_format'] = os.path.join(path, 'columns and formatting.xlsx')
    output_paths['path_assigned'] = os.path.join(path, 'status de ordenes.xlsx')
    output_paths['path_report'] = os.path.join(path, 'reporte de manufactura.xlsx')
    return output_paths

def set_col_rel(output_paths):
    df_columns = read_excel(output_paths['path_xl_format'], sheet_name='column_format')
    df_col_rel = df_columns[~df_columns['std_name'].isnull()].copy()
    column_equivalence = read_excel(
        st.session_state.output_paths["path_xl_format"], sheet_name="column_equivalence"
    )
    master_operation_relation = read_excel(
            st.session_state.output_paths["path_xl_format"], sheet_name="master_operation_relation"
        )    
    return {'col_rel': df_col_rel, 
            'columns': df_columns,
            'column_equivalence':column_equivalence,
            'master_operation_relation':master_operation_relation}

# =============================================================================
# Excel Management Functions
# =============================================================================

def read_excel(path=None, sheet_name=0, header=0, keep_default_na=True, dtype=None):
    with pd.ExcelFile(path) as xls:
        df = pd.read_excel(path, sheet_name=sheet_name, header=header, keep_default_na=keep_default_na, dtype=dtype)
    return df

def load_excel_with_header_key(file_path, sheet_name=0, key_text='', dtype=None, **kwargs):
    df = read_excel(file_path, sheet_name=sheet_name, keep_default_na=False, dtype=dtype)
    header_row = None
    if key_text in df.columns:
        header_row = 0
    else:
        for col in df.columns:
            for i, cell in df[col].items():
                if pd.notna(cell) and key_text in str(cell):
                    header_row = i
                    break
            if header_row is not None:
                break
        if header_row is None:
            raise ValueError(f"Key text '{key_text}' not found in the sheet.")
        df.columns = df.iloc[header_row]
        df = df.iloc[header_row + 1:].reset_index(drop=True)
    return df

def find_cell_by_text(ws, text):
    """
    Find the cell containing the specified date in the worksheet.
    
    :param ws: The worksheet object from openpyxl
    :param date_value: The date value to search for (datetime object or string)
    :return: The cell address (e.g., 'B2') or None if not found
    """
    # for row in ws.iter_rows():
    for row in ws[ws.calculate_dimension()]:
        for cell in row:
            if cell.value == text:
                return cell.coordinate  # Return cell address (e.g., 'B2')
    return None 

def get_column_info(ws, col_name, raise_error=True):
    """
    Returns a dictionary with the column cell and the data range for the column.
    """
    col_cell=find_cell_by_text(ws, col_name)
    if not col_cell:
        if not raise_error:
            return False
        st.error(f"Column '{col_name}' not found in the sheet.")
        st.stop()
    col_cell=ws[col_cell]
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    data_range=ws[col_cell.offset(1,0).coordinate:ws.cell(ws[last_cell].row,col_cell.offset(1,0).column).coordinate]
    data_range_list=[cell[0] for cell in data_range]
    return {'data_range':data_range_list,
            'col_cell':col_cell}

def get_col_sizes(wb):
    """
    Obtains a dictionary with the column sizes of each sheet in a workbook
    """
    col_sizes={}
    for sheet in wb.sheetnames:
        col_sizes[sheet]=wb[sheet].column_dimensions
    return col_sizes

def apply_col_sizes(wb,col_sizes):
    """
    Applies the column sizes obtained by function get_col_sizes
    """
    for sheet in col_sizes.keys():
        if sheet not in wb.sheetnames:
            continue
        ws=wb[sheet]
        for column_letter in col_sizes[sheet].keys():
            if column_letter==0:
                continue
            ws.column_dimensions[column_letter].width=col_sizes[sheet][column_letter].width
    return wb

def get_cell_properties(cell):
    properties = {}
    fill = cell.fill
    properties["background_color"] = fill
    properties["font"]=cell.font
    properties["alignment"] = cell.alignment
    return properties

def format_cell(cell,properties):
    "Apply properties based on the dict of prooperties"
    cell.fill=copy(properties['background_color'])
    cell.font = copy(properties["font"])
    cell.alignment = copy(properties["alignment"])

def get_xl_formatting(table_name=None):
    """
    Returns the formatting for the excel file defined in columns and formatting.xlsx
    return example
    {'oor':
    {
        'columna':{
            'head':{properties},
            'data':{properties}
        }
    }
    }
    """
    wb = load_workbook(st.session_state.output_paths['path_xl_format'])
    ws = wb['column_format']
    col_info=get_column_info(ws,'column_name')
    cell_properties={}
    for cell in col_info['data_range']:
        head_properties=get_cell_properties(cell)
        data_properties=get_cell_properties(cell.offset(0, 1))
        if cell.offset(0, -1).value not in cell_properties:
            cell_properties[cell.offset(0, -1).value]={} #Table name
        if cell.value not in cell_properties[cell.offset(0, -1).value]:
            cell_properties[cell.offset(0, -1).value][cell.value]= {} #Column name
        cell_properties[cell.offset(0, -1).value][cell.value]['head_properties']=head_properties
        cell_properties[cell.offset(0, -1).value][cell.value]['data_properties']=data_properties
    ws = wb['special_format']
    col_info=get_column_info(ws,'format_name')
    cell_properties['special_format']={}
    for cell in col_info['data_range']:
        cell_properties['special_format'][cell.value]=get_cell_properties(cell)

    if table_name:
        cell_properties=cell_properties[table_name]
    return cell_properties  
# =============================================================================
# Persistence / State Management
# =============================================================================

def save_state_pickle(state, filename='folder_state.pkl'):
    with open(filename, 'wb') as f:
        pickle.dump(state, f)

def load_state_pickle(filename='folder_state.pkl'):
    try:
        with open(filename, 'rb') as f:
            return pickle.load(f)
    except FileNotFoundError:
        return {
            "folder_output": None,
            "selections": {}
        }
def is_file_open(filepath):
    # Check if the file exists
    if not os.path.exists(filepath):
        return False  # File does not exist, so treat it as "open" for your logic

    try:
        # Try to open the file in write mode
        with open(filepath, 'a'):
            pass
        return False  # File is not open (no exception raised)
    except PermissionError:
        return True 
    
def close_xl_if_open(path):
    if is_file_open(path):
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            workbook = excel.Workbooks(path)
            workbook.Save()
            workbook.Close()
        except:
            st.error(f"Cerrar el archivo: {path}")
            st.stop()   


# =============================================================================
# DataFrame Management & Data Processing
# =============================================================================

def rename_columns(df, df_col_rel, table_from='std_name', sheet_from='only', table_to='std_name', sheet_to='only'):
    required_cols = {'table', 'sheet', 'column_name', 'std_name'}
    missing_cols = required_cols - set(df_col_rel.columns)
    if missing_cols:
        raise ValueError(f"df_col_rel is missing required columns: {missing_cols}")
    mapping = {}
    if table_to == 'std_name':
        filtered = df_col_rel[(df_col_rel['table'] == table_from) & (df_col_rel['sheet'] == sheet_from)]
        if filtered.empty:
            raise ValueError("No matching mapping found for given table_from and sheet_from.")
        mapping = filtered.set_index('column_name')['std_name'].to_dict()
    elif table_from == 'std_name':
        filtered = df_col_rel[(df_col_rel['table'] == table_to) & (df_col_rel['sheet'] == sheet_to)]
        if filtered.empty:
            raise ValueError("No matching mapping found for given table_to and sheet_to.")
        mapping = filtered.set_index('std_name')['column_name'].to_dict()
    else:
        filtered_from = df_col_rel[(df_col_rel['table'] == table_from) & (df_col_rel['sheet'] == sheet_from)]
        if filtered_from.empty:
            raise ValueError("No matching mapping found for given table_from and sheet_from.")
        mapping_from = filtered_from.set_index('column_name')['std_name'].to_dict()
        df = df.rename(columns=mapping_from)
        filtered_to = df_col_rel[(df_col_rel['table'] == table_to) & (df_col_rel['sheet'] == sheet_to)]
        if filtered_to.empty:
            raise ValueError("No matching mapping found for given table_to and sheet_to.")
        mapping = filtered_to.set_index('std_name')['column_name'].to_dict()
    return df.rename(columns=mapping)

def format_dates(df, date_cols=[],type='iso'):
    for col in date_cols:
        if not col in df.columns:
            continue
        if type=='iso':
            df[col]=pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
        else:
            df[col]=pd.to_datetime(df[col], errors='coerce')
    return df

def read_predefined_excel(path,df_columns=pd.DataFrame(),table='',sheet='only',check_mandatory=False):
    """
    If the file exists, it reads it and returns the dataframe, 
    otherwise it returns an empty dataframe but with the columns of the file
    it checks mandatory columns according to columns and formatting file but rises it only if indicated
    """
    if not os.path.exists(path):
        df=get_predefined_df(df_columns=df_columns,
                             table=table,
                             sheet=sheet)
    else:
        df=read_excel(path)    
        if check_mandatory:
            check_mandatory_columns_df(df.columns,df_columns=df_columns,table=table,sheet=sheet)
    return df

def get_predefined_df(df_columns=pd.DataFrame(),table='',sheet='only'):
    cols=df_columns[(df_columns['table']==table)&
                                        (df_columns['sheet']==sheet)]['column_name'].to_list()
    if len(cols)==0:
        st.error(f"Tabla: {table} no definida en archivo columns and formatting")
        st.stop()
    df=pd.DataFrame(columns=cols)
    return df

def check_mandatory_columns_df(cols=[],df_columns=pd.DataFrame(),table='',sheet='only'):
    """
    Check mandatory columns against the columns marked in
    columns and formatting file
    """
    mandatory_cols=df_columns[(df_columns['table']==table)&
                                        (df_columns['sheet']==sheet)&
                                        (~df_columns['mandatory_column'].isna())]['column_name'].to_list()
    missing_columns = [col for col in mandatory_cols if col not in cols]
    if len(missing_columns)>0:
        st.error(f"No se encontraron las siguientes columnas en el archivo {table}: {missing_columns}")
        st.stop()
def append_df_to_df(df_new=pd.DataFrame(),df_old=pd.DataFrame(),table='',keys=[],date_cols=[],allow_duplicates=False):
    """
    Appends a dataframe to an existing Dataframe
    keys: Fields that should not be duplicated, the old records are kept
    table: Table which is source of the dataframe, just to show it in the error
    date_cols: Columns transformed to date
    """
    if df_new.empty:
        return
    if not keys:
        keys=df_new.columns
    df_new=df_new.copy()
    df_new=format_dates(df_new,date_cols)
    df_old=format_dates(df_old,date_cols)
    df_new=get_common_records(df_new,df_old,keys=keys,how='uncommon')
    df_old=pd.concat([df_old,df_new])  
    df_old.reset_index(inplace=True,drop=True)
    df_grp=df_old.groupby(keys).count()
    if allow_duplicates==True:
        return df_old
    df_grp=df_grp[df_grp[df_grp.columns[0]]>1]
    if len(df_grp)>0:
        df_grp.reset_index(inplace=True)
        st.error(f"Hay duplicados en el archivo {table} para la llave {keys}:")
        st.error(df_grp[keys].sort_values(keys))
        st.stop()
    return df_old

def get_common_records(df_new=pd.DataFrame(),df_old=pd.DataFrame(),keys=[],how='common'):
    """
    Gets common records in two dataframes based on a key of columns
    when getting uncommon only records the records of the new are returned
    """
    df_old=df_old.copy()
    df_new=df_new.copy()
    df_old['composite_key'] = list(zip(*(df_old[col] for col in keys)))
    df_new['composite_key'] = list(zip(*(df_new[col] for col in keys)))
    if how=='common':
        df_new=df_new[df_new['composite_key'].isin(df_old['composite_key'])]
    else:
        df_new=df_new[~df_new['composite_key'].isin(df_old['composite_key'])]
    df_new.drop(columns=['composite_key'],inplace=True)
    return df_new


    
# =============================================================================
# Main Functions
# =============================================================================

def orders_priority():
    path_order_list_proposed=os.path.join(state['folder_output'],'Lista de Ordenes sugerida.xlsx')
    path_sales = st.session_state.selected_paths['sales']
    path_end_of_period = st.session_state.selected_paths['end_of_period']
    path_16_wk = st.session_state.selected_paths['16_wk']
    msg_orders_priority=st.empty()
    if not path_16_wk:
        msg_orders_priority.error("Seleccionar 16 wk")
        st.stop()
    if not path_sales:
        msg_orders_priority.error("Seleccionar Top ventas")
        st.stop()
    if not path_end_of_period:
        msg_orders_priority.error("Seleccionar End of period")
        st.stop()        

    close_xl_if_open(path_order_list_proposed)
    df_columns=st.session_state.df_columns
    df_col_rel=st.session_state.df_col_rel
    df_16_wk=read_excel(path=path_16_wk)
    df_16_wk=rename_columns(df_16_wk,df_col_rel=df_col_rel,table_from='16 WK')
    df_16_wk['pn']=df_16_wk['pn'].astype(str)
    df_16_wk=df_16_wk[~df_16_wk['site'].str.contains('SVT')]
    df_16_wk=df_16_wk[df_16_wk['wk5']<0]
    df_16_wk.sort_values(['wk5'],inplace=True)
    # Sales
    df_sales=read_excel(path_sales)
    df_sales=rename_columns(df_sales,df_col_rel,table_from="Top ventas")
    df_sales['pn']=df_sales['pn'].astype(str)
    df_sales['tot_value'] = pd.to_numeric(df_sales['tot_value'], errors='coerce')
    df_sales=df_sales.dropna(subset=['tot_value'])
    df_sales = (
        df_sales
        .dropna(subset=['pn'])
        .sort_values('tot_value', ascending=False)
    )
    df_sales['cumperc'] = df_sales['tot_value'].cumsum() / df_sales['tot_value'].sum()
    cut_idx = df_sales['cumperc'].gt(0.8).idxmax()
    df_sales = df_sales.loc[:cut_idx]

    # End of period
    df_end_of_period=load_excel_with_header_key(file_path=path_end_of_period,key_text='Report Date')
    df_end_of_period=rename_columns(df_end_of_period,df_col_rel=df_col_rel,table_from='End of Period')
    df_end_of_period=df_end_of_period[~df_end_of_period['site_code'].str.contains('SVT')]
    df_end_of_period['pn']=df_end_of_period['pn'].astype(str)
    df_end_of_period=df_end_of_period.merge(df_16_wk[['pn','wk5']],how='left',on='pn')
    df_end_of_period=df_end_of_period.merge(df_sales[['pn','tot_value']],how='left',on='pn')
    df_order_list_proposed=df_end_of_period[['pn','wo','pzas_x_hacer','wk5','tot_value','create_wo']].sort_values(by=['wk5','tot_value','create_wo','pn'],ascending=[True,True,False,True])
    df_order_list_proposed.reset_index(inplace=True,drop=True)
    df_order_list_proposed.reset_index(inplace=True,names='priority')
    df_order_list_proposed['machine']=''
    df_order_list_proposed['operation_description']=''
    df_order_list_proposed['status']=''
    df_order_list_proposed = df_order_list_proposed.loc[:, ~df_order_list_proposed.columns.duplicated(keep='last')]
    df_order_list_proposed=rename_columns(df_order_list_proposed,df_col_rel=df_col_rel,table_from='End of Period',table_to='Lista de ordenes')
    df_order_list_proposed=df_order_list_proposed[df_columns.loc[df_columns['table']=='Lista de ordenes','column_name'].to_list()]
    df_order_list_proposed.to_excel(path_order_list_proposed,index=False)    
    os.startfile(path_order_list_proposed)

# @note Seleccion de maquinas
def machine_selection():
    if 'editor_id' not in st.session_state:
        st.session_state.editor_id=''
    st.session_state.editor_id=str(uuid.uuid4())
    df_order_list=st.session_state.df_order_list
    valid_part_numbers=df_order_list['pn'].tolist()
    if 'df_plan_old' not in st.session_state:
        st.error('Favor de verificar las ordenes')
        return
    normalizer = ExcelNormalizer(st.session_state.df_column_equivalence)
    df_master_operation_relation=st.session_state.df_master_operation_relation
    df_selected_operations=st.session_state.df_selected_operations
    df_routing=st.session_state.df_routing
    df_master = normalizer.normalize_folder(state["folder_master"])
    df_master['operation_description']=df_master['operation_description'].str.upper().fillna('')
    df_master = df_master[df_master["pn"].isin(valid_part_numbers)]
    blanks_idx=(df_master['operation_description']=='')
    df_master_1=df_master[blanks_idx]
    df_master_2=df_master[~blanks_idx]
    df_master_1=df_master_1.drop(columns=['operation_description']).merge(df_master_operation_relation,how='left',on=['file_name_like'],)
    df_master_1['operation_description']=df_master_1['operation_description'].fillna('')
    df_master=pd.concat([df_master_1,df_master_2])
    df_master.reset_index(drop=True,inplace=True)
    df_master=df_master[df_master["operation_description"].isin(df_selected_operations['operation_description'])]
    df_master.fillna("", inplace=True)
    df_master = df_master[df_master["pn"] != ""]
    df_master['composite_key']=list(zip(*(df_master[col] for col in ['pn','operation_description'])))
    df_routing['composite_key']=list(zip(*(df_routing[col] for col in ['pn','operation_description'])))
    df_master=df_master[df_master['composite_key'].isin(df_routing['composite_key'])]
    df_master.drop(columns=['composite_key'],inplace=True)
    machine_cols = [c for c in df_master.columns if c.startswith("maq_")]
    df_master_long = df_master.melt(
    id_vars=['pn','file_name_like','operation_description'],
    value_vars=machine_cols,
    var_name='maq_col',
    value_name='maq'
            )
    df_master_long=df_master_long[df_master_long['maq']!='']
    df_master_long.drop(columns=['file_name_like'],inplace=True)
    df_master_long['Disponible']=True
    st.session_state.df_master_long = df_master_long.copy()
    path_available_hours=st.session_state.selected_paths['available_hours']
    df_avail_hours=read_excel(path=path_available_hours)
    df_avail_hours['dia']=pd.to_datetime(df_avail_hours['dia'],errors='coerce').dt.strftime('%Y-%m-%d')
    df_avail_hours['dia'].fillna('default',inplace=True)   

    st.session_state.df_avail_hours=df_avail_hours

# @note Verify order list
def verify_order_list():
    df_columns=st.session_state.df_columns
    df_col_rel=st.session_state.df_col_rel
    #% Open order list
    path_order_list = st.session_state.selected_paths['order_file']
    df_order_list = load_excel_with_header_key(path_order_list, key_text='Priority')
    check_mandatory_columns_df(df_order_list.columns,df_columns=df_columns,table='Lista de ordenes')
    df_order_list = rename_columns(df_order_list, df_col_rel, table_from='Lista de ordenes')
    df_order_list = df_order_list[df_order_list['status']!='Planeada']
    st.session_state.df_order_list=df_order_list

    # Get routing list
    path_routing = st.session_state.selected_paths['routing_file']
    df_routing = load_excel_with_header_key(path_routing, sheet_name='Operations', key_text='Routing')
    df_routing = rename_columns(df_routing, st.session_state.df_col_rel, table_from='Routing', sheet_from='Operations')
    df_routing['operation_description']=df_routing['operation_description'].str.upper()
    df_routing['pn']=df_routing['pn'].astype(str)
    st.session_state.df_routing = df_routing
    df_available_operations=df_routing[df_routing['pn'].isin(df_order_list['pn'])][['operation_description']].drop_duplicates()
    df_available_operations['Programar']=False
    st.session_state.df_available_operations=df_available_operations

    #% Open old plan
    path_plan=os.path.join(st.session_state.folder_output,f"{st.session_state.plan_name}.xlsx")
    df_plan_old=read_predefined_excel(path_plan,df_columns=st.session_state.df_columns,table='Manufacturing plan',check_mandatory=True)
    st.session_state.df_plan_old=df_plan_old
    if len(df_plan_old)==0:
        st.info("Ok")
        return
    df_plan_old=df_plan_old[~df_plan_old['status'].isna()]
    st.session_state.df_plan_old=df_plan_old
    #% Show orders already planned
    df_already_planned=get_common_records(df_new=df_order_list,df_old=df_plan_old,keys=['wo','pn'])
    if len(df_already_planned)==0:
        st.info("Ok")
    else:
        st.info("Las siguientes ordenes ya estan planeadas, continue si desea agregarlas al nuevo plan con cantidad diferente")
        st.dataframe(df_already_planned)


# @note Create plan
def create_plan():
    if not os.path.exists(st.session_state.output_paths['path_xl_format']):
        st.error("No se encuentra el archivo: columns and formatting.xlsx")
        st.stop()
    if 'df_plan_old' not in st.session_state:
        st.error('Favor de verificar las ordenes')
        return
    if len(st.session_state.df_missing_avail)>0:
        st.error("Se requiere disponibilidad en el archivo Horas disponibles para las maquinas:")
        st.dataframe(df_missing_avail)
        return
    if ('plan_name' not in st.session_state) | (st.session_state.plan_name==''):
        st.error("Definir nombre para el plan")
        return
    df_avail_hours=st.session_state.df_avail_hours
    dict_avail_hours=df_avail_hours.set_index(['dia','maquina']).to_dict(orient='index')

    df_col_rel=st.session_state.df_col_rel
    path_plan=os.path.join(st.session_state.folder_output,f"{st.session_state.plan_name}.xlsx")
    df_columns=st.session_state.df_columns
    close_xl_if_open(path_plan)
    path_master_doblado = st.session_state.selected_paths['master_file']
    df_master_doblado = load_excel_with_header_key(path_master_doblado, sheet_name='00. Formato para Master de WC', key_text='PN')
    check_mandatory_columns_df(df_master_doblado.columns,df_columns=df_columns,table='Master Doblado',sheet='00. Formato para Master de WC')
    df_master_doblado = rename_columns(df_master_doblado, st.session_state.df_col_rel, table_from='Master Doblado', sheet_from='00. Formato para Master de WC')
    df_master_doblado.replace('/','_',regex=True,inplace=True)
    df_master_doblado['pn']=df_master_doblado['pn'].astype(str)

    df_routing=st.session_state.df_routing 

    # Equivalencias
    path_equiv = st.session_state.selected_paths['equivalencias_file']
    df_equiv=read_predefined_excel(path_equiv,df_columns,table='Equivalencias')
    df_equiv['pn']=df_equiv['pn'].astype(str)
    dict_equiv=df_equiv.set_index('pn')['equivalencia'].to_dict()
    dict_equiv_inv={v: k for k, v in dict_equiv.items()}
    # Lista de ordenes
    path_order_list = st.session_state.selected_paths['order_file']
    df_order_list = load_excel_with_header_key(path_order_list, key_text='Priority')
    df_order_list = rename_columns(df_order_list, st.session_state.df_col_rel, table_from='Lista de ordenes')
    df_order_list['pn']=df_order_list['pn'].astype(str)
    # df_order_list['operation_description'].fillna('DOBLADO',inplace=True)
    # df_order_list.loc[df_order_list['operation_description']=='','operation_description']='DOBLADO'
    st.session_state.df_order_list=df_order_list
    df_order_list['pn_orig']=df_order_list['pn']
    df_order_list['pn']=df_order_list['pn'].replace(dict_equiv)
    # df_order_list['operation_description']=df_order_list['operation_description'].str.upper()
    # Missing routings or machine definition alert
    #df_missing_rout=get_common_records(df_order_list,df_routing,keys=['pn','operation_description'],how='uncommon')
    # df_order_list['composite_key'] = list(zip(*(df_order_list[col] for col in ['pn','operation_description'])))
    # df_routing['composite_key'] = list(zip(*(df_routing[col] for col in ['pn','operation_description'])))
    # df_missing_rout=df_order_list[~df_order_list['composite_key'].isin(df_routing['composite_key'])]
    # df_missing_rout['composite_key'] = list(zip(*(df_missing_rout[col] for col in ['pn_orig','operation_description'])))
    # df_missing_rout=df_missing_rout[~df_missing_rout['composite_key'].isin(df_routing['composite_key'])]
    # if 'composite_key' in df_missing_rout.columns:
    #     df_missing_rout.drop(columns=['composite_key'],inplace=True)
    # df_missing_machine=df_order_list[(~df_order_list['pn'].isin(df_master_doblado['pn']))&
    #                                 (~df_order_list['pn_orig'].isin(df_master_doblado['pn']))&
    #                                 (df_order_list['machine']=='')]
    # df_order_list.drop(columns=['composite_key'],inplace=True)
    
    # if len(df_missing_rout)>0:
    #     st.error("Falta definir router para las lineas:")
    #     st.dataframe(df_missing_rout)
    # if len(df_missing_machine)>0:
    #     st.error("Falta definir maquina para la lineas:")
    #     st.dataframe(df_missing_machine)

    # part_numbers=st.session_state.df_master_long
    part_numbers = {}
    for pn, grp in st.session_state.df_selected_master.groupby("pn"):
        ops = {}
        for idx, (op_name, sub) in enumerate(grp.groupby("operation_description"), start=1):
            # Tomamos la lista de máquinas de la columna 'maq', quitamos NaN/strings vacíos
            machines_raw = sub["maq"].dropna().astype(str).tolist()
            machines_raw = [m for m in machines_raw if m]
            # Preservamos orden y unicidad
            unique_machines = list(dict.fromkeys(machines_raw))
            # Construimos la lista [{m: "1"} …]
            machines = [{m: "1"} for m in unique_machines]
            ops[str(idx)] = {
                "name": op_name,
                "machines": machines
            }
        part_numbers[pn] = {"operations": ops}
    df_missing=df_order_list[~df_order_list['pn'].isin(list(part_numbers.keys()))]
    if len(df_missing)>0:
        st.error("Falta definir router o maquina para las lineas:")
        st.dataframe(df_missing[['priority','pn','wo','pzas_x_hacer']])

    #% Open part master
    path_part_master = st.session_state.selected_paths['part_master']
    df_part_master = load_excel_with_header_key(path_part_master, key_text='Site')
    check_mandatory_columns_df(df_part_master.columns,df_columns=df_columns,table='Part Master')
    df_part_master = rename_columns(df_part_master, df_col_rel, table_from='Part Master')
    df_part_master=df_part_master[['pn','unit_selling_price']].drop_duplicates(subset=['pn'],keep='last')
    df_part_master_orig=df_part_master.copy()
    df_part_master['pn']=df_part_master['pn'].replace(dict_equiv_inv)
    df_part_master=pd.concat([df_part_master_orig,df_part_master])
    df_part_master=df_part_master.drop_duplicates(subset=['pn'],keep='last')
    st.session_state.df_part_master=df_part_master
    df_plan_old=read_predefined_excel(path_plan,df_columns=df_columns,table='Manufacturing plan',check_mandatory=True)
    df_plan_old=rename_columns(df_plan_old,df_col_rel=df_col_rel,table_from='Manufacturing plan')


    # helper to iterate business‐day calendar
    def bdate_range_str(start, end):
        return [d.strftime('%Y-%m-%d') 
                for d in pd.bdate_range(start=start, end=end)]
    initial_date_str = pd.to_datetime(state['initial_date']).strftime('%Y-%m-%d')
    limit_date_ts    = pd.to_datetime(state['limit_date'])
    all_dates = bdate_range_str(initial_date_str, limit_date_ts.strftime('%Y-%m-%d'))
    machines = set(m for _, m in dict_avail_hours.keys())
    slots = {}
    for avail_dt in all_dates:
        for m in machines:
            dict_avail_hours_copy=deepcopy(dict_avail_hours)
            slots[(avail_dt, m)]=dict_avail_hours_copy.get((avail_dt, m),dict_avail_hours_copy.get(('default',m)))

    # 2) Prepare
    assignments    = []
    pn_to_machine  = {}    # first‐machine lock per PN
    wo_next_start  = {}    # earliest allowed timestamp per WO
    wo_last_machine = {}    # last machine used per WO
    machine_status = {}
    df_order_list.sort_values('priority', inplace=True)
    
    # prebuild full set of dates we may need
    # 3) Process each order
    for _, order in df_order_list.iterrows():
        pn_orig       = order['pn_orig']
        pn            = order['pn']
        # qty           = order['pzas_x_hacer']
        wo            = order['wo']
        # routing_name  = order['operation_description']
        # machine_def   = order['machine']
        pty           = order['priority']
        if pn not in part_numbers:
            continue
        for _,operation in part_numbers[pn]['operations'].items():
            qty = order['pzas_x_hacer']
            routing_name=operation['name']
            # routing info
            pn_info = df_routing[
                (df_routing['pn']==pn) & 
                (df_routing['operation_description']==routing_name)
            ]
            if pn_info.empty:
                pn_info = df_routing[
                    (df_routing['pn']==pn_orig) & 
                    (df_routing['operation_description']==routing_name)
                ]
                if pn_info.empty:
                    continue
            run_time, setup_time = pn_info.iloc[0][['run_time','setup_time']]
            # select machine
            # if machine_def:
            #     m_list = [machine_def]
            # elif pn in pn_to_machine:
            if pn in pn_to_machine:
                m_list = [pn_to_machine[pn]]
            else:
                # rows = df_master_doblado[df_master_doblado['pn']==pn]
                # if rows.empty:
                #     rows = df_master_doblado[df_master_doblado['pn']==pn_orig]
                #     if rows.empty:
                #         continue
                # m_list = [v for k,v in rows.iloc[0].items() if 'maq_opc' in k and v]
                m_list = [
                            m_name
                            for op in part_numbers[pn]['operations'].values()
                            for m_dict in op['machines']
                            for m_name, valid in m_dict.items()
                            if valid == '1'
                        ]                

            # determine earliest start
            initial_start={"finish_ts":pd.to_datetime(initial_date_str)}
            start_ts = wo_next_start.get(wo, initial_start)
            start_ts = start_ts.get("finish_ts")
            start_shift = wo_next_start.get(wo, initial_start)
            start_shift = start_shift.get("next_shift","PRIMER TURNO")
            # start_ts = start_ts.get("finish_ts")
            last_m   = wo_last_machine.get(wo)

            # assignment loop
            for m in m_list:
                # lock PN → machine on first assignment
                if pn not in pn_to_machine:
                    pn_to_machine[pn] = m
                # walk through dates until qty is 0 or we pass limit
                for date_str in all_dates:
                    if qty <= 0:
                        break
                    date_ts = pd.to_datetime(date_str)
                    # if date_ts < start_ts.normalize():
                    #     continue
                    if last_m is None or m != last_m:
                        if date_ts < start_ts.normalize():
                            continue                
                    if date_ts > limit_date_ts:
                        break

                    # get or init that day’s shifts
                    key = (date_str, m)
                    if key in slots:
                        avail_shifts = slots[key]
                    # iterate shifts in sorted order
                    for shift in sorted(avail_shifts):
                        if qty <= 0:
                            break

                        # compute setup only if PN changed vs last service on this machine
                        last_pn = machine_status.get((m,'last_pn'), None)
                        last_oper = machine_status.get((m,'last_oper'), None)
                        need_setup = setup_time if (last_pn != pn or last_oper != routing_name) else 0
                        # can we fit setup+one run?
                        if avail_shifts[shift] < need_setup + run_time:
                            avail_shifts[shift] = 0
                            continue
                        shifts_list=list(avail_shifts.keys())
                        if (date_ts == start_ts.normalize())&\
                            (shifts_list.index(start_shift)>shifts_list.index(shift))&\
                            (last_m is None or m != last_m):
                            continue
                        # assign as many as fit this slot
                        effective = avail_shifts[shift] - need_setup
                        pieces   = qty if run_time==0 else min(qty, effective // run_time)
                        if pieces<10:
                            continue
                        run_used = pieces * run_time
                        total    = run_used + need_setup

                        # record it
                        assignments.append({
                            'date': date_str,
                            'machine': m,
                            'shift': shift,
                            'operation_description': routing_name,
                            'pn': pn_orig,
                            'wo': wo,
                            'priority': pty,
                            'pzas_x_hacer': pieces,
                            'time_used': run_used,
                            'setup_time': need_setup
                        })

                        # update slot availability
                        avail_shifts[shift] -= total
                        qty -= pieces
                        # track last_pn for setup logic
                        machine_status[(m,'last_pn')] = pn
                        machine_status[(m,'last_oper')] = routing_name
                        # compute finish timestamp to enforce sequencing
                        finish_ts = date_ts  # + shift end offset if you track that
                        # bump next‐start only when changing machines
                        if last_m is None or m != last_m:
                            shifts_list=list(avail_shifts.keys())
                            shidx=shifts_list.index(shift)
                            if len(shifts_list)>shidx+1:
                                next_shift=shifts_list[shidx+1]
                            else:
                                finish_ts=finish_ts + BDay(1)
                                next_shift=shifts_list[0]                
                            next_start={"finish_ts":finish_ts,
                                        "next_shift":next_shift}
                            wo_next_start[wo] = next_start
                        # record this machine as last used for the WO
                        wo_last_machine[wo] = m

                    # end for shift
                # end for date

                break  # once we’ve tried this machine, stop (PN locked)

    df_plan_new=get_predefined_df(df_columns=df_columns,table='Manufacturing plan')
    if len(assignments)==0:
        msg_create_plan.error("No se lograron asignar ordenes, revise errores previos")
        st.stop()
    df_plan_new = pd.concat([df_plan_new,pd.DataFrame(assignments)],ignore_index=True)
    df_plan_new.sort_values(['date', 'machine', 'shift', 'priority', 'wo'], inplace=True)
    df_plan_new['pzas_x_hora']=(df_plan_new['pzas_x_hacer']/df_plan_new['time_used']).astype(float).round(2)
    df_plan_new['time_used']=df_plan_new['time_used'].astype(float).round(2)
    df_plan_old=df_plan_old[~df_plan_old['status'].isnull()]
    df_plan_old=append_df_to_df(df_new=df_plan_new,df_old=df_plan_old,table='Manufacturing plan',keys=['wo','pn','pzas_x_hacer'],allow_duplicates=True)
    df_part_master=st.session_state.df_part_master
    if 'unit_selling_price' in df_plan_old.columns:
        df_plan_old.drop(columns=['unit_selling_price'],inplace=True)
    df_plan_old=df_plan_old.merge(df_part_master,on='pn',how='left') 
    df_plan_old['unit_selling_price']=df_plan_old['unit_selling_price']*df_plan_old['pzas_x_hacer']
    df_plan_old['unit_selling_price'] = df_plan_old['unit_selling_price'].apply(lambda x: '${:,.2f}'.format(x))
    plan_cols=df_columns[(df_columns['table']=='Manufacturing plan')&
            (~df_columns['mandatory_column'].isna())]['std_name'].to_list()
    df_plan_old=df_plan_old[plan_cols]
    df_plan_old['date']=pd.to_datetime(df_plan_old['date'],format='mixed').dt.strftime('%B, %#d, %Y')
    st.session_state.df_plan_old=df_plan_old
    df_plan_old=rename_columns(df_plan_old,df_col_rel=st.session_state.df_col_rel,table_to='Manufacturing plan')    
    path_plan=os.path.join(st.session_state.folder_output,f"{st.session_state.plan_name}.xlsx")
    df_plan_old.to_excel(path_plan,sheet_name='Manufacturing plan',index=False)
    os.startfile(path_plan)
    msg_create_plan.info("Plan creado")


def generate_reports():
    msg_generate_reports=st.info("Generando...")
    df_columns=st.session_state.df_columns
    df_col_rel=st.session_state.df_col_rel
    df_plan_old=st.session_state.df_plan_old.copy()
    folder_output=st.session_state.folder_output
    machines=df_plan_old['machine'].drop_duplicates().tolist()
    df_plan_old=rename_columns(df_plan_old,df_col_rel=df_col_rel,table_to='Machine Report')
    plan_report_cols=df_columns[(df_columns['table']=='Machine Report')&
        (~df_columns['mandatory_column'].isna())]['column_name'].to_list()
    numeric_cols=df_columns[(df_columns['table']=='Machine Report')&
        (df_columns['data_type']=='sub_total')]['column_name'].to_list()
    machine_col=df_columns[(df_columns['table']=='Machine Report')&
        (df_columns['std_name']=='machine')]['column_name'].to_list()[0]  
    date_col=df_columns[(df_columns['table']=='Machine Report')&
        (df_columns['std_name']=='date')]['column_name'].to_list()[0]  
    group_cols=[date_col]
    df_plan_old=df_plan_old[plan_report_cols]
      
    dict_formats=get_xl_formatting()
    special_formats=dict_formats['special_format']
    col_sizes=None
    for file in os.listdir(folder_output):
        if 'reporte de manufactura' in file:
            wb=load_workbook(os.path.join(folder_output,file))
            col_sizes=get_col_sizes(wb)
            break

    for machine in machines:
        df=df_plan_old[df_plan_old[machine_col]==machine].copy()
        if len(df)==0:
            continue
        df.loc[df['Operacion']=='DOBLADO','Operacion']=''
        wb = Workbook()
        ws = wb.active
        ws.title = 'Report'

        # Write headers
        titles = list(df.columns)
        for idx, title in enumerate(titles, start=1):
            ws.cell(row=1, column=idx, value=title)

        current_row = 2
        black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
        df.sort_values(group_cols,inplace=True)
        
        groups = df.groupby(group_cols)
        for date, group in sorted(groups, key=lambda x: pd.to_datetime(x[0])):
            start_row = current_row
            # Data rows
            for idx, row in group.sort_index().iterrows():
                for col_idx, col in enumerate(titles, start=1):
                    if col not in group_cols:
                        cell=ws.cell(row=current_row, column=col_idx, value=row[col])
                        cell.value=row[col]
                        if row[col] in special_formats.keys():
                            format=special_formats[row[col]]
                            format_cell(cell,format)
                current_row += 1
            end_data_row = current_row - 1

            # Subtotal row
            for col_idx, col in enumerate(titles, start=1):
                if col in numeric_cols:
                    col_letter = get_column_letter(col_idx)
                    formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_data_row})"
                    ws.cell(row=current_row, column=col_idx, value=formula)
            current_row += 1
        
            # Blank line with black fill across data range
            for col_idx in range(1, len(titles) + 1):
                ws.cell(row=current_row, column=col_idx).fill = black_fill
            current_row += 1

            # Merge date cells and rotate text
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_data_row+1, end_column=1)
            date_cell = ws.cell(row=start_row, column=1, value=date[0])
            date_cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
            date_cell.number_format = 'mmmm, d, yyyy'
        if col_sizes:
            wb=apply_col_sizes(wb,col_sizes)            
        base, ext = os.path.splitext(st.session_state.output_paths['path_report'])
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        wb.save(f"{base} {machine} {timestamp}{ext}")
    msg_generate_reports.success("Reportes listos")
 
# @note Validate plan
def validate_plan():
    #% Validar plan
    if ('plan_name' not in st.session_state) | (st.session_state.plan_name==''):
        st.error("Definir nombre para el plan")
        return
    path_plan=os.path.join(st.session_state.folder_output,f"{st.session_state.plan_name}.xlsx")
    path_order_list = st.session_state.selected_paths['order_file']
    df_order_list=st.session_state.df_order_list
    close_xl_if_open(path_order_list)
    close_xl_if_open(path_plan)
    df_plan=read_excel(path_plan)
    df_plan_grp=df_plan.groupby(['pn','wo']).sum(['pzas_x_hacer'])[['pzas_x_hacer']]
    df_plan_grp.reset_index(inplace=True)
    df_order_list['composite_key']=list(zip(*(df_order_list[col] for col in ['pn','wo'])))
    df_plan['composite_key']=list(zip(*(df_plan[col] for col in ['pn','wo'])))
    df_not_planned=df_order_list[~df_order_list['composite_key'].isin(df_plan['composite_key'])]
    if len(df_not_planned)>0:
        df_not_planned.drop(columns=['composite_key'],inplace=True)
        st.warning("Las siguientes ordenes no han sido planeadas")
        st.dataframe(df_not_planned)
    # df_order_assignment=df_order_list.merge(df_plan_grp,how='left',on=['pn','wo'],suffixes=('','_assigned'))
    # df_order_assignment['pzas_x_asignar']=df_order_assignment['pzas_x_hacer']-df_order_assignment['pzas_x_hacer_assigned']
    # df_order_assignment.loc[df_order_assignment['pzas_x_asignar']==0,'status']='Planeada'
    # df_order_assignment.loc[df_order_assignment['pzas_x_asignar']>0,'status']='Incompleta'
    # df_order_assignment.loc[df_order_assignment['pzas_x_asignar'].isna(),'status']='Pendiente'
    # df_order_assignment=df_order_assignment.drop(columns=['pzas_x_hacer_assigned','pzas_x_asignar'])
    # orders_cols=df_columns[(df_columns['table']=='Lista de ordenes')]['std_name'].to_list()
    # df_order_assignment=df_order_assignment[orders_cols]
    # df_order_assignment=rename_columns(df_order_assignment,df_col_rel,table_to='Lista de ordenes')
    # df_order_assignment.to_excel(path_order_list,index=False)
    df_plan.loc[df_plan['status'].isna(),'status']='Planeada'
    df_plan.to_excel(path_plan,sheet_name='Manufacturing plan',index=False)
    os.startfile(path_order_list)
    os.startfile(path_plan)

def manage_file_selector(selector_key, display_label, state):
    if not state["selections"].get(selector_key):
        if st.button(f"{display_label} File", key=f"select_{selector_key}"):
            files = open_file_selection(initialdir=state["folder_output"] or os.getcwd())
            if files:
                state["selections"][selector_key] = files[0]
                save_state_pickle(state,filename=path_pickle)
                st.rerun()
        st.info(f"{display_label} no seleccionado.")
    else:
        st.success(f"Selected {display_label} File: {state['selections'][selector_key]}")
        if st.button(f"Change {display_label} File", key=f"change_{selector_key}"):
            state["selections"][selector_key] = ""
            save_state_pickle(state,filename=path_pickle)
            st.rerun()

# =============================================================================
# Main Script: Streamlit App UI
# =============================================================================
path_pickle=os.path.join(Path(__file__).parent,'folder_state_planner.pkl')
state = load_state_pickle(path_pickle)
st.session_state.folder_output = state['folder_output']
st.session_state.selected_paths = state['selections']
try:
    st.session_state.output_paths = set_paths(st.session_state.folder_output)
    st.session_state.col_rel = set_col_rel(st.session_state.output_paths)
    st.session_state.df_col_rel = st.session_state.col_rel['col_rel']
    st.session_state.df_columns = st.session_state.col_rel['columns']
    st.session_state.dict_formats=get_xl_formatting()
    st.session_state.df_column_equivalence = st.session_state.col_rel['column_equivalence']
    st.session_state.df_master_operation_relation = st.session_state.col_rel['master_operation_relation']
except Exception as e:
    st.warning("Seleccione la carpeta de trabajo y reinicie")
    st.error(f"Error: {e}")
try:
    st.set_page_config(page_title="Plan de manufactura", page_icon=":factory:")
except StreamlitAPIException:
    pass
st.markdown("<div style='position: absolute; top: 10px; left: 10px; font-size: 14px; color: gray;'>V15. 2025-08-05</div>", unsafe_allow_html=True)
st.markdown(
    r"""
    <style>
    .stAppDeployButton {
            visibility: hidden;
        }
    </style>
    """, unsafe_allow_html=True
)
st.title("Plan de manufactura")

custom_css = """
<style>
.stTabs [data-baseweb="tab-list"] button [data-testid="stMarkdownContainer"] p {
    font-size: 1.5rem; /* Adjust this value as needed */
}
</style>
"""
st.markdown(custom_css, unsafe_allow_html=True)

tab1, tab2 = st.tabs(["Seleccion de Archivos", "Planeacion"])
with tab1:
    st.header("Seleccionar carpeta de trabajo")
    if st.button("Seleccionar carpeta", key="select_folder"):
        folder = select_directory(initialdir=state["folder_output"] or os.getcwd())
        if folder:
            state["folder_output"] = folder
            save_state_pickle(state,filename=path_pickle)
            st.rerun()
    if state["folder_output"]:
        st.success(f"Carpeta de trabajo: {state['folder_output']}")
    else:
        st.info("No seleccionado.")


    st.header("Masters folder")
    if "folder_master" not in state:
        state["folder_master"]=""
    if st.button("Seleccionar carpeta con reportes Master", key="select_folder_master"):
        folder = select_directory(initialdir=state["folder_master"] or os.getcwd())
        if folder:
            state["folder_master"] = folder
            save_state_pickle(state,filename=path_pickle)
            st.rerun()

    if state["folder_master"]:
        st.success(f"Carpeta de trabajo: {state['folder_master']}")
    else:
        st.info("No seleccionado.")



    st.header("Seleccion de archivos")
    file_selectors = [
        ("16_wk", "16Wk Gap"),
        ("sales", "Top ventas"),
        ("end_of_period", "End Of Period"),
        ("master_file", "Master Doblado"),
        ("equivalencias_file", "Equivalencias"),
        ("order_file", "Lista de Ordenes"),
        ("routing_file", "Routing"),
        ("available_hours", "Horas disponibles"),
        ("part_master", "Part Master")
    ]
    for selector_key, display_label in file_selectors:
        manage_file_selector(selector_key, display_label, state)

with tab2:

    st.header("Nombre del plan:")
    st.session_state.plan_name = st.text_input("",value="",key='txt_plan_name')

    st.header("Definicion de ordenes")

    if st.button("Sugerir prioridades"):
        orders_priority()

    if st.button("Verificar ordenes"):
        verify_order_list()
        
    st.header("Operaciones a planear")
    msg_error_avail=st.empty()
    if 'df_available_operations' in st.session_state:
        df_available_operations=st.session_state.df_available_operations
        query_operation = st.text_input("Filtrar Operacion",value="",key='query_operation')
        if query_operation!='':
            df_available_operations=df_available_operations[df_available_operations['operation_description'].str.contains(query_operation.upper())]
        df_selected_operations=st.data_editor(df_available_operations,key="available_operations")
        st.session_state.df_selected_operations=df_selected_operations[df_selected_operations['Programar']==True][['operation_description']]

    if st.button("Buscar maquinas disponibles"):
        machine_selection()

    # @note Master management
    if 'df_master_long' in st.session_state:
        msg_error_avail=st.empty()
        df_filtered_master=st.session_state.df_master_long
        query_master = st.text_input("Filtrar Componente",value="",key='query_master')
        if query_master!='':
            df_filtered_master=df_filtered_master[df_filtered_master['pn'].str.contains(query_master.upper())]
        df_filtered_master=st.data_editor(df_filtered_master,key=f"master_long{st.session_state.editor_id}")
        idx_unselected_master=df_filtered_master['Disponible']==False
        df_selected_master=st.session_state.df_master_long
        if len(idx_unselected_master)>0:
            df_selected_master=df_selected_master[~idx_unselected_master]
        st.session_state.df_selected_master=df_selected_master
        df_avail_hours=st.session_state.df_avail_hours
        df_missing_avail=df_selected_master[(~df_selected_master['maq'].isin(df_avail_hours['maquina']))&(df_selected_master['Disponible']==True)]
        df_missing_avail=df_missing_avail[['maq']].drop_duplicates()
        st.session_state.df_missing_avail=df_missing_avail
        if len(df_missing_avail)>0:
            msg_error_avail=st.error("Se requiere disponibilidad en el archivo Horas disponibles para las maquinas:")
            st.dataframe(df_missing_avail)


    selected_date = st.date_input("Fecha inicial de programacion", value=state.get("initial_date", datetime.date.today()))
    if selected_date != state.get("initial_date", datetime.date.today()):
        state["initial_date"] = selected_date
        save_state_pickle(state,filename=path_pickle)
        st.rerun()

    selected_lim_date = st.date_input("Fecha final de programacion", value=state.get("limit_date", datetime.date.today()))
    if selected_lim_date != state.get("limit_date", datetime.date.today()):
        state["limit_date"] = selected_lim_date
        save_state_pickle(state,filename=path_pickle)
        st.rerun()

    if st.button("Crear Plan"):
        if not (state.get("folder_output") and 
                state["selections"].get("master_file") and 
                state["selections"].get("order_file") and 
                state["selections"].get("routing_file")):
            st.error("Por favor seleccione los archivos mandatorios.")
        else:
            msg_create_plan=st.success("Creando plan...")
            create_plan()

    if st.button("Generar reportes"):
        generate_reports()

    if st.button("Validar plan"):
        validate_plan()

