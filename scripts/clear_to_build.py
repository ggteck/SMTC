# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.16.7
#   kernelspec:
#     display_name: Python 3
#     language: python
#     name: python3
# ---

# %% [markdown]
# # Clear to build process
#  Toma varios archivos de inputs en excel para generar el analisis de lanzamiento prellenado, permite agregar datos al mismo y a partir de este y otros inputs genera el archivo CTB.xlsx con informacion necesaria para el analisis de la definicion de lo que se planea lanzar
# - V20. 2025-03-27
#   - Solo los Cortos se ponen en el reporte de gating parts
# - V19. 2025-03-23
#   - Creacion de gating parts report
# - V18. 2025-02-04
#   - Correccion de duplicados consolidando el independent demands
#   - Se conserva el ancho de columnas en el CTB y Analisis de lanzamiento
# - V17. 2025-01-22
#   - Manejo de error si no se encuentra ningun bom para la demanda
#   - Correccion en el uso IDemand REQ para las tablillas
# - V16. 2025-01-20
#   - Se cambia el orden de las versiones de mayor a menor
#   - Se corrige error al crear df_korrus vacio
#
# - V15. 2025-01-06  
#   - Se corrige el formato de fecha en el análisis de lanzamiento, hoja cortos-resumen
#
# - V14. 2024-12-18  
#   - Corrección en cantidad del archivo korrus  
#   - Cambio de nombre de col P.O. Unit Price  
#   - Filtrar las órdenes CANCELADAS en el WOS
#
# - V13. 2024-12-18  
#   - No se filtra el área BOX en las tablillas  
#   - En el Modelo en korrus no se toma en cuenta minusculas para coincidir con el independend demands  
#   - Se filtran registros en blanco del WOS
#
# - V12. 2024-12-17  
#   - Las tablas de alternos, WOS, PO WO Info no son mandatoria  
#   - Se crea un solo paso para el proceso de generación de CTB  
#   - Mejora el manejo de archivos  
#   - Uso de Work Order Action Report para generar la demanda de tablillas y su sugerencia de lanzamientos  
#   - Cambia en el reporte de cortos resumen, muestra los cortos restando la demanda de lo disponible, no del detalle de cortos  
#   - Uso del Korrus file para el requerimiento de todo el proceso, el requerimiento del independent demands se deja como IDemand REQ pero no se utiliza
#
# - V11. 2024-12-02  
#   - Corrección al generar pivote por familia, se ignora la demanda negativa  
#   - Se evita que se bloquee el CTB
#
# - V10. 2024-12-02  
#   - Corrección al guardar un df en excel
#
# - V9. 2024-11-29  
#   - Se agrega Llave  
#   - Corrección, no se estaba utilizando el Pendiente de recibo seleccionado
#
# - V8. 2024-11-26  
#   - Corrección de suma acumulativa en el reporte de cortos  
#   - Manejo de archivos para evitar que se queden bloqueados
#
# - V7. 2024-11-25  
#   - Se usa el formato real de llegadas de componentes  
#   - Corrección en Bom details  
#   - El CTB se abre al final  
#   - Formulas de Cobertura
#
# - V6. 2024-11-24  
#   - El notebook guarda y cierra CTB y Analisis de lanzamients  
#   - Se usa el analisis de lanzamiento como reporte de sugerencia de lanzamientos, el orden en este archivo es el orden de asignacion  
#   - Se ordena por Prioridad manual, fecha y cantidad menor para asignar lanzamientos  
#   - Se omiten del calculo requerimientos con componentes criticos faltantes  
#   - Los componentes criticos se definen por P Family  
#   - Se omite del calculo lo que tiene mas cortos de los permitidos en el selector "Cantidad maxima de cortos permitidos"  
#   - Se agregan llegadas de componente (Necesario corregir con un reporte real)  
#   - Se dio formato a los reportes
#
# - V5. 2024-11-22  
#   - Se cambia el metodo de seleccion de archivos de trabajo  
#   - Correcciones a las formulas del CTB  
#   - Se agregan reportes de cortos, link al reporte de cortos, resumen de cortos  
#   - En el proceso de sugerencia de lanzamientos, los requerimientos se asignan completos
#
# - V4. 2024-11-18  
#   - Se elimina el proceso de Recibos, se tomara solo el archivo Pendiente de Recibos que generara el usuario  
#   - En el analisis de Lanzamiento se toma en cuenta solo lo del Area BOX  
#   - Se elimina el ultimo guion en las familias  
#   - Se agrega la cantidad por bom de cada familia y la cantidad mayor en un bom por componente (QTY PER MAX GRAL)  
#   - Se agrega la sumatoria de cada familia  
#   - Se agrega la cantidad en el Forecast de cada componente  
#   - Se agregan las columnas de Simulacion y CTB
#
# - V3. 2024-11-07  
#   - Carpeta adicional para las Work Orders  
#   - Se agrega el BOM Details como input para la columna Primary Stock
#
# - V2. 2024-10-23  
#   - Mejoras y correcciones:  
#     - Se definio el nombre de hoja a leer en el archivo de BOM  
#     - Se cambio el archivo Manifest por Pendiente  
#     - Se corrigieron f' por f" para correr en jupyterlab  
#     - Se agrego formato a celdas del CTB  
#     - Se agrego el proceso de Propuesta de Lanzamiento  
#     - Se agrego pend, recibo a la formula de OH Dispo (Validar)  
#     - Se pone 0 en lugar de valores nulos en CTB  
#     - Genera sugerencia de lanzamientos
#
# - V1. 2024-10-21  
#   - Version inicial
#

# %% [markdown]
# # 1. Seleccionar archivos de trabajo
# Ejecutar la celda siguiente y seleccione la carpeta y los archivos a procesar

# %% jupyter={"source_hidden": true}
# Seleccion de archivos de trabajo
import pandas as pd
import os
import ipywidgets as widgets
from ipywidgets import  Layout,HTML
from IPython.display import display, Markdown, clear_output
import pickle
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.formatting.rule import CellIsRule
from tqdm.notebook import tqdm
from tkinter import Tk, filedialog as fd
import win32com.client
import warnings
from copy import copy 
import gc

warnings.filterwarnings("ignore")
def open_file_selection(initialdir='',filter_name='*.*'):
    """
    Opens a file selection dialog and returns the selected file paths.
    
    :param initialdir: The initial directory to open in the dialog.
    :return: A tuple of selected file paths or an empty tuple if no file is selected.
    """
    # Create a hidden root window
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)  # Ensure the dialog appears on top
    
    filetypes = (
        (filter_name,f"*{filter_name.split(' ')[0]}*.*"),
        ('All files', '*.*'),
        ('Excel', '*.xlsx'),
        ('CSV', '*.doc')
    )
    files = fd.askopenfilenames(
        filetypes=filetypes,
        initialdir=initialdir
    )
    
    # Destroy the root window after use
    root.destroy()
    
    return files
def select_directory(initialdir):
    """
    Opens a directory selection dialog and returns the selected directory path.
    
    :param initialdir: The initial directory to open in the dialog.
    :return: The selected directory path as a string, or an empty string if no directory is selected.
    """
    # Create a hidden root window
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)  # Ensure the dialog appears on top

    # Open the directory selection dialog
    directory = fd.askdirectory(initialdir=initialdir)

    # Destroy the root window after use
    root.destroy()

    return directory

# Function to show a pop-up message
def show_popup_message(message):
    display(Markdown(f"### **{message}**"))
    

# Decorator to handle the permission error
def handle_permission_error_with_popup(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except PermissionError as e:
            if e.errno == 13:  # Permission denied error
                show_popup_message(f"Error: {e}\nFavor de cerrar el archivo.")
    return wrapper

# @handle_permission_error_with_popup
# def save_df(df, filepath,sheet_name,index):
#     df.to_excel(filepath,sheet_name=sheet_name,index=index)

def save_df(df, filepath, sheet_name, index=False):
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index)

@handle_permission_error_with_popup
def save_df_multiple(df_dict=dict(), filepath='',index=False):
    with pd.ExcelWriter(filepath) as writer:
        for key in df_dict.keys():
            df_dict[key].to_excel(writer, sheet_name=key, index=index)

@handle_permission_error_with_popup
def save_wb(wb, filepath):
    wb.save(filepath)

@handle_permission_error_with_popup
def append_sheet(df,path,sheet_name,index):
    with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index) 


def read_excel(path=None,sheet_name=0,header=0):
    with pd.ExcelFile(path) as xls:
        df=pd.read_excel(path,sheet_name=sheet_name,header=header)
    return df

def verify_selections(file_selectors):
    not_selected=[]
    # not_selected=[file_selector.children[0].description[0:-1] for file_selector in file_selectors if (file_selector.children[1].value.strip()=='Not selected') and (file_selector.children[0].description[0:-1] in mandatory_selectors)]
    for file_selector in file_selectors:
        selector=file_selector.children[0].description[0:-1]
        selected=file_selector.children[1].value.strip()
        if not os.path.exists(selected):
            file_selector.children[1].value='Not selected'
            selected='Not selected'
        if selector not in mandatory_selectors:
            continue
        if selected=='Not selected':
            not_selected.append(selector)
            continue

    if len(not_selected)>0:
        show_popup_message(f'Favor de seleccionar los siguientes archivos:{not_selected}')
        raise SystemExit()

def close_xl_if_open(path):
    if is_file_open(path):
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            workbook = excel.Workbooks(path)
            workbook.Save()
            workbook.Close()
        except:
            show_popup_message(f"Cerrar el archivo: {path}")
            raise SystemExit()               

def get_bom():
    path_bom=get_path(file_selectors,'BOM')
    if not os.path.exists(path_bom):
        show_popup_message(f"No se ha encontrado el archivo {path_bom}")
        raise SystemExit()   
    df_bom=read_excel(path_bom,sheet_name='Flat Bill Browser - Cost Roll U')
    check_mandatory_cols(df_bom.columns,'BOM')
    df_bom=df_bom[mandatory_cols['BOM']]
    df_bom[df_bom['Type']!='Y']
    df_bom.rename({'BOM':'MODELO'},axis=1,inplace=True)
    return df_bom

##### Excel management functions
def find_cell_by_text(ws, text):
    """
    Find the cell containing the specified date in the worksheet.
    
    :param ws: The worksheet object from openpyxl
    :param date_value: The date value to search for (datetime object or string)
    :return: The cell address (e.g., 'B2') or None if not found
    """
    # for row in ws.iter_rows():
    for row in ws[ws.calculate_dimension()]:
        for cell in row:
            if cell.value == text:
                return cell.coordinate  # Return cell address (e.g., 'B2')
    return None 

def find_all_cells_by_text(ws, text):
    """
    Find the cell containing the specified date in the worksheet.
    
    :param ws: The worksheet object from openpyxl
    :param date_value: The date value to search for (datetime object or string)
    :return: The cell address (e.g., 'B2') or None if not found
    """
    # for row in ws.iter_rows():
    cells=[]
    for row in ws[ws.calculate_dimension()]:
        for cell in row:
            if cell.value == text:
                cells.append(cell.coordinate)
    return cells 

def set_number_format(ws,col_name,format):
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    col_head_address=find_cell_by_text(ws,col_name)
    for row in ws[f"{col_head_address}:{last_cell}"]:
        row[0].number_format = format

def fill_column(ws,col_name,fill):
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    col_head_address=ws[find_cell_by_text(ws,col_name)].offset(1,0).coordinate
    for row in ws[f"{col_head_address}:{last_cell}"]:
        row[0].fill=fill

def font_column(ws,col_name,font):
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    col_head_address=ws[find_cell_by_text(ws,col_name)].offset(1,0).coordinate
    for row in ws[f"{col_head_address}:{last_cell}"]:
        row[0].font=font    



def fill_formula(ws,col_name,formula):
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    col_head_address=ws[find_cell_by_text(ws,col_name)].offset(1,0).coordinate
    for row in ws[f"{col_head_address}:{last_cell}"]:
        row[0].value=formula

def get_col_sizes(wb):
    """
    Obtains a dictionary with the column sizes of each sheet in a workbook
    """
    col_sizes={}
    for sheet in wb.sheetnames:
        col_sizes[sheet]=wb[sheet].column_dimensions
    return col_sizes


def apply_col_sizes(wb,col_sizes):
    """
    Applies the column sizes obtained by function get_col_sizes
    """
    for sheet in col_sizes.keys():
        if sheet not in wb.sheetnames:
            continue
        ws=wb[sheet]
        for column_letter in col_sizes[sheet].keys():
            if column_letter==0:
                continue
            ws.column_dimensions[column_letter].width=col_sizes[sheet][column_letter].width

# Cell colors
light_green='99FF99'
avocato_green='E2EFDA'
dark_blue='002060'
light_blue='DDEBF7'
grey='595959'
dark_red='9C0006'
melon='FCE4D6'
white='FFFFFF'
black='000000'
light_pink='FFC7CE'
light_yellow='FFF2CC'

def is_file_open(filepath):
    # Check if the file exists
    if not os.path.exists(filepath):
        return False  # File does not exist, so treat it as "open" for your logic

    try:
        # Try to open the file in write mode
        with open(filepath, 'a'):
            pass
        return False  # File is not open (no exception raised)
    except PermissionError:
        return True 
    
def get_path(file_selectors,selector_name):
    file_selector=[file_selector for file_selector in file_selectors if file_selector.children[0].description[0:-1]==selector_name]
    if len(file_selector)==0:
        show_popup_message(f"No se encontro el selector")
        raise SystemExit()
    file_selector=file_selector[0]
    if not file_selector.children[1].value:
        return None
    selected_path=file_selector.children[1].value.strip()
    return selected_path

# Save and load state using pickle
def save_state_pickle(state, filename='folder_state.pkl'):
    with open(filename, 'wb') as f:
        pickle.dump(state, f)

def load_state_pickle(filename='folder_state.pkl'):
    try:
        with open(filename, 'rb') as f:
            return pickle.load(f)
    except FileNotFoundError:
        return {"folder_input": None, "folder_output": None, "selections": {}}

def set_opposites(name_a,name_b):
    opposite = {name_a: name_b, name_b: name_a}
    return opposite
    

def update_oposite_selectors(filter_name,selector_name_a,selector_name_b):
    opposite = {selector_name_a: selector_name_b, selector_name_b: selector_name_a}
    if filter_name in [selector_name_a, selector_name_b]:
        # Remove the opposite action from mandatory selectors
        if opposite[filter_name] in mandatory_selectors:
            mandatory_selectors.remove(opposite[filter_name])
        # Add current filter name if it's not already in the list
        if filter_name not in mandatory_selectors:
            mandatory_selectors.append(filter_name)
        
        # Update file selectors to set the opposite field to 'Not selected'
        for file_selector in file_selectors:
            if file_selector.children[0].description[:-1] == opposite[filter_name]:
                file_selector.children[1].value = 'Not selected'
                state[opposite[filter_name]]=f" Not selected"
                save_state_pickle(state)
    
# Agregar las columnas criticas por archivo
mandatory_cols={'Pendiente':[
                            'Tracking No',
                            'Part No',
                            'Qty',
                            'PO',
                            'Tally No',
                            'Date Shipped'
                            ],
                'Recibos':[
                                'Part No.',
                                'Qty. of TRX',
                                'Reference'
                            ],
                'WOS':[
                    'PO cliente',
                    'Modelo',
                    'WO\n QTY',
                    'WO'
                ],
                'Independent Demands':[
                        'Reference Notes',
                        'Part Name',
                        'P Family',
                        'FCST / RR  Unit Price',
                        'Qty',
                        'Due Date'
                    ],
                'Korrus':[
                    'PurchaseOrder',
                    'ProductServiceID',
                    'Quantity'
                ],
                'Work Order Action':[
                        'Part',
                        'Type',
                        'Due Date',
                        'WO QTY'
                ],
                'BOM':[
                    'BOM',
                    'Component',
                    'Qty Per',
                    'Type'
                ],
                'BOM Detail':[
                    'Flat Component',
                    'Primary Stock'
                ],                
                'Consumption':[
                    'Site',
                    'Name',
                    'Description',
                    'Note',
                    'Std Unit Cost',
                    'Total',
                    ' PP',
                    'NR',
                    'Allocation',
                    'IssueToWo',
                    'ABC Code',
                    'Buyer Code',
                    'Manufacturer',
                    'MPN',
                    'LT',
                    'UOM',
                    'WhereUsed'
                ],
                'On Hand Detail':[
                    'Part',
                    'Quantity'
                ],
                'Alternos':[
                    'HL1Z (New LED)',
                    'CCT/CRI',
                    'APC (ZES Old LED)'
                ],
                'PO WO Info':[
                    'Order Type',
                    'Part Name',
                    'Due',
                    'Quantity'
                ],
                'Tablillas':[
                    'MODELO',
                    'REQ'
                ],
                'Component Allocation':[
                    'Component',
                    'Qty To be Issued'
                ]

                }
mandatory_selectors=[
    'Consumption', 
    'Independent Demands',
    'Korrus',
    'Work Order Action', 
    'On Hand Detail', 
    'Pendiente', 
    'BOM']

excluding_selectors=['Work Order Action','Independent Demands']

def check_mandatory_cols(cols,selector_name):
    missing_columns = [col for col in mandatory_cols[selector_name] if col not in cols]
    if len(missing_columns)>0:
        show_popup_message(f"No se encontraron las siguientes columnas en el archivo {selector_name}: {missing_columns}")
        raise SystemExit()
    return
state = load_state_pickle()
if ('folder_output' not in state.keys()):
    state['folder_output']=''

folder_output_button = widgets.Button(description="Folder de salidas:")
if state['folder_output']:
    initialdir=state['folder_output']
else:
    initialdir='Not selected'
folder_output_label = widgets.Label(value=initialdir)

def on_output_button_click(b):
    if state['folder_output']:
        initialdir=state['folder_output']
    else:
        initialdir='/'
    selected_dir = select_directory(initialdir=initialdir)
    if selected_dir:
        folder_output_label.value = f"{selected_dir}"
        state['folder_output']=selected_dir
        save_state_pickle(state)
    else:
        folder_output_label.value = "Not selected"

# Attach the event to the folder_output_button
folder_output_button.on_click(on_output_button_click)

filters = ['Consumption', 
           'Independent Demands', 
           'Korrus',
           'On Hand Detail', 
           'Pendiente', 
           'WOS', 
           'BOM', 
           'BOM Detail', 
           'Alternos',
           'PO WO Info',
           'Work Order Action', 
           'Tablillas',
           'Component Allocation']

# Create an array of button and label widgets
file_selectors = []
for filter_name in filters:
    # Create button and label
    button = widgets.Button(description=f"{filter_name}:")
    if state.get(filter_name):
        label = widgets.Label(value=f" {state.get(filter_name)}")
    else:
        label = widgets.Label(value=f" Not selected")
    
    # Define the button click event
    def on_button_click(filter_name=filter_name, label=label):
        if (filter_name in state.keys()) and (state[filter_name]):
            initialdir=os.path.dirname(state[filter_name][0])
        else:
            initialdir='/'
        selected_dir = open_file_selection(initialdir=initialdir,filter_name=filter_name)  # Adjust initialdir as needed
        if selected_dir:
            label.value = f" {selected_dir[0]}"
            state[filter_name]=selected_dir[0]
            update_oposite_selectors(filter_name,excluding_selectors[0],excluding_selectors[1])
            
        else:
            label.value = f" Not selected"
            state[filter_name]=f" Not selected"
        save_state_pickle(state)

    # Attach the event to the button
    button.on_click(lambda b, f=filter_name, l=label: on_button_click(f, l))
    
    # Add the button and label as a horizontal box
    file_selectors.append(widgets.HBox([button, label]))



ui = widgets.VBox([widgets.HBox([folder_output_button, folder_output_label])]+file_selectors)
display(ui)

for file_selector in file_selectors:
    filter_name=file_selector.children[0].description[:-1]
    if filter_name in excluding_selectors:
        file_selector.children[1].value = 'Not selected'
        state[filter_name]=f" Not selected"
        save_state_pickle(state)


# %% [markdown]
# # CTB calculo completo

# %% [markdown]
# ## 2. Analisis de lanzamiento
# - Utiliza los siguientes inputs para el analisis de requerimiento por lanzar
#     - Independent Demands
#     - WOS Lanzadas
# - Agregue comentarios en las columnas "Comentarios" y "Status de Lineas"
# - El proceso obtendra los comentarios del ultimo analisis de lanzamiento generado por lo que se puede correr varias veces
# - No conserva comentarios de registros eliminados, es decir, si la demanda no existe, no apareceran sus comentarios

# %%
# Verificar selecciones
verify_selections(file_selectors)

# %%
# Analisis de lanzamiento
path_demanda_tablillas=get_path(file_selectors,'Work Order Action')
if path_demanda_tablillas=='Not selected':
    path_launch=os.path.join(folder_output_label.value,'Analisis_lanzamiento.xlsx')
else:
    path_launch=os.path.join(folder_output_label.value,'Analisis_lanzamiento_tablillas.xlsx')
close_xl_if_open(path_launch)

df_notes=pd.DataFrame()
if os.path.exists(path_launch):
    df_notes=pd.read_excel(path_launch,sheet_name='Analisis de lanzamiento')
    df_notes.dropna(subset=['Comentarios', 'Status de Lineas'], how='all', inplace=True)
    df_notes=df_notes[['PO','MODELO','Comentarios','Status de Lineas']].drop_duplicates()

path_lanzadas=get_path(file_selectors,'WOS')
df_lanzadas=pd.DataFrame(columns=['PO', 'MODELO', 'WO', 'QTY WO', 'Total Wos lanzadas'])
if (path_lanzadas) and (path_lanzadas!='Not selected'):
    df_lanzadas_wb=pd.read_excel(path_lanzadas,sheet_name=None)
    df_lanzadas=pd.DataFrame()
    for sheet in df_lanzadas_wb.keys():
        if 'Seguimiento' in sheet:
            df=df_lanzadas_wb[sheet]
            if 'Unnamed' in df.columns[0]:
                df.columns=df.iloc[0]
                df.drop(0,inplace=True)
            check_mandatory_cols(df.columns,selector_name='WOS')
            df=df[~df['Modelo'].isna()]
            df_lanzadas=pd.concat([df_lanzadas,df])
    if path_demanda_tablillas!='Not selected':
        df_lanzadas=df_lanzadas[df_lanzadas['Area']=='BOX']
    df_lanzadas_raw=df_lanzadas.copy()
    df_lanzadas['status Orden'].fillna('',inplace=True)
    df_lanzadas=df_lanzadas[~df_lanzadas['status Orden'].str.upper().str.contains('CANCELADA')]
    df_lanzadas.rename({"PO cliente":"PO",
                    "Modelo":"MODELO"
                    },axis=1,inplace=True)
    df_lanzadas=df_lanzadas.groupby(['PO','MODELO']).agg({
        'WO':'first',
        'WO\n QTY':['first','sum']

    })
    df_lanzadas.columns = ['_'.join(col).strip() for col in df_lanzadas.columns.values]
    df_lanzadas.rename({"WO_first":"WO",
                    "WO\n QTY_first":"QTY WO",
                    "WO\n QTY_sum":"Total Wos lanzadas"
                    },axis=1,inplace=True)
    df_lanzadas.reset_index(inplace=True)

if path_demanda_tablillas!='Not selected':
    # Demanda de tablillas
    df_demand=read_excel(path_demanda_tablillas,sheet_name='Work Order Action Report')
    check_mandatory_cols(df_demand.columns,'Work Order Action')
    df_demand.rename({"UDF Ref":"P Family",
                      "Part":"MODELO",
                      "WO QTY":"REQ"},axis=1,inplace=True)
    df_demand['PO']='PO'+df_demand.index.astype(str)
    df_demand['Price']=0
    df_demand['IDemand REQ']=df_demand['REQ']
    df_demand=df_demand[(df_demand['MODELO'].str[0:4]=='APC-') & (df_demand['P Family']=='RISE')]


df_korrus=pd.DataFrame(columns=['PO','MODELO','Quantity'])
path_demand=get_path(file_selectors,'Independent Demands')
if path_demand!='Not selected':
    # Demanda de componentes finales
    df_demand=read_excel(path_demand,sheet_name='Independent Demands')
    check_mandatory_cols(df_demand.columns,'Independent Demands')
    df_demand.rename({"Reference Notes":"PO",
                  "Part Name":"MODELO",
                  "FCST / RR  Unit Price":"Price",
                  "Qty":"REQ"
                  },axis=1,inplace=True)
    
    df_demand=df_demand.groupby(['P Family','PO','MODELO','Due Date']).agg({
        'REQ':'sum',
        'Price':'first'
    })
    df_demand.reset_index(inplace=True)

    path_korrus=get_path(file_selectors,'Korrus')
    df_korrus=read_excel(path_korrus,sheet_name='Data')
    check_mandatory_cols(df_korrus.columns,'Korrus')
    df_korrus.rename({'PurchaseOrder':'PO','ProductServiceID':'MODELO'},axis=1,inplace=True)
    df_korrus['MODELO']=df_korrus['MODELO'].str.upper()
    df_korrus=df_korrus[['PO','MODELO','Quantity']].groupby(['PO','MODELO']).sum()[['Quantity']]
    df_korrus.reset_index(inplace=True)
    df_demand=df_demand.merge(df_korrus,how='left',on=['MODELO','PO'])
    df_demand.rename({'REQ':'IDemand REQ','Quantity':'REQ'},axis=1,inplace=True)
else:
    df_demand=df_demand.groupby(['P Family','PO','MODELO','Due Date']).agg({
        'IDemand REQ':'sum',
        'REQ':'sum',
        'Price':'first'
    })
    df_demand.reset_index(inplace=True)
    df_demand=df_demand.merge(df_korrus,how='left',on=['MODELO','PO'])
df_demand_raw=df_demand.copy()


df_demand['Sales']=df_demand['Price']*df_demand['REQ']
df_demand=df_demand.merge(df_lanzadas,how='left',on=['PO','MODELO'])

df_demand['WO']=df_demand['WO'].fillna(0).astype(int)
#Familia para la demanda por lanzar
idx=df_demand['PO'].str.contains('FC')

#Familia para el forecast
if path_demand!='Not selected':
    df_demand.loc[~idx,'Familia']=df_demand.loc[~idx,'MODELO'].str[:8]
    df_demand.loc[idx,'Familia']='FC-'+df_demand.loc[idx,'P Family']
    df_demand['Familia']=df_demand['Familia'].str.rstrip('-')
else:
    # No hay familias para el proceso de tablillas
    df_demand['Familia']=df_demand['MODELO']
df_demand=df_demand[['MODELO','PO','Due Date','P Family','Familia','IDemand REQ','REQ','WO','QTY WO','Total Wos lanzadas']]
df_demand[['QTY WO','Total Wos lanzadas']]=df_demand[['QTY WO','Total Wos lanzadas']].fillna(0).astype(int)
df_demand['QTY Pend. Lanzar']=df_demand['REQ']-df_demand['Total Wos lanzadas']

df_demand['Due Date']=df_demand['Due Date'].dt.date


if len(df_notes)>0:
    df_demand=df_demand.merge(df_notes,how='left',on=['PO','MODELO'])
else:
    df_demand['Comentarios']=''
    df_demand['Status de Lineas']=''
df_demand=df_demand.sort_values(['MODELO','PO'])
df_demand_launch=df_demand[~df_demand['PO'].str.contains('FC')]
df_demand_launch['Omitir']=''
df_demand_launch['Estatus']=''
df_demand_launch['Alta Prioridad']=''
save_df(df=df_demand_launch,filepath=path_launch,sheet_name='Analisis de lanzamiento',index=False)
if len(df_lanzadas)>0:
    append_sheet(df_lanzadas_raw, path_launch,'WOS Lanzadas',index=False)
append_sheet(df_demand_raw, path_launch,'Independent Demands',index=False)
# try:
#     os.startfile(path_launch)
# except:
#     show_popup_message(f"El Analisis de lanzamiento esta listo.")

# %% [markdown]
# ## 3. Consumption
# - Crea el archivo CTB tomando los siguientes inputs:
#     - BOMs de Lanzamiento
#     - Consumption
#     - Tabla de alternos
#     - Calendario. Es importante llenar el calendario para fechas mas alla de la demanda

# %% [markdown]
# ### 3.1 Consolidacion de BOM y Demanda

# %%
#BOM y Demanda para CTB

df_bom=get_bom()
df_bom_demand=df_demand.copy()
df_bom_demand['Comentarios']=df_bom_demand['Comentarios'].fillna('')
df_bom_demand=df_bom_demand[~df_bom_demand['Comentarios'].str.upper().str.contains('LANZAMIENTO COMPLETO')]
df_bom_demand=df_bom_demand[['PO','MODELO','P Family','Familia','QTY Pend. Lanzar']]
df_bom_demand=df_bom_demand.merge(df_bom,how='left',on='MODELO')

df_missing_boms_demand=df_bom_demand[df_bom_demand['Qty Per'].isnull()]
df_bom_demand['Type'].fillna('',inplace=True)
df_bom_demand['Qty Per']=df_bom_demand['Qty Per'].fillna(0)
# Llenar Forecast BOM, el modelo es el componente y la cantidad es 1
df_bom_demand.loc[df_bom_demand['PO'].str.contains('FC'),'Component']=df_bom_demand.loc[df_bom_demand['PO'].str.contains('FC'),'MODELO']
df_bom_demand.loc[df_bom_demand['PO'].str.contains('FC'),'Qty Per']=1
df_bom_demand['REQ. Total']=df_bom_demand['Qty Per']*df_bom_demand['QTY Pend. Lanzar']

df_bom_max=df_bom_demand.groupby(['Component']).max()[['Qty Per']]
df_bom_max.reset_index(inplace=True)
df_bom_max.rename({'Qty Per':'QTY PER MAX GRAL'},axis=1,inplace=True)

path_bom_demand=os.path.join(folder_output_label.value,'bom_demand.xlsx')
save_df(df_bom_demand,path_bom_demand,'bom_demand',index=False)


df_bom_demand=df_bom_demand[df_bom_demand['REQ. Total']>0]

if len(df_bom_demand)==0:
    show_popup_message('No se encontraron BOMS para la demanda, favor de revisar el archivo de BOMS')
    raise SystemExit()
    


pivot_bom_demand = pd.pivot_table(df_bom_demand, 
                       values=['REQ. Total','Qty Per'],  
                       index=['Component'],         
                       columns=['P Family','Familia'],       
                       aggfunc={'REQ. Total': 'sum', 'Qty Per': 'last'},
                       fill_value=0
                        ) 


# Flatten the columns after the pivot
pivot_bom_demand.columns = [' '.join(col).strip() for col in pivot_bom_demand.columns.values]
# Sort columns by name to intercalate 'REQ. Total' and 'Qty Per' for each 'Familia'
sorted_columns = sorted(pivot_bom_demand.columns, key=lambda x: (x.split()[-1], x.split()[0]))
pivot_bom_demand = pivot_bom_demand[sorted_columns]
pivot_bom_demand.reset_index(inplace=True)
# Lista de columnas para sumatorias por categoria de familia, no se suman las columnas Forecast (FC)
dict_req_total={}
for pfamily in df_bom_demand['P Family'].drop_duplicates().to_list():
    cols_to_sum=[x for x in pivot_bom_demand.columns if ((f"REQ. Total {pfamily}" in x)&('FC-' not in x))]
    pivot_bom_demand[pfamily]=pivot_bom_demand[cols_to_sum].sum(axis=1)
fc_qty_cols=[x for x in pivot_bom_demand.columns if (('FC-' in x)&('Qty Per' in x))]
fc_req_cols=[x for x in pivot_bom_demand.columns if (('FC-' in x)&('REQ. Total' in x))]
pivot_bom_demand.drop(fc_qty_cols,axis=1,inplace=True)
# I will use family columns for the format
family_cols=pivot_bom_demand.columns
family_cols=family_cols[1:]
pivot_bom_demand=pivot_bom_demand.merge(df_bom_max,how='left',on='Component')


# %% [markdown]
# ### 3.2 Consolidar Consumption
# Este proceso debe tomar unos segundos, si llega a un minuto, revisar si hay mensajes en excel

# %% jupyter={"source_hidden": true}
#Consumption
def get_mondays(n):
    today = datetime.today()
    
    # Calculate the Monday of the current week
    monday = today - timedelta(days=today.weekday())  # `today.weekday()` gives 0 for Monday, so we subtract it
    
    # Generate the next `n` Mondays including the current week's Monday
    mondays = [monday + timedelta(weeks=i) for i in range(n + 1)]

    return mondays

path_ctb=os.path.join(folder_output_label.value,'CTB KRS.xlsx')

if path_demanda_tablillas=='Not selected':
    path_ctb=os.path.join(folder_output_label.value,'CTB KRS.xlsx')
else:
    path_ctb=os.path.join(folder_output_label.value,'CTB KRS_tablillas.xlsx')

col_sizes=[]
if os.path.exists(path_ctb):
    wb=load_workbook(path_ctb)
    col_sizes=get_col_sizes(wb)

path_consumption=get_path(file_selectors,'Consumption')

df_cons=pd.read_excel(path_consumption,header=1)
check_mandatory_cols(df_cons.columns,'Consumption')
#Crear el CTB inicial solo con el RL
wb = load_workbook(path_consumption)
save_wb(wb=wb,filepath=path_ctb)
df_missing_boms_demand
df_missing_boms_demand=df_missing_boms_demand[~df_missing_boms_demand['PO'].str.contains('FC')]
if len(df_missing_boms_demand)>0:
    append_sheet(df_missing_boms_demand,path_ctb,'Missing boms',index=False)
demand_cols=[x for x in df_cons.columns if "Demand" in x]
scd_cols=[x for x in df_cons.columns if "Schd" in x]
general_cols=['Site',
'Name',
'Description',
'Note',
'Std Unit Cost',
'Total',
' PP',
'NR',
'Allocation',
'IssueToWo',
'ABC Code',
'Buyer Code',
'Manufacturer',
'MPN',
'LT',
'UOM',
'WhereUsed']

ctb_cols=[
    'Name',
    'Description',
    'Note',
    'Total',
    ' PP',
    'NR',
    'Allocation',
]

mondays = get_mondays(len(demand_cols)-1)

df_cols=pd.DataFrame(columns=['demand_cols'],data=demand_cols)
df_cols['scd_cols']=scd_cols
df_cols['monday_date']=mondays
df_cols['monday_date']=pd.to_datetime(df_cols['monday_date']).dt.date.astype(str)
df_cols['week']=pd.to_datetime(df_cols['monday_date']).dt.isocalendar().week
df_calendar=pd.read_excel('calendario.xlsx')
df_calendar['monday_date']=df_calendar['monday_date'].astype(str)
df_calendar['demand_year_month']=df_calendar['year'].astype(str)+'-'+df_calendar['closing_month'].astype(str)
df_calendar['scd_year_month']=df_calendar['demand_year_month'].shift(-1)
df_cols=df_cols.merge(df_calendar,on='monday_date',how='left')
df_rl=pd.DataFrame()
for index,row in df_cols.iterrows():
    df=df_cons[general_cols+[row['demand_cols'],row['scd_cols']]]
    df['demand_year_month']=row['demand_year_month']
    df['scd_year_month']=row['scd_year_month']
    df['monday_date']=row['monday_date']
    df['year']=row['year']
    df.rename({row['demand_cols']:'Demand',row['scd_cols']:'SchdRcpt'},axis=1,inplace=True)
    df_rl=pd.concat([df,df_rl])
    

if len(df_rl[df_rl['demand_year_month'].isnull()])>0:
    show_popup_message("Cuidado, se encontraron fechas sin cierre, favor de completar el calendario")



df_d=df_rl[general_cols+['Demand','demand_year_month','monday_date','year']]
df_d['type']='Demand'
df_d.rename({'Demand':'qty','demand_year_month':'year_month'},axis=1,inplace=True)
df_r=df_rl[general_cols+['SchdRcpt','scd_year_month','monday_date','year']]
df_r['type']='SchdRcpt'
df_r.rename({'SchdRcpt':'qty','scd_year_month':'year_month'},axis=1,inplace=True)
df_rl_raw=pd.concat([df_d,df_r])

df_rl_raw[general_cols]=df_rl_raw[general_cols].fillna('')
df_rl[general_cols]=df_rl[general_cols].fillna('')
rl_raw_path=os.path.join(folder_output_label.value,'rl_raw.xlsx')
save_df(df=df_rl_raw,filepath=rl_raw_path,sheet_name='RL',index=False)

df_rl_raw.fillna('',inplace=True)

# %% [markdown]
# ### 3.3 On hand Detail

# %% jupyter={"source_hidden": true}
#On hand detail
if is_file_open(path_ctb):
    show_popup_message(f"Favor de cerrar un archivo {path_ctb}")
    raise SystemExit()  

path_on_hand=get_path(file_selectors,'On Hand Detail')
df_onhand=pd.read_excel(path_on_hand)
check_mandatory_cols(df_onhand.columns,'On Hand Detail')
df_onhand=df_onhand.groupby(['Part']).sum()[['Quantity']]
df_onhand.reset_index(inplace=True)
append_sheet(df_onhand,path_ctb,'On Hand Detail',False)

df_onhand=df_onhand.groupby(['Part']).sum()[['Quantity']]
df_onhand.reset_index(inplace=True)
append_sheet(df_onhand,path_ctb,'On Hand Detail',False)

# %% [markdown]
# ### 3.4 Recibos

# %% jupyter={"source_hidden": true}
#Recibos
path_manifest=get_path(file_selectors,'Pendiente')
df_manifest=read_excel(path_manifest)
check_mandatory_cols(df_manifest,'Pendiente')
df_manifest=df_manifest[mandatory_cols['Pendiente']]

# %% [markdown]
# ### 3.5 CTB
# - No se conservan cambios manuales

# %% jupyter={"source_hidden": true}
#CTB
if is_file_open(path_ctb):
    show_popup_message(f"Favor de cerrar un archivo {path_ctb}")
    raise SystemExit()  

df_rl_raw=read_excel(rl_raw_path)
df_rl_raw.fillna('',inplace=True)

df_alloc=df_rl_raw.drop_duplicates(['Name'])[['Name','Allocation']]
df_alloc.rename({'Name':'Alterno','Allocation':'Aloc Aterno'},axis=1,inplace=True)

#Agregar alternos
path_alternos=get_path(file_selectors,'Alternos')
if (path_alternos) and (path_alternos!='Not selected'):
    df_altern=read_excel(path_alternos)
    check_mandatory_cols(df_altern.columns,'Alternos')
    df_altern=df_altern.merge(df_onhand,left_on='APC (ZES Old LED)',right_on='Part')
    df_altern.drop(['APC (ZES Old LED)','CCT/CRI'],axis=1,inplace=True)
    df_altern.rename({'HL1Z (New LED)':'Name','Part':'Alterno','Quantity':'ON hand'},axis=1,inplace=True)
    df_altern=df_altern.merge(df_alloc,how='left',on='Alterno')
    df_rl_raw=df_rl_raw.merge(df_altern,how='left',on='Name')
else:
    df_rl_raw[['Alterno','ON hand','Aloc Aterno']]=["",0,0]
    
#Agregar tablillas adicionales
path_adicionales=get_path(file_selectors,'Tablillas')
if (path_adicionales) and (path_adicionales!='Not selected'):
    df_adicionales=read_excel(path_adicionales)
    check_mandatory_cols(df_adicionales.columns,'Tablillas')
    df_adicionales=df_adicionales.groupby('MODELO').sum('REQ')[['REQ']]
    df_adicionales.reset_index(inplace=True)
    df_adicionales.rename({"REQ":"Adicional","MODELO":"Name"},axis=1,inplace=True)
    if 'Omitir' in df_adicionales.columns:
        df_adicionales=df_adicionales[df_adicionales['Omitir'].isnull()]
    path_adic_alloc=get_path(file_selectors,'Component Allocation')
    if (path_adic_alloc) and (path_adic_alloc!='Not selected'):
        df_adic_alloc=read_excel(path_adic_alloc,sheet_name='Workorder Component Allocations')
        check_mandatory_cols(df_adic_alloc.columns,'Component Allocation')
        df_adic_alloc=df_adic_alloc.groupby('Component').sum('Qty To be Issued')[['Qty To be Issued']]
        df_adic_alloc.reset_index(inplace=True)
        df_adic_alloc.rename({"Component":"Name","Qty To be Issued":"Aloc Adicional"},axis=1,inplace=True)
        df_adicionales=df_adicionales.merge(df_adic_alloc,how='left',on='Name')
    df_adicionales=df_adicionales.merge(df_onhand,how='left',left_on='Name',right_on='Part')
    df_adicionales.drop('Part',inplace=True,axis=1)
    df_adicionales.rename({'Quantity':'On Hand Adicional'},axis=1,inplace=True)
    df_rl_raw=df_rl_raw.merge(df_adicionales,how='left',on='Name')
    df_rl_raw['Adicional'].fillna(0,inplace=True)
else:
    df_rl_raw[['Adicional','On Hand Adicional','Aloc Adicional']]=["",0,0]



#Agregar pendiente de recibo
pend_recibo=df_manifest.groupby(['Part No']).sum('Qty')
pend_recibo.reset_index(inplace=True)
pend_recibo.rename({'Part No':'Name','Qty':'pend, recibo'},axis=1,inplace=True)
df_rl_raw=df_rl_raw.merge(pend_recibo,how='left',on='Name')

# Formula columns:
formula_cols=['Req. total','Delta','OH Disp']
for col in formula_cols:
    df_rl_raw[col]=''

path_bom_detail=get_path(file_selectors,'BOM Detail')

if (path_bom_detail) and (path_bom_detail!='Not selected'):
    df_bomdetail=read_excel(path_bom_detail)
    check_mandatory_cols(df_bomdetail.columns,'BOM Detail')
    df_bomdetail.drop('BOM',axis=1,inplace=True)
    df_bomdetail=df_bomdetail.drop_duplicates(['Flat Component'])
    df_rl_raw=df_rl_raw.merge(df_bomdetail,left_on='Name',right_on='Flat Component')
    df_rl_raw.drop('Flat Component',axis=1,inplace=True)
else:
    df_rl_raw['Primary Stock']=''

df_rl_raw.fillna('',inplace=True)
df_ctb = pd.pivot_table(df_rl_raw, 
                       values=['qty'],  
                       index=ctb_cols+formula_cols+['Alterno','ON hand','Primary Stock','Aloc Aterno','Adicional','On Hand Adicional','Aloc Adicional','pend, recibo'],         
                       columns=['year_month','type'],       
                       aggfunc='sum',
                       fill_value=0
                        )  
df_ctb.reset_index(inplace=True)

# Delta comp for columns
cols=df_ctb['qty'].columns
cols=[col[0] for col in cols]
cols=set(cols)

for col in list(cols):
    if not 'Demand' in df_ctb['qty',col].columns:
        df_ctb.insert(df_ctb.columns.get_loc(('qty',col,'SchdRcpt')),('qty',col,'Demand'),0)

for col in list(cols):
    if not 'SchdRcpt' in df_ctb['qty',col].columns:
        df_ctb.insert(df_ctb.columns.get_loc(('qty',col,'Demand'))+1,('qty',col,'SchdRcpt'),0)

for col in list(cols):
    if not 'Delta Comp' in df_ctb['qty',col].columns:
        df_ctb.insert(df_ctb.columns.get_loc(('qty',col,'SchdRcpt'))+1,('qty',col,'Delta Comp'),0)

df_ctb.rename({
    "Name":"Component",
    "Description":"Component Description"
},axis=1,inplace=True)


if len(pivot_bom_demand)>0:
    existing_cols=df_ctb.columns
    if not isinstance(pivot_bom_demand.columns, pd.MultiIndex):
        tuple_columns = [(col, '', '') for col in pivot_bom_demand.columns]
        multiindex_columns = pd.MultiIndex.from_tuples(tuple_columns)
        pivot_bom_demand.columns=multiindex_columns
    pivot_bom_demand['Simulacion']=''
    pivot_bom_demand['CTB']=''
    df_ctb=df_ctb.merge(pivot_bom_demand,how='left',on='Component')
    remaining_columns = [col for col in df_ctb.columns if col not in pivot_bom_demand.columns]
    df_ctb=df_ctb[list(pivot_bom_demand.columns)+list(remaining_columns)]


# Ordeno columnas de Forecast al inicio
family_cols = family_cols.tolist()
req_family_cols=[x for x in family_cols if (('REQ. Total ' in x)&('FC-' not in x))]
for item in reversed(fc_req_cols):
    if item in family_cols:
        family_cols.remove(item)
        family_cols.insert(0, item)
family_cols = pd.Index(family_cols)

ordered_cols=[('Component', '', '')]+\
            [(col, '', '') for col in family_cols]+\
            [
            (           'Req. total','',''),
            (                'Delta','',''),
            (                'QTY PER MAX GRAL','',''),
            (                'CTB','',''),
            (              'Alterno','',''),
            (              'ON hand','',''),
            (          'Aloc Aterno','',''),
            (        'Primary Stock','',''),
            ('Component Description','',''),
            (                 'Note','',''),
            (              'OH Disp','',''),
            (                'Total','',''),
            (                  ' PP','',''),
            (                   'NR','',''),
            (           'Allocation','',''),
            (          'Adicional','',''),
            (          'On Hand Adicional','',''),
            (          'Aloc Adicional','',''),
            (         'pend, recibo','',''),
            (         'Simulacion','','')]+\
            [col for col in df_ctb.columns if col[0]=='qty']

          
df_ctb=df_ctb[ordered_cols]

modified_columns = [(col[0], col[1], col[2]) if 'qty' in col[0] else ('', '', col[0]) for col in ordered_cols]
multiindex_columns = pd.MultiIndex.from_tuples(modified_columns)
df_ctb.columns=multiindex_columns
df_ctb.fillna(0,inplace=True)
append_sheet(df_ctb,path_ctb,'CTB',index=True)


# %% [markdown]
# #### 3.4.1 Formato de CTB

# %% jupyter={"source_hidden": true}
#Formato al CTB
if is_file_open(path_ctb):
    show_popup_message(f"Favor de cerrar un archivo {path_ctb}")
    raise SystemExit()  


wb = load_workbook(path_ctb)
ws = wb['CTB']

init=find_cell_by_text(ws,'Component')
#Reposition headers only if not done already
if ws[init].column!=1:
    merged_cells = list(ws.merged_cells.ranges)  
    for merged_cell in merged_cells:
            ws.unmerge_cells(str(merged_cell))  
    ws.delete_cols(1)
    ws.delete_rows(4)

    for merged_cell in merged_cells:
        min_col, min_row, max_col, max_row = merged_cell.min_col, merged_cell.min_row, merged_cell.max_col, merged_cell.max_row
        if min_col > 1:
            new_range = ws.cell(min_row, min_col - 1).coordinate + ":" + ws.cell(max_row, max_col - 1).coordinate
            ws.merge_cells(new_range)  

#Fonts and fills
for row in ws[ws.calculate_dimension()]:
    for cell in row:
        cell.font=Font(size=8,name='Arial')

demand_start=find_cell_by_text(ws,'Demand')
last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
fill = PatternFill(start_color=light_pink, end_color=light_pink, fill_type="solid")
font = Font(size=8,name='Arial',color=dark_red) 
ws.conditional_formatting.add(f"{demand_start}:{last_cell}",  
    CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=fill, font=font))

fill = PatternFill(start_color=grey, end_color=grey, fill_type="solid")
cell=ws[find_cell_by_text(ws,'Component')]
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)

cell=ws[find_cell_by_text(ws,'Primary Stock')]
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True) 
cell.alignment = Alignment(text_rotation=90,wrap_text=True, horizontal='center', vertical='center')  

cell=ws[find_cell_by_text(ws,'OH Disp')]
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)   
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

cell=ws[find_cell_by_text(ws,' PP')]
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)   
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center') 

cell=ws[find_cell_by_text(ws,'NR')]
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)   
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center') 

cell=ws[find_cell_by_text(ws,'Allocation')]
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)   

fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")

cell=ws[find_cell_by_text(ws,'Component Description')]
cell.alignment = Alignment(text_rotation=90,wrap_text=True, horizontal='center', vertical='center')
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)

cell=ws[find_cell_by_text(ws,'Note')]
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  

cell=ws[find_cell_by_text(ws,'Req. total')]
cell.alignment = Alignment(text_rotation=90,wrap_text=True, horizontal='center', vertical='center')
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)

cell=ws[find_cell_by_text(ws,'Delta')]
cell.alignment = Alignment(text_rotation=90,wrap_text=True, horizontal='center', vertical='center')
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)    


cell=ws[find_cell_by_text(ws,'pend, recibo')]
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)  
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  

cell=ws[find_cell_by_text(ws,'Simulacion')]
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=white, bold=True)  
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')      

fill = PatternFill(start_color=light_green, end_color=light_green, fill_type="solid")
# Alterno puede no existir
cell=ws[find_cell_by_text(ws,'Alterno')]
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
cell.fill=fill
cell=ws[find_cell_by_text(ws,'Aloc Aterno')]
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
cell.fill=fill
cell=ws[find_cell_by_text(ws,'ON hand')]
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
cell.fill=fill
cell=ws[find_cell_by_text(ws,'Total')]
cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
cell.fill=fill

cell=ws[find_cell_by_text(ws,'QTY PER MAX GRAL')]
cell.alignment = Alignment(text_rotation=90,wrap_text=True, horizontal='center', vertical='center')
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=black, bold=False)    

cell=ws[find_cell_by_text(ws,'CTB')]
cell.alignment = Alignment(text_rotation=90,wrap_text=True, horizontal='center', vertical='center')
cell.fill=fill
cell.font=Font(size=8,name='Arial',color=black, bold=False)  

fill = PatternFill(start_color=avocato_green, end_color=avocato_green, fill_type="solid")
cell1=ws[find_cell_by_text(ws,'Demand')]
for row in ws[f"{cell1.offset(-1,0).coordinate}:{ws[ws.calculate_dimension()][cell1.row-1][-1].coordinate}"]:
    for cell in row:
        cell.fill=fill
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

set_number_format(ws,'Delta',format='#,##0.00_);[Red](#,##0.00)')
set_number_format(ws,'CTB',format='#,##0.00_);[Red](#,##0.00)')
set_number_format(ws,'OH Disp',format='#,##0.00_);[Red](#,##0.00)')
font=Font(size=8,name='Arial', bold=True) 
font_column(ws,'Delta',font)
font_column(ws,'OH Disp',font)

fill = PatternFill(start_color=light_yellow, end_color=light_yellow, fill_type="solid")
fill_column(ws,'Total',fill)

fill = PatternFill(start_color=melon, end_color=melon, fill_type="solid")
fill_column(ws,'Req. total',fill)
fill_column(ws,'Delta',fill)
fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
fill_column(ws,'OH Disp',fill)
fill_column(ws,'pend, recibo',fill)

for col in family_cols:
        cell=ws[find_cell_by_text(ws,col)]
        if 'REQ. Total' in col:
            cell.font=Font(size=8,name='Arial',color=black, bold=False)  
            cell.fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")
        elif 'Qty Per' in col:
            cell.font=Font(size=8,name='Arial',color=white, bold=True)  
            cell.fill = PatternFill(start_color=grey, end_color=grey, fill_type="solid")
        cell.alignment=Alignment(text_rotation=90,horizontal='center', vertical='center')

cell=ws[find_cell_by_text(ws,"Adicional")]
cell.fill = PatternFill(start_color=melon, end_color=melon, fill_type="solid")
cell.alignment=Alignment(text_rotation=90,horizontal='center', vertical='center')
cell=ws[find_cell_by_text(ws,"On Hand Adicional")]
cell.fill = PatternFill(start_color=melon, end_color=melon, fill_type="solid")
cell.alignment=Alignment(text_rotation=90,horizontal='center', vertical='center')
cell=ws[find_cell_by_text(ws,"Aloc Adicional")]
cell.fill = PatternFill(start_color=melon, end_color=melon, fill_type="solid")
cell.alignment=Alignment(text_rotation=90,horizontal='center', vertical='center')

# Formulas

# Suma de familias
req_cells=[find_cell_by_text(ws,x) for x in req_family_cols]
col_head_address=ws[find_cell_by_text(ws,'Req. total')].offset(1,0).coordinate
for row in ws[f"{col_head_address}:{last_cell}"]:
    req_cells_str=""
    for cell in req_cells:
        req_cells_str=f"{req_cells_str}{ws.cell(row[0].row,ws[cell].column).coordinate},"
    req_cells_str=req_cells_str[:-1]
    row[0].value=f"=sum({req_cells_str})"

# Formula Delta
cell_a=find_cell_by_text(ws,'OH Disp')
cell_b=find_cell_by_text(ws,'Req. total')


col_head_address=ws[find_cell_by_text(ws,'Delta')].offset(1,0).coordinate
for row in ws[f"{col_head_address}:{last_cell}"]:
    row[0].value=f"={ws.cell(row[0].row,ws[cell_a].column).coordinate}-{ws.cell(row[0].row,ws[cell_b].column).coordinate}"

# Formula CTB
cell_a=find_cell_by_text(ws,'OH Disp')
cell_b=find_cell_by_text(ws,'QTY PER MAX GRAL')
col_head_address=ws[find_cell_by_text(ws,'CTB')].offset(1,0).coordinate
for row in ws[f"{col_head_address}:{last_cell}"]:
    row[0].value=f"={ws.cell(row[0].row,ws[cell_a].column).coordinate}/{ws.cell(row[0].row,ws[cell_b].column).coordinate}"

# Formula OH Dispo
cell_a=find_cell_by_text(ws,'Total')
cell_b=find_cell_by_text(ws,'Allocation')
cell_c=find_cell_by_text(ws,'pend, recibo')
cell_d=find_cell_by_text(ws,'Simulacion')
cell_e=find_cell_by_text(ws,'Adicional')
cell_f=find_cell_by_text(ws,'On Hand Adicional')
cell_g=find_cell_by_text(ws,'Aloc Adicional')
col_head_address=ws[find_cell_by_text(ws,'OH Disp')].offset(1,0).coordinate
for row in ws[f"{col_head_address}:{last_cell}"]:
    row[0].value=f"={ws.cell(row[0].row,ws[cell_a].column).coordinate}-{ws.cell(row[0].row,ws[cell_b].column).coordinate}+{ws.cell(row[0].row,ws[cell_c].column).coordinate}+{ws.cell(row[0].row,ws[cell_d].column).coordinate}+{ws.cell(row[0].row,ws[cell_e].column).coordinate}+{ws.cell(row[0].row,ws[cell_f].column).coordinate}-{ws.cell(row[0].row,ws[cell_g].column).coordinate}"

# Formula for Delta Comp first column
cell_a=find_cell_by_text(ws,'Total')
cell_b=find_cell_by_text(ws,'Demand')
cell_c=find_cell_by_text(ws,'SchdRcpt')
col_head_address=ws[find_cell_by_text(ws,'Delta Comp')].offset(1,0).coordinate
for row in ws[f"{col_head_address}:{last_cell}"]:
    row[0].value=f"={ws.cell(row[0].row,ws[cell_a].column).coordinate}+{ws.cell(row[0].row,ws[cell_c].column).coordinate}-{ws.cell(row[0].row,ws[cell_b].column).coordinate}"

# Formula for Delta Comp
cells_b=find_all_cells_by_text(ws,'Demand')
cells_c=find_all_cells_by_text(ws,'SchdRcpt')
cells_d=find_all_cells_by_text(ws,'Delta Comp')
df=pd.DataFrame({'Demand':cells_b,
                'SchdRcpt':cells_c,
                'DeltaComp':cells_d})

df['PrevDeltaComp']=df['DeltaComp'].shift(1)
# The first formula is different so we drop the first row
df.drop(0,inplace=True)
for index,dfrow in df.iterrows():
    for row in ws[f"{ws[dfrow['DeltaComp']].offset(1,0).coordinate}:{last_cell}"]:
        row[0].value=f"={ws.cell(row[0].row,ws[dfrow['PrevDeltaComp']].column).coordinate}+{ws.cell(row[0].row,ws[dfrow['SchdRcpt']].column).coordinate}-{ws.cell(row[0].row,ws[dfrow['Demand']].column).coordinate}"
ws.freeze_panes = 'B4'
if col_sizes:
    apply_col_sizes(wb,col_sizes)
save_wb(wb=wb,filepath=path_ctb)
wb.close()
try:
    os.startfile(path_ctb)
except:
    show_popup_message(f"El archivo CTB esta listo.")


# %%
# Maximo de cortos permitidos
display(HTML('''<style>
    .widget-label { min-width: 40ex !important; }
</style>'''))
max_allowed_shorts = widgets.IntSlider(
    value=3,  # Cantidad por default de cortos maximos aceptables
    min=0,    
    max=10,   
    step=1,   
    description='Cantidad maxima de cortos permitidos:',  # Description of the slider
    orientation='horizontal',  # Orientation of the slider
    readout=True,  # Display the current value
    layout=Layout(width='500px')
)
max_allowed_shorts


# %% [markdown]
# # Sugerencia de lanzamientos
# - Guardar el archivo CTB antes de correr esta seccion, incluso si no se han realizado cambios en el archivo con el fin de que las formulas se evaluen
# - Toma los archivos de excel: Analisis de lanzamiento y CTB KRS
#     - Determina el requerimiento con menos cortos
#     - Lo asigna como primer lanzamiento y resta el BOM de los componentes disponibles (OH Dispo del CTB)
#     - Evalua el siguiente requerimiento con menos cortos y realiza el proceso hasta que asigne todos los lanzamientos
#     - Genera un reporte de cortos por requerimiento, uno de cortos globales y la sugerencia de lanzamientos
#     - La columna Cortos aparece en blanco si el requerimiento esta completo o un link al reporte de cortos
# - POs con cortos criticos o con maximo de cortos permitidos no se toman en cuenta, no se descuentandel material disponible

# %%
# Sugerencia
close_xl_if_open(path_launch)
verify_selections(file_selectors)
if path_demanda_tablillas=='Not selected':
    path_ctb=os.path.join(folder_output_label.value,'CTB KRS.xlsx')
else:
    path_ctb=os.path.join(folder_output_label.value,'CTB KRS_tablillas.xlsx')
close_xl_if_open(path_ctb)

# Definir componentes criticos por cliente
critical_components={"TROV1.0":["APC-"],
                     "RISE":["PCB-"]}

def substract_bom(df_demand, df_mat):
    """
    Subtracts the Bill of Materials (BOM) demand from available materials and calculates completion percentages and shortages.

    Parameters:
    df_demand (DataFrame): DataFrame containing demand data with at least 'component' and 'QTY' columns.
    df_mat (DataFrame): DataFrame containing material availability data with at least 'component' and 'AVAIL' columns.

    Returns:
    df_demand_new (DataFrame): Updated demand DataFrame with additional columns 'completion', 'new_avail', 'short'.
    df_mat_new (DataFrame): Updated material availability DataFrame with adjusted 'AVAIL' after subtracting demand.
    """
    df_demand_new=df_demand.copy()
    df_mat_new=df_mat.copy()
    df_mat_new['OH Disp']=df_mat_new['OH Disp'].clip(lower=0) # Ignoro disponibilidad negativa
    df_demand_new = df_demand_new.merge(df_mat_new[['Component', 'OH Disp']], how='left', on='Component')
    df_demand_new['OH Disp'].fillna(0,inplace=True)
    df_demand_new['completion'] = (df_demand_new['OH Disp'] / df_demand_new['REQ. Total']).clip(upper=1)
    df_demand_new['new_avail'] = (df_demand_new['OH Disp'] - df_demand_new['REQ. Total']).clip(lower=0)
    df_demand_new['short'] = (df_demand_new['OH Disp'] - df_demand_new['REQ. Total']).clip(upper=0)
    df_demand_total = df_demand_new.groupby('Component')['REQ. Total'].sum().reset_index()
    df_mat_new = df_mat_new.merge(df_demand_total, how='left', on='Component')
    df_mat_new['REQ. Total'] = df_mat_new['REQ. Total'].fillna(0)
    df_mat_new['OH Disp'] = (df_mat_new['OH Disp'] - df_mat_new['REQ. Total']).clip(lower=0)
    df_mat_new = df_mat_new.drop(columns=['REQ. Total'])
    df_demand_new.fillna(0,inplace=True)
    df_demand_new.reset_index(drop=True,inplace=True)
    df_mat_new.fillna(0,inplace=True)
    critical_short=False
    for pfam in critical_components.keys():
        df=df_demand_new[(df_demand_new['Component'].str.contains('|'.join(critical_components[pfam])))&
                        (df_demand_new['P Family']==pfam)&
                        (df_demand_new['short']<0)]
        if len(df)>0:
            critical_short=True
            break
    # Revisar cantidad maxima de cortos
    max_shorts=False
    if len(df_demand_new[df_demand_new['short']<0])>max_allowed_shorts.value:
        max_shorts=True
    if critical_short | max_shorts:
        df_mat_new=df_mat # Si hay un corto critico o mas cortos de los permitidos, no se hacen cambios en el material disponible     
    return dict(df_demand_new=df_demand_new,
                df_mat_new=df_mat_new,
                critical_short=critical_short,
                max_shorts=max_shorts)

# Cargar datos de excel incluyendo modificaciones manuales
wb=load_workbook(path_ctb, data_only=True,read_only=True)
ws=wb['CTB']
data = ws.values
for _ in range(2):  # Skip the first two rows
    next(data)
columns = next(data)
df_ctb = pd.DataFrame(data, columns=columns)
if df_ctb[df_ctb['OH Disp'].isnull()].shape[0]>0:
    show_popup_message(f"Favor de Guardar el archivo {path_ctb}")
    wb.close()
    raise SystemExit()
wb.close()
# Actualizar el bom para lo pendiente de lanzar
wb_demand_launch=pd.read_excel(path_launch,sheet_name=None)
df_demand_launch=wb_demand_launch['Analisis de lanzamiento']
df_demand_launch.sort_values(['Alta Prioridad','Due Date','REQ'],inplace=True)
df_demand_launch['Cortos(link)']=''
df_demand_launch['Estatus']=''
df_bom=get_bom()
df_demand_launch_rdy=df_demand_launch[(df_demand_launch['QTY Pend. Lanzar']>0)&(df_demand_launch['REQ']>0)&(df_demand_launch['Omitir'].isnull())]
df_bom_demand=df_demand_launch_rdy.merge(df_bom,how='left',on=['MODELO'])
df_missing=df_bom_demand[df_bom_demand['Component'].isnull()]
df_bom_demand=df_bom_demand[~df_bom_demand['Component'].isnull()]
df_bom_demand['REQ. Total']=df_bom_demand['QTY Pend. Lanzar']*df_bom_demand['Qty Per']
df_bom_demand=df_bom_demand[['PO','MODELO','P Family','QTY Pend. Lanzar','Component','Qty Per','REQ. Total']]

# Componentes disponibles segun el CTB
df_available=df_ctb[['Component','OH Disp']]

alloc_priority=0
completion_left=1
max_req=0
old_req=0

df_available_left=df_available.copy()
df_short_detail=pd.DataFrame()
df_short_resume=df_bom_demand.groupby('Component').sum('REQ. Total')[['REQ. Total']]
df_short_resume.reset_index(inplace=True)

df_short_resume=df_short_resume.merge(df_available,how='left',on='Component')
df_short_resume['short']=df_short_resume['OH Disp']-df_short_resume['REQ. Total']
df_short_resume=df_short_resume[df_short_resume['short']<0]


for idx_po,po_row in df_demand_launch_rdy.iterrows():
    idx=(df_bom_demand['PO']==po_row['PO']) & (df_bom_demand['MODELO']==po_row['MODELO'])
    result=substract_bom(df_bom_demand[idx],df_available_left)
    df=result['df_demand_new']
    df_available_left=result['df_mat_new']
    df=df[df['short']<0]
    df_short_detail=pd.concat([df,df_short_detail])
    if result['critical_short']:
        df_demand_launch.loc[idx_po,'Estatus']='Corto Critico, no lanzar'
        continue
    if result['max_shorts']:
        df_demand_launch.loc[idx_po,'Estatus']='Maximo de cortos alcanzado, no lanzar'
        continue    
    if len(df)==0:
        df_demand_launch.loc[idx_po,'Estatus']='Listo'
    else:
        df_demand_launch.loc[idx_po,'Estatus']='Cortos'

df_demand_launch.loc[df_demand_launch['QTY Pend. Lanzar']<=0,'Estatus']='Lanzado'

df_short_detail.reset_index(inplace=True,drop=True)
df_short_detail.reset_index(inplace=True)
df_indexes=df_short_detail.drop_duplicates(['PO','MODELO'],keep='first')[['PO','MODELO','index']] 
df_demand_launch=df_demand_launch.merge(df_indexes,how='left',on=['PO','MODELO'])
df_short_detail.drop('index',axis=1,inplace=True)
idx=~df_demand_launch['index'].isnull()
df_demand_launch.loc[idx,'index']="=HYPERLINK(\"#'Cortos-Detalle'!A"+(df_demand_launch.loc[idx,'index']+2).astype(int).astype(str)+"\",'Cortos-Detalle'!E"+(df_demand_launch.loc[idx,'index']+2).astype(int).astype(str)+")"
if 'Cortos(link)' in df_demand_launch.columns:
    df_demand_launch['Cortos(link)']=df_demand_launch['index']
    df_demand_launch.drop(['index'],axis=1,inplace=True)
else:
    df_demand_launch.rename({'index':'Cortos(link)'},axis=1,inplace=True)
#------------------------------------------------------
# Agregar llegada de material
path_arrivals=get_path(file_selectors,'PO WO Info')
if (path_arrivals) and (path_arrivals!='Not selected'):
    df_arrivals=pd.read_excel(path_arrivals,sheet_name="POWO Info Update",header=1)
else:
    df_arrivals=pd.DataFrame(columns=mandatory_cols['PO WO Info'])
check_mandatory_cols(df_arrivals.columns,'PO WO Info')
df_arrivals=df_arrivals[mandatory_cols['PO WO Info']]
if len(df_arrivals)>0:
    df_arrivals=df_arrivals.groupby(['Part Name','Due']).sum('Quantity')[['Quantity']]
    df_arrivals.reset_index(inplace=True)
    # Fecha de cobertura

    df_arrivals.rename({'Part Name':'Component'},axis=1,inplace=True)
    df_arrivals['Cumulative Sum'] = df_arrivals.groupby(['Component'])['Quantity'].cumsum()
    df_coverage = df_short_resume.merge(df_arrivals, on='Component', how='left')
    df_coverage['short']=-df_coverage['short']
    

    def filter_first_exceeding(group):
        mask = group['Cumulative Sum'] > group['short'].iloc[0]
        return group[mask].head(1) 
    df_coverage = df_coverage.groupby('Component').apply(filter_first_exceeding).reset_index(drop=True)
    df_coverage=df_coverage[['Component','Due']]
    df_coverage.rename({'Due':'Cobertura'},axis=1, inplace=True)
    try:
        df_arrivals['Due']=df_arrivals['Due'].dt.date.astype(str)
    except:
        show_popup_message(f"Favor de revisar el archivo: {path_arrivals}")
    
    pivot_arrivals = pd.pivot_table(df_arrivals, 
                        values=['Quantity'],  
                        index=['Component'],         
                        columns=['Due'],       
                        aggfunc='sum',
                        fill_value=0
                            ) 
    pivot_arrivals.columns=[col[1] for col in pivot_arrivals.columns.values]
    pivot_arrivals.reset_index(inplace=True)
    df_short_resume=df_short_resume.merge(df_coverage,how='left',on='Component')
    df_short_resume['Cobertura'].fillna('No disponible',inplace=True)
    df_short_resume['Qty']=''
    df_short_resume=df_short_resume.merge(pivot_arrivals,how='left',on=['Component']).fillna(0)
    df_short_resume['Cobertura'] = df_short_resume['Cobertura'].astype(str).str.replace(' 00:00:00', '')
df_demand_launch['Llave']=df_demand_launch['MODELO']+df_demand_launch['PO']
col = df_demand_launch.pop('Llave')
df_demand_launch.insert(0, 'Llave', col)
wb_demand_launch['Analisis de lanzamiento']=df_demand_launch
wb_demand_launch['Cortos-Detalle']=df_short_detail
wb_demand_launch['Cortos-Resumen']=df_short_resume
if len(df_missing)>0:
    wb_demand_launch['BOMS Faltantes']=df_missing
col_sizes=[]
if os.path.exists(path_launch):
    wb=load_workbook(path_launch)
    col_sizes=get_col_sizes(wb)
save_df_multiple(df_dict=wb_demand_launch,filepath=path_launch)

# %% [markdown]
# ## Formato sugerencia de lanzamientos

# %%
# Formato
wb = load_workbook(path_launch)
ws = wb['Cortos-Resumen']
ws.freeze_panes = 'E2'
ws.auto_filter.ref = f"{ws[ws.calculate_dimension()][0][0].coordinate}:{ws[ws.calculate_dimension()][0][-1].coordinate}"
last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
cell_cober=find_cell_by_text(ws,'Cobertura')
if cell_cober:
    col_head_address=ws[cell_cober].offset(1,0).coordinate
    for row in ws[f"{col_head_address}:{last_cell}"]: 
        # row[0].value=f"=IFERROR(INDEX({ws.cell(1,ws[col_head_address].column+2).coordinate}:{ws.cell(1,ws[last_cell].column).coordinate},MATCH(TRUE,SCAN(0,INDEX({ws.cell(row[0].row,ws[col_head_address].column+2).coordinate}:{ws.cell(row[0].row,ws[last_cell].column).coordinate},0,0),LAMBDA(a,x,IF(a>=-{ws.cell(row[0].row,ws[col_head_address].column-1).coordinate},FALSE,a+x)))>=-{ws.cell(row[0].row,ws[col_head_address].column-1).coordinate},0)),\"No disponible\")"
        row[1].value=f"=IF({ws.cell(row[0].row,ws[col_head_address].column).coordinate}=\"No disponible\", 0, SUM({ws.cell(row[0].row,ws[col_head_address].column+2).coordinate}:INDEX({ws.cell(row[0].row,ws[col_head_address].column+2).coordinate}:{ws.cell(row[0].row,ws[last_cell].column).coordinate}, MATCH({ws.cell(row[0].row,ws[col_head_address].column).coordinate}, {ws.cell(1,ws[col_head_address].column+2).coordinate}:{ws.cell(1,ws[last_cell].column).coordinate}, 0))))"
ws = wb['Cortos-Detalle']
ws.freeze_panes = 'B2'

ws = wb['Analisis de lanzamiento']

ws.auto_filter.ref = f"{ws[ws.calculate_dimension()][0][0].coordinate}:{ws[ws.calculate_dimension()][0][-1].coordinate}"
ws.freeze_panes = 'A2'
if col_sizes:
    apply_col_sizes(wb,col_sizes)
save_wb(wb=wb,filepath=path_launch)

wb.close()

os.startfile(path_launch)
os.startfile(path_ctb)


# %% [markdown]
# ## Gating parts Report

# %%
# Reporte Gating Parts
path_launch=os.path.join(folder_output_label.value,'Analisis_lanzamiento.xlsx')
close_xl_if_open(path_launch)
dict_launch=read_excel(path=path_launch,sheet_name=None)
df_launch=dict_launch['Analisis de lanzamiento']
df_shorts=dict_launch['Cortos-Detalle']

df_launch_ready=df_launch[df_launch['Estatus']=='Listo']
# Si la decision es no tomar en cuenta cortos criticos, cambiar por igual a Corto
df_launch_short=df_launch[(df_launch['Estatus']=='Cortos')]
df_launch_short=df_launch_short.merge(df_shorts,how='left',on=['PO','MODELO'])
df_launch_short=df_launch_short[['PO','MODELO','Component','REQ','short','Estatus']]
df_launch_short['Cumulative Sum']=-df_launch_short.groupby('Component')['short'].cumsum()
df_launch_short['Cumulative Sum']=df_launch_short['Cumulative Sum'].astype(int)
df_arrivals['Cumulative Sum']=df_arrivals['Cumulative Sum'].astype(int)
df_cover = pd.merge_asof(
    df_launch_short.sort_values('Cumulative Sum'),
    df_arrivals.sort_values('Cumulative Sum'),
    left_on='Cumulative Sum',
    right_on='Cumulative Sum',
    by='Component',
    direction='forward'
)
df_cover=df_cover[['PO','MODELO','Component','REQ','short','Estatus','Cumulative Sum','Due','Quantity']]
df_launch_ready=df_launch_ready[['PO','MODELO','REQ']]
dict_gating={'Shorts':df_cover,
             'Ready':df_launch_ready,
             'Arrivals':df_arrivals}
path_gating=os.path.join(folder_output_label.value,'Gating Parts Report.xlsx')
save_df_multiple(dict_gating,path_gating)
