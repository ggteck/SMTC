#%%
import pandas as pd
import os
from copy import copy, deepcopy
from importlib import reload
from manufacturing_plan import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pandas.tseries.offsets import BDay
state = load_state_pickle()

folder_output = state['folder_output']
file_selectors = state['selections']
output_paths = set_paths(folder_output)
col_rel = set_col_rel(output_paths)
df_col_rel = col_rel['col_rel']
df_columns=col_rel['columns']
#%%

if not os.path.exists(output_paths['path_xl_format']):
    print("No se encuentra el archivo: columns and formatting.xlsx")
    raise SystemExit()

path_available_hours=file_selectors['available_hours']
df_avail_hours=read_excel(path=path_available_hours)
df_avail_hours['dia']=pd.to_datetime(df_avail_hours['dia'],errors='coerce').dt.strftime('%Y-%m-%d')
df_avail_hours['dia'].fillna('default',inplace=True)
dict_avail_hours=df_avail_hours.set_index(['dia','maquina']).to_dict(orient='index')

#%
path_master_doblado = file_selectors['master_file']
df_master_doblado = load_excel_with_header_key(path_master_doblado, sheet_name='00. Formato para Master de WC', key_text='PN')
df_master_doblado = rename_columns(df_master_doblado, df_col_rel, table_from='Master Doblado', sheet_from='00. Formato para Master de WC')
df_master_doblado.replace('/','_',regex=True,inplace=True)
df_master_doblado['pn']=df_master_doblado['pn'].astype(str)
#%
path_routing = file_selectors['routing_file']
df_routing = load_excel_with_header_key(path_routing, sheet_name='Operations', key_text='Routing')
df_routing = rename_columns(df_routing, df_col_rel, table_from='Routing', sheet_from='Operations')
df_routing['operation_description']=df_routing['operation_description'].str.upper()
# Equivalencias
path_equiv = st.session_state.selected_paths['equivalencias_file']
df_equiv=read_predefined_excel(path_equiv,df_columns,table='Equivalencias')
dict_equiv=df_equiv.set_index('pn')['equivalencia'].to_dict()
dict_equiv_inv={v: k for k, v in dict_equiv.items()}
#% Open order list
path_order_list = file_selectors['order_file']
df_order_list = load_excel_with_header_key(path_order_list, key_text='Priority')
df_order_list = rename_columns(df_order_list, df_col_rel, table_from='Lista de ordenes')
df_order_list['operation_description']=df_order_list['operation_description'].str.upper()
df_order_list['pn_orig']=df_order_list['pn']
df_order_list['pn']=df_order_list['pn'].replace(dict_equiv)
check_mandatory_columns_df(df_order_list.columns,df_columns=df_columns,table='Lista de ordenes')

df_missing_rout=get_common_records(df_order_list,df_routing,keys=['pn','operation_description'],how='uncommon')
df_order_list['composite_key'] = list(zip(*(df_order_list[col] for col in ['pn','operation_description'])))
df_routing['composite_key'] = list(zip(*(df_routing[col] for col in ['pn','operation_description'])))
df_missing_rout=df_order_list[~df_order_list['composite_key'].isin(df_routing['composite_key'])]
df_missing_rout['composite_key'] = list(zip(*(df_missing_rout[col] for col in ['pn_orig','operation_description'])))
df_missing_rout=df_missing_rout[~df_missing_rout['composite_key'].isin(df_routing['composite_key'])]
df_missing_machine=df_order_list[(~df_order_list['pn'].isin(df_master_doblado['pn']))&
                                 (~df_order_list['pn_orig'].isin(df_master_doblado['pn']))&
                                 (df_order_list['machine']=='')]
df_order_list.drop(columns=['composite_key'],inplace=True)

#% Open part master
path_part_master = file_selectors['part_master']
df_part_master = load_excel_with_header_key(path_part_master, key_text='Site')
df_part_master = rename_columns(df_part_master, df_col_rel, table_from='Part Master')
check_mandatory_columns_df(df_part_master.columns,df_columns=df_columns,table='Part Master')
df_part_master=df_part_master[['pn','unit_selling_price']].drop_duplicates(subset=['pn'],keep='last')
df_part_master_orig=df_part_master.copy()
df_part_master['pn']=df_part_master['pn'].replace(dict_equiv_inv)
df_part_master=pd.concat([df_part_master_orig,df_part_master])
df_part_master=df_part_master.drop_duplicates(subset=['pn'],keep='last')
df_part_master['pn'].replace(dict_equiv)

# df_part_master=get_common_records(df_part_master,df_order_list,keys=['pn'])
#% Open old plan
path_plan=output_paths['path_plan']
df_plan_old=read_predefined_excel(path_plan,df_columns=df_columns,table='Manufacturing plan',check_mandatory=True)
df_plan_old=rename_columns(df_plan_old,df_col_rel=df_col_rel,table_from='Manufacturing plan')
df_plan_old=df_plan_old[~df_plan_old['status'].isna()]

#% Show orders already planned
df_already_planned=get_common_records(df_new=df_order_list,df_old=df_plan_old,keys=['wo','pn'])
st.info("Las siguientes ordenes ya estan planeadas, continue si desea agregarlas al nuevo plan")
st.info(df_already_planned)

#%%
# helper to iterate business‐day calendar
def bdate_range_str(start, end):
    return [d.strftime('%Y-%m-%d') 
            for d in pd.bdate_range(start=start, end=end)]
initial_date_str = pd.to_datetime(state['initial_date']).strftime('%Y-%m-%d')
limit_date_ts    = pd.to_datetime(state['limit_date'])
all_dates = bdate_range_str(initial_date_str, limit_date_ts.strftime('%Y-%m-%d'))
machines = set(m for _, m in dict_avail_hours.keys())
slots = {}
for avail_dt in all_dates:
    for m in machines:
        dict_avail_hours_copy=deepcopy(dict_avail_hours)
        slots[(avail_dt, m)]=dict_avail_hours_copy.get((avail_dt, m),dict_avail_hours_copy.get(('default',m)))

# 2) Prepare
assignments    = []
pn_to_machine  = {}    # first‐machine lock per PN
wo_next_start  = {}    # earliest allowed timestamp per WO
wo_last_machine = {}    # last machine used per WO
machine_status = {}
df_order_list.sort_values('priority', inplace=True)

# prebuild full set of dates we may need
# 3) Process each order
for _, order in df_order_list.iterrows():
    pn_orig       = order['pn_orig']
    pn            = order['pn']
    qty           = order['pzas_x_hacer']
    wo            = order['wo']
    routing_name  = order['operation_description']
    machine_def   = order['machine']
    pty           = order['priority']

    # routing info
    pn_info = df_routing[
        (df_routing['pn']==pn) & 
        (df_routing['operation_description']==routing_name)
    ]
    if pn_info.empty:
        pn_info = df_routing[
            (df_routing['pn']==pn_orig) & 
            (df_routing['operation_description']==routing_name)
        ]
        if pn_info.empty:
            continue
    run_time, setup_time = pn_info.iloc[0][['run_time','setup_time']]

    # select machine
    if machine_def:
        m_list = [machine_def]
    elif pn in pn_to_machine:
        m_list = [pn_to_machine[pn]]
    else:
        rows = df_master_doblado[df_master_doblado['pn']==pn]
        if rows.empty:
            rows = df_master_doblado[df_master_doblado['pn']==pn_orig]
            if rows.empty:
                continue
        m_list = [v for k,v in rows.iloc[0].items() if 'maq_opc' in k and v]

    # determine earliest start
    initial_start={"finish_ts":pd.to_datetime(initial_date_str)}
    start_ts = wo_next_start.get(wo, initial_start)
    start_ts = start_ts.get("finish_ts")
    start_shift = wo_next_start.get(wo, initial_start)
    start_shift = start_shift.get("next_shift","PRIMER TURNO")
    # start_ts = start_ts.get("finish_ts")
    last_m   = wo_last_machine.get(wo)
    
    # assignment loop
    for m in m_list:
        # lock PN → machine on first assignment
        if pn not in pn_to_machine:
            pn_to_machine[pn] = m

        # walk through dates until qty is 0 or we pass limit
        for date_str in all_dates:
            if qty <= 0:
                break
            date_ts = pd.to_datetime(date_str)
            # if date_ts < start_ts.normalize():
            #     continue
            if last_m is None or m != last_m:
                if date_ts < start_ts.normalize():
                    continue                
            if date_ts > limit_date_ts:
                break

            # get or init that day’s shifts
            key = (date_str, m)
            if key in slots:
                avail_shifts = slots[key]


            # iterate shifts in sorted order
            for shift in sorted(avail_shifts):
                if qty <= 0:
                    break

                # compute setup only if PN changed vs last service on this machine
                last_pn = machine_status.get((m,'last_pn'), None)
                need_setup = setup_time if last_pn != pn else 0
                # can we fit setup+one run?
                if avail_shifts[shift] < need_setup + run_time:
                    avail_shifts[shift] = 0
                    continue
                shifts_list=list(avail_shifts.keys())
                if (date_ts == start_ts.normalize())&\
                    (shifts_list.index(start_shift)>shifts_list.index(shift))&\
                    (last_m is None or m != last_m):
                    continue
                # assign as many as fit this slot
                effective = avail_shifts[shift] - need_setup
                pieces   = qty if run_time==0 else min(qty, effective // run_time)
                if pieces<10:
                    continue
                run_used = pieces * run_time
                total    = run_used + need_setup

                # record it
                assignments.append({
                    'date': date_str,
                    'machine': m,
                    'shift': shift,
                    'operation_description': routing_name,
                    'pn': pn_orig,
                    'wo': wo,
                    'priority': pty,
                    'pzas_x_hacer': pieces,
                    'time_used': run_used,
                    'setup_time': need_setup
                })

                # update slot availability
                avail_shifts[shift] -= total
                qty -= pieces
                # track last_pn for setup logic
                machine_status[(m,'last_pn')] = pn
                # compute finish timestamp to enforce sequencing
                finish_ts = date_ts  # + shift end offset if you track that
                # bump next‐start only when changing machines
                if last_m is None or m != last_m:
                    shifts_list=list(avail_shifts.keys())
                    shidx=shifts_list.index(shift)
                    if len(shifts_list)>shidx+1:
                        next_shift=shifts_list[shidx+1]
                    else:
                        finish_ts=finish_ts + BDay(1)
                        next_shift=shifts_list[0]                
                    next_start={"finish_ts":finish_ts,
                                "next_shift":next_shift}
                    wo_next_start[wo] = next_start
                # record this machine as last used for the WO
                wo_last_machine[wo] = m

            # end for shift
        # end for date

        break  # once we’ve tried this machine, stop (PN locked)


df_plan_new=get_predefined_df(df_columns=df_columns,table='Manufacturing plan')
df_plan_new = pd.concat([df_plan_new,pd.DataFrame(assignments)],ignore_index=True)
df_plan_new.sort_values(['date', 'machine', 'shift', 'priority', 'wo'], inplace=True)
df_plan_new['pzas_x_hora']=(df_plan_new['pzas_x_hacer']/df_plan_new['time_used']).astype(float).round(2)
df_plan_old=df_plan_old[~df_plan_old['status'].isnull()]
df_plan_old=append_df_to_df(df_new=df_plan_new,df_old=df_plan_old,table='Manufacturing plan',keys=['wo','pn'],allow_duplicates=True)

if 'unit_selling_price' in df_plan_old.columns:
    df_plan_old.drop(columns=['unit_selling_price'],inplace=True)
df_plan_old=df_plan_old.merge(df_part_master,on='pn',how='left')    
df_plan_old['unit_selling_price']=df_plan_old['unit_selling_price']*df_plan_old['pzas_x_hacer']
plan_cols=df_columns[(df_columns['table']=='Manufacturing plan')&
           (~df_columns['mandatory_column'].isna())]['std_name'].to_list()
df_plan_old=df_plan_old[plan_cols]
df_plan_old['date']=pd.to_datetime(df_plan_old['date'],format='mixed').dt.strftime('%B, %#d, %Y')
df_plan_old=rename_columns(df_plan_old,df_col_rel=df_col_rel,table_to='Manufacturing plan')
df_plan_old.to_excel(path_plan,index=False)
os.startfile(path_plan)
#%%
df_plan_old=rename_columns(df_plan_old,df_col_rel=st.session_state.df_col_rel,table_from='Manufacturing plan')    
#% Reportes
machines=df_plan_old['machine'].drop_duplicates().tolist()
group_cols=['date']
plan_report_cols=df_columns[(df_columns['table']=='Machine Report')&
    (~df_columns['mandatory_column'].isna())]['std_name'].to_list()
df_plan_old=df_plan_old[plan_report_cols]
machine_col=df_columns[(df_columns['table']=='Machine Report')&
    (df_columns['std_name']=='machine')]['column_name'].to_list()[0]

df_columns[(df_columns['table']=='Machine Report')]

dict_formats=get_xl_formatting()
special_formats=dict_formats['special_format']
col_sizes=None
for file in os.listdir(folder_output):
    if 'reporte de manufactura' in file:
        wb=load_workbook(os.path.join(folder_output,file))
        col_sizes=get_col_sizes(wb)
        break

for machine in machines:
    df=df_plan_old[df_plan_old['machine']==machine].copy()
    if len(df)==0:
        continue
    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'

    # Write headers
    titles = list(df.columns)
    for idx, title in enumerate(titles, start=1):
        ws.cell(row=1, column=idx, value=title)

    current_row = 2
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

    # Identify numeric columns for subtotals
    numeric_cols = ['pzas_x_hacer','time_used']

    groups = df.groupby(group_cols)
    for date, group in sorted(groups, key=lambda x: pd.to_datetime(x[0])):
        start_row = current_row
        # Data rows
        for _, row in group.iterrows():
            for col_idx, col in enumerate(titles, start=1):
                if col not in group_cols:
                    cell=ws.cell(row=current_row, column=col_idx, value=row[col])
                    cell.value=row[col]
                    if row[col] in special_formats.keys():
                        format=special_formats[row[col]]
                        format_cell(cell,format)
            current_row += 1
        end_data_row = current_row - 1

        # Subtotal row
        for col_idx, col in enumerate(titles, start=1):
            if col in numeric_cols:
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}{start_row}:{col_letter}{end_data_row})"
                ws.cell(row=current_row, column=col_idx, value=formula)
        current_row += 1
    
        # Blank line with black fill across data range
        for col_idx in range(1, len(titles) + 1):
            ws.cell(row=current_row, column=col_idx).fill = black_fill
        current_row += 1

        # Merge date cells and rotate text
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_data_row+1, end_column=1)
        date_cell = ws.cell(row=start_row, column=1, value=date[0])
        date_cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
        date_cell.number_format = 'mmmm, d, yyyy'
    if col_sizes:
        wb=apply_col_sizes(wb,col_sizes)
    base, ext = os.path.splitext(output_paths['path_report'])
    wb.save(f"{base} {machine}{ext}")
# %%

#%% Validar plan
close_xl_if_open(path_order_list)
close_xl_if_open(path_plan)
df_plan=read_excel(path_plan)
df_plan=rename_columns(df_plan,df_col_rel=df_col_rel,table_from='Manufacturing plan')
df_plan_grp=df_plan.groupby(['pn','wo']).sum(['pzas_x_hacer'])[['pzas_x_hacer']]
df_plan_grp.reset_index(inplace=True)
df_order_assignment=df_order_list.merge(df_plan_grp,how='left',on=['pn','wo'],suffixes=('','_assigned'))
df_order_assignment['pzas_x_asignar']=df_order_assignment['pzas_x_hacer']-df_order_assignment['pzas_x_hacer_assigned']
df_order_assignment.loc[df_order_assignment['pzas_x_asignar']==0,'status']='Planeada'
df_order_assignment.loc[df_order_assignment['pzas_x_asignar']>0,'status']='Incompleta'
df_order_assignment.loc[df_order_assignment['pzas_x_asignar'].isna(),'status']='Pendiente'
df_order_assignment=df_order_assignment.drop(columns=['pzas_x_hacer_assigned','pzas_x_asignar'])
orders_cols=df_columns[(df_columns['table']=='Lista de ordenes')]['std_name'].to_list()
df_order_assignment=df_order_assignment[orders_cols]
df_order_assignment=rename_columns(df_order_assignment,df_col_rel,table_to='Lista de ordenes')
df_order_assignment.to_excel(path_order_list,index=False)
df_plan.loc[df_plan['status'].isna(),'status']='Planeada'
df_plan.to_excel(path_plan,sheet_name='Manufacturing plan',index=False)

#%%
# %%
# path_isar=file_selectors['isar']
# df_isar=read_excel(path=path_isar)
# df_isar=rename_columns(df_isar,df_col_rel=df_col_rel,table_from='ISAR')
# df_isar['wo'].fillna('',inplace=True)
# df_isar[(df_isar.duplicated(['wo','pn'],keep=False))&(df_isar['wo']!='')&(df_isar['wo']!='Planned')]
#%% 16 wk
path_16_wk=file_selectors['16_wk']
df_16_wk=read_excel(path=path_16_wk)
df_16_wk=rename_columns(df_16_wk,df_col_rel=df_col_rel,table_from='16 WK')
df_16_wk['pn']=df_16_wk['pn'].astype(str)
df_16_wk=df_16_wk[~df_16_wk['site'].str.contains('SVT')]
df_16_wk=df_16_wk[df_16_wk['wk5']<0]
df_16_wk.sort_values(['wk5'],inplace=True)
#%%
df_16_wk['site'].drop_duplicates()
#%%
file_selectors
#%%
# Sales
path_sales=file_selectors['sales']
df_sales=read_excel(path_sales)
df_sales=rename_columns(df_sales,df_col_rel,table_from="Top ventas")
df_sales['pn']=df_sales['pn'].astype(str)
df_sales = (
    df_sales
    .dropna(subset=['pn'])
    .sort_values('tot_value', ascending=False)
)
df_sales['cumperc'] = df_sales['tot_value'].cumsum() / df_sales['tot_value'].sum()
cut_idx = df_sales['cumperc'].gt(0.8).idxmax()
df_sales = df_sales.loc[:cut_idx]

path_end_of_period=file_selectors['end_of_period']
df_end_of_period=load_excel_with_header_key(file_path=path_end_of_period,key_text='Report Date')
df_end_of_period=rename_columns(df_end_of_period,df_col_rel=df_col_rel,table_from='End of Period')
#%%
df_end_of_period['pn']=df_end_of_period['pn'].astype(str)
df_end_of_period=df_end_of_period.merge(df_16_wk[['pn','wk5']],how='left',on='pn')
df_end_of_period=df_end_of_period.merge(df_sales[['pn','tot_value']],how='left',on='pn')
df_order_list_proposed=df_end_of_period[['pn','wo','pzas_x_hacer','wk5','tot_value','create_wo']].sort_values(by=['wk5','tot_value','create_wo','pn'])
df_order_list_proposed.reset_index(inplace=True,drop=True)
df_order_list_proposed.reset_index(inplace=True,names='priority')
df_order_list_proposed['machine']=''
df_order_list_proposed['operation_description']=''
df_order_list_proposed['status']=''
df_order_list_proposed = df_order_list_proposed.loc[:, ~df_order_list_proposed.columns.duplicated(keep='last')]
df_order_list_proposed=rename_columns(df_order_list_proposed,df_col_rel=df_col_rel,table_from='End of Period',table_to='Lista de ordenes')
df_order_list_proposed=df_order_list_proposed[df_columns.loc[df_columns['table']=='Lista de ordenes','column_name'].to_list()]
df_order_list_proposed.to_excel(os.path.join(state['folder_output'],'Lista de Ordenes.xlsx'))
#%%
df_plan_new=get_predefined_df(df_columns=df_columns,table='Manufacturing plan')
#%%%
df_plan_new.columns