# ---
# jupyter:
#   jupytext:
#     formats: ipynb,py:percent
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.16.7
#   kernelspec:
#     display_name: Python 3
#     language: python
#     name: python3
# ---

# %% [markdown]
# # Manufacturing 
# - V1. 2025-04-07
#     - Version inicial, calculo de plan de produccion

# %% [markdown]
# ## Seleccionar archivos

# %%
# Script functions
# -*- coding: utf-8 -*-
import pandas as pd
from pandas.tseries.offsets import BDay
import os
import pickle
import warnings
from datetime import datetime, timedelta, date
from ipytree import Node

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import TableStyleInfo, Table
from openpyxl.utils.cell import coordinate_from_string, get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter import Tk, filedialog as fd
import win32com.client
import shutil
from copy import copy
import ipywidgets as widgets
from IPython.display import display, Markdown, clear_output
import xml.etree.ElementTree as ET
import zipfile
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from types import SimpleNamespace

warnings.filterwarnings("ignore")

# File/Directory & System Utilities

def open_file_selection(initialdir='', filter_name='*.*'):
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    filetypes = (
        (filter_name, f"*{filter_name.split(' ')[0]}*.*"),
        ('All files', '*.*'),
        ('Excel', '*.xlsx'),
        ('CSV', '*.doc')
    )
    files = fd.askopenfilenames(filetypes=filetypes, initialdir=initialdir)
    root.destroy()
    return files

def select_directory(initialdir):
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    directory = fd.askdirectory(initialdir=initialdir)
    root.destroy()
    return directory

def is_file_open(filepath):
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, 'a'):
            pass
        return False
    except PermissionError:
        return True

def get_path(file_selectors, selector_name):
    file_selector = [fs for fs in file_selectors if fs.children[0].description[:-1] == selector_name]
    if len(file_selector) == 0:
        show_popup_message(f"No se encontro el selector: {selector_name}")
        raise SystemExit()
    file_selector = file_selector[0]
    if not file_selector.children[1].value:
        return None
    selected_path = file_selector.children[1].value.strip()
    return selected_path

def close_xl_if_open(path):
    if is_file_open(path):
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            workbook = excel.Workbooks(path)
            workbook.Save()
            workbook.Close()
        except:
            show_popup_message(f"Cerrar el archivo: {path}")
            raise SystemExit()

def set_paths(path):
    global output_paths
    output_paths = {}
    output_paths['path_xl_format'] = os.path.join(path, 'columns and formatting.xlsx')
    output_paths['path_plan'] = os.path.join(path, 'manufacturing plan.xlsx')

def set_col_rel(output_paths):
    global df_col_rel,df_columns,col_names
    df_columns=read_excel(output_paths['path_xl_format'],sheet_name='column_format')
    df_col_rel=df_columns[~df_columns['std_name'].isnull()].copy()

# User Interface Management & Event Handling

def show_popup_message(message, df=pd.DataFrame()):
    display(Markdown(f"### **{message}**"))
    if len(df) > 0:
        display(Markdown(df.to_markdown(index=False)))

def verify_selections(file_selectors):
    not_selected = []
    for file_selector in file_selectors:
        selector = file_selector.children[0].description[:-1]
        selected = file_selector.children[1].value.strip()
        if not os.path.exists(selected):
            file_selector.children[1].value = 'Not selected'
            selected = 'Not selected'
        if selector not in mandatory_selectors:
            continue
        if selected == 'Not selected':
            not_selected.append(selector)
            continue
    if len(not_selected) > 0:
        show_popup_message(f'Favor de seleccionar los siguientes archivos:{not_selected}')
        raise SystemExit()
    
def on_output_button_click(b):
    if state['folder_output']:
        initialdir = state['folder_output']
    else:
        initialdir = '/'
    selected_dir = select_directory(initialdir=initialdir)
    if selected_dir:
        folder_output_label.value = f"{selected_dir}"
        state['folder_output'] = selected_dir
        save_state_pickle(state)
    else:
        folder_output_label.value = "Not selected"
    set_paths(folder_output_label.value)

def on_date_change(change):
    state['initial_date'] = change['new']
    save_state_pickle(state)

def on_folder_select(change):
    if tree.selected_nodes:
        state['mail_folder'] = str(tree.selected_nodes[0].name)
        global mail_folder
        mail_folder = state['mail_folder']
        save_state_pickle(state)

def get_folder_structure(folder, default_name):
    node = Node(folder.Name)
    found_node = None
    for subfolder in folder.Folders:
        sub_node, found = get_folder_structure(subfolder, default_name)
        node.add_node(sub_node)
        if found:
            found_node = found
    if folder.Name == default_name:
        node.selected = True
        found_node = node
    return node, found_node

# Error Handling / Decorators

def handle_permission_error_with_popup(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except PermissionError as e:
            if e.errno == 13:
                show_popup_message(f"Error: {e}\nFavor de cerrar el archivo.")
    return wrapper

# Excel Management Functions

def save_df(df, filepath, sheet_name='Sheet', index=False):
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index)

@handle_permission_error_with_popup
def save_df_multiple(df_dict=dict(), filepath='', index=False):
    with pd.ExcelWriter(filepath) as writer:
        for key in df_dict.keys():
            df_dict[key].to_excel(writer, sheet_name=key, index=index)

@handle_permission_error_with_popup
def save_wb(wb, filepath):
    wb.save(filepath)
    wb.close()

def read_excel(path=None, sheet_name=0, header=0, keep_default_na=True, dtype=None):
    with pd.ExcelFile(path) as xls:
        df = pd.read_excel(path, sheet_name=sheet_name, header=header, keep_default_na=keep_default_na, dtype=dtype)
    return df

def load_excel_with_header_key(file_path, sheet_name=0, key_text='', dtype=None, **kwargs):
    df = read_excel(file_path, sheet_name=sheet_name, keep_default_na=False, dtype=dtype)
    header_row = None
    if key_text in df.columns:
        header_row=0
    else:
        for col in df.columns:
            for i, cell in df[col].items():
                if pd.notna(cell) and key_text in str(cell):
                    header_row = i
                    break
            if header_row is not None:
                break
        if header_row is None:
            raise ValueError(f"Key text '{key_text}' not found in the sheet.")
        df.columns = df.iloc[header_row]
        df = df.iloc[header_row + 1:].reset_index(drop=True)
    return df

def find_cell_by_text(ws, text):
    for row in ws[ws.calculate_dimension()]:
        for cell in row:
            if cell.value == text:
                return cell.coordinate
    return None

def get_cell_properties(cell):
    properties = {}
    fill = cell.fill
    properties["background_color"] = fill
    properties["font"] = cell.font
    properties["alignment"] = cell.alignment
    return properties

def format_cell(cell, properties):
    cell.fill = copy(properties['background_color'])
    cell.font = copy(properties["font"])
    cell.alignment = copy(properties["alignment"])

def format_on_change(zip_cols, ws, start_row=1, format1=None, format2=None):
    if format1 is None:
        format1 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    if format2 is None:
        format2 = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    previous_values = None
    current_format = format1
    for row_idx, cells_rows in enumerate(zip_cols, start=start_row):
        combined_values = ""
        for cell in cells_rows:
            combined_values += str(cell.value)
        if previous_values != combined_values:
            current_format = format1 if current_format == format2 else format2
            previous_values = combined_values
        for cell in cells_rows:
            format_cell(cell, current_format)

def get_column_info(ws, col_name, raise_error=True):
    col_cell = find_cell_by_text(ws, col_name)
    if not col_cell:
        if not raise_error:
            return False
        show_popup_message(f"Column '{col_name}' not found in the sheet.")
        raise SystemExit()
    col_cell = ws[col_cell]
    last_cell = ws[ws.calculate_dimension()][-1][-1].coordinate
    data_range = ws[col_cell.offset(1, 0).coordinate:ws.cell(ws[last_cell].row, col_cell.offset(1, 0).column).coordinate]
    data_range_list = [cell[0] for cell in data_range]
    return {'data_range': data_range_list, 'col_cell': col_cell}

def get_xl_formatting(table_name=None):
    output_paths  # assumed to be global
    wb = load_workbook(output_paths['path_xl_format'])
    ws = wb['column_format']
    col_info = get_column_info(ws, 'column_name')
    cell_properties = {}
    for cell in col_info['data_range']:
        head_properties = get_cell_properties(cell)
        data_properties = get_cell_properties(cell.offset(0, 1))
        table = cell.offset(0, -1).value
        if table not in cell_properties:
            cell_properties[table] = {}
        cell_properties[table][cell.value] = {
            'head_properties': head_properties,
            'data_properties': data_properties
        }
    ws = wb['special_format']
    col_info = get_column_info(ws, 'format_name')
    cell_properties['special_format'] = {}
    for cell in col_info['data_range']:
        cell_properties['special_format'][cell.value] = get_cell_properties(cell)
    if table_name:
        cell_properties = cell_properties[table_name]
    return cell_properties

def get_col_sizes(wb):
    col_sizes = {}
    for sheet in wb.sheetnames:
        col_sizes[sheet] = wb[sheet].column_dimensions
    return col_sizes

def get_table_styles(wb):
    table_styles = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tables = ws._tables
        table_info = None
        for table in tables:
            if hasattr(table, "displayName"):
                style_name = table.tableStyleInfo.name if table.tableStyleInfo else "No Style"
                table_info = style_name
                break
        if table_info:
            table_styles[sheet_name] = table_info
    return table_styles

def apply_col_sizes(wb, col_sizes):
    for sheet in col_sizes.keys():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for column_letter in col_sizes[sheet].keys():
            if column_letter == 0:
                continue
            ws.column_dimensions[column_letter].width = col_sizes[sheet][column_letter].width
    return wb

def apply_table_styles(wb, table_styles):
    for sheet_name, style_name in table_styles.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws._tables:
                continue
            table_range = ws.calculate_dimension()
            table_name = f"{sheet_name}_Table".replace(" ", "")
            table = Table(displayName=table_name, ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name=style_name,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
    return wb

def get_locations(df, ws, column):
    coord = {}
    for text in df[column].drop_duplicates():
        coord[text] = find_cell_by_text(ws, text)
    return coord

def get_worksheet_df(ws, key_text=None, data_only=False):
    if key_text:
        header_row_idx = None
        for col in ws.iter_cols(min_row=ws.min_row, max_row=ws.max_row):
            for cell in col:
                if cell.value is not None and key_text in str(cell.value):
                    header_row_idx = cell.row
                    break
            if header_row_idx is not None:
                break
        if header_row_idx is None:
            show_popup_message(f"No se encontro el header: {key_text}")
            raise SystemExit()
    else:
        header_row_idx = 1
    header = [cell for cell in ws[header_row_idx]]
    num_cols = len(header)
    formula_dict = {}
    if not data_only:
        first_data_rows = list(ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 1, max_col=num_cols))
        if first_data_rows:
            first_data_row = first_data_rows[0]
            for i, cell in enumerate(first_data_row):
                col_name = header[i].value
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_dict[col_name] = cell.value
    data_rows = []
    cell_properties_dict = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, max_col=num_cols):
        row_dict = {}
        for i, cell in enumerate(row):
            col_name = header[i].value
            row_dict[col_name] = cell.value
            if data_only:
                row_dict[f"{col_name}_cell"] = cell.coordinate
            else:
                if (cell.fill.patternType) or ((cell.font.color) and (str(cell.font.color.rgb) != "Values must be of type <class 'str'>")):
                    cell_properties = get_cell_properties(cell)
                    if cell.row not in cell_properties_dict:
                        cell_properties_dict[cell.row] = {}
                    cell_properties_dict[cell.row][cell.coordinate] = cell_properties
                row_dict[f"{col_name}_cell"] = cell.coordinate
        data_rows.append(row_dict)
    df = pd.DataFrame(data_rows)
    header_dict = {}
    for cell in header:
        header_dict[cell.value] = {"coordinate": cell.coordinate, "properties": get_cell_properties(cell)}
    if len(df) == 0:
        df = pd.DataFrame(columns=header_dict.keys())
    worksheet_df = {
        "df": df,
        "header": header_dict,
        "formulas": formula_dict,
        "properties": cell_properties_dict
    }
    return worksheet_df

def update_sheet(ws_dict, ws, apply_formats=True):
    df = ws_dict["df"]
    header_info = ws_dict["header"]
    formulas = ws_dict.get("formulas", {})
    properties_info = ws_dict.get("properties", {})
    first_header = next(iter(header_info.values()))
    _, header_row = coordinate_from_string(first_header["coordinate"])
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)
    data_columns = [col for col in df.columns if col is not None and not col.endswith('_cell')]
    format_columns = [col for col in df.columns if col is not None and col.endswith('_cell')]
    header_columns = sorted(
        [col for col in header_info.keys() if col in data_columns],
        key=lambda x: column_index_from_string(coordinate_from_string(header_info[x]['coordinate'])[0])
    )
    remaining_columns = [col for col in data_columns if col not in header_columns]
    ordered_columns = header_columns + remaining_columns
    df_ordered = df.copy().reset_index(drop=True)
    df_ordered = df_ordered[ordered_columns]
    df_formats = df[format_columns]
    for col_index, header in enumerate(ordered_columns, start=1):
        ws.cell(row=header_row, column=col_index, value=header)
    start_data_row = header_row + 1
    for i, r in enumerate(dataframe_to_rows(df_ordered, index=False, header=False)):
        ws.append(r)
        if not apply_formats:
            continue
        formats = df_formats.loc[i, format_columns].dropna()
        for col in formats.keys():
            current_row = start_data_row + i
            cell = ws.cell(row=current_row, column=header_columns.index(col[:-5]) + 1)
            format_cell(cell, formats[col])
    for col, formula in formulas.items():
        origin_cell = ws.cell(start_data_row, ordered_columns.index(col) + 1)
        origin_cell.value = formula
        last_cell = origin_cell.offset(i, 0)
        cell_range = ws[origin_cell.offset(1, 0).coordinate:last_cell.coordinate]
        for cell in cell_range:
            cell[0].value = Translator(formula, origin=origin_cell.coordinate).translate_formula(cell[0].coordinate)
    return ws

def append_to_sheet(ws_dict, ws):
    df = ws_dict['df']
    cell_cols = [col for col in df.columns if col is not None and col.endswith('_cell')]
    data_cols = list(ws_dict["header"].keys())
    df_to_append = df[df[cell_cols[0]].isnull()][data_cols]
    for r in dataframe_to_rows(df_to_append, index=False, header=False):
        ws.append(r)
    return ws

def update_column(ws_dict, ws, column=None):
    df = ws_dict['df'].copy()
    if not column:
        return ws
    df.dropna(subset=[f'{column}_cell'], inplace=True)
    for index, row in df[[column, f'{column}_cell']].iterrows():
        ws[row[f'{column}_cell']].value = row[column]
    return ws

def create_new_sheet(wb, sheet_name='Sheet1', df=pd.DataFrame()):
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    return wb

def format_xl_dates(wb, sheet_name='', date_columns=[]):
    ws = wb[sheet_name]
    for col in date_columns:
        col_info = get_column_info(ws, col)
        for cell in col_info['data_range']:
            cell.number_format = 'mm/dd/yyyy'
    return wb

def extract_selected_sheets(xlsx_file, sheets_to_keep, keep_original=True):
    current_sheets = pd.ExcelFile(xlsx_file).sheet_names
    if set(sheets_to_keep) == set(current_sheets):
        return False
    file_dir = os.path.dirname(xlsx_file)
    file_name = os.path.basename(xlsx_file).replace(".xlsx", "")
    temp_dir = os.path.join(file_dir, "temp_xlsx")
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir, exist_ok=True)
    with zipfile.ZipFile(xlsx_file, "r") as zip_ref:
        zip_ref.extractall(temp_dir)
    original_temp_dir = os.path.join(file_dir, file_name)
    if os.path.exists(original_temp_dir):
        shutil.rmtree(original_temp_dir.strip())
    shutil.copytree(temp_dir, original_temp_dir.strip())
    workbook_xml_path = os.path.join(temp_dir, "xl", "workbook.xml")
    tree = ET.parse(workbook_xml_path)
    root = tree.getroot()
    ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    sheets_elem = root.find("ns:sheets", ns)
    sheet_mapping = {}
    if sheets_elem is not None:
        for idx, sheet in enumerate(sheets_elem.findall("ns:sheet", ns), start=1):
            sheet_name = sheet.get("name")
            sheet_mapping[sheet_name] = f"sheet{idx}.xml"
    worksheets_dir = os.path.join(temp_dir, "xl", "worksheets")
    for sheet_name, sheet_xml_file in sheet_mapping.items():
        sheet_xml_path = os.path.join(worksheets_dir, sheet_xml_file)
        if sheet_name not in sheets_to_keep:
            if os.path.exists(sheet_xml_path):
                os.remove(sheet_xml_path)
    reduced_xlsx_file = os.path.join(file_dir, xlsx_file)
    with zipfile.ZipFile(reduced_xlsx_file, "w", zipfile.ZIP_DEFLATED) as new_zip:
        for root_dir, _, files in os.walk(temp_dir):
            for file in files:
                file_path = os.path.join(root_dir, file)
                arcname = os.path.relpath(file_path, temp_dir)
                new_zip.write(file_path, arcname)
    if not keep_original:
        shutil.rmtree(original_temp_dir.strip())
    shutil.rmtree(temp_dir)
    return sheet_mapping

def integrate_modified_file_with_backup(modified_xlsx_file, sheets_mapping):
    file_dir = os.path.dirname(modified_xlsx_file)
    base_name = os.path.basename(modified_xlsx_file).replace(".xlsx", "")
    temp_modified = os.path.join(file_dir, "temp_xlsx_reduced")
    if os.path.exists(temp_modified):
        shutil.rmtree(temp_modified)
    os.makedirs(temp_modified, exist_ok=True)
    with zipfile.ZipFile(modified_xlsx_file, "r") as zip_ref:
        zip_ref.extractall(temp_modified)
    mod_workbook_xml_path = os.path.join(temp_modified, "xl", "workbook.xml")
    mod_tree = ET.parse(mod_workbook_xml_path)
    mod_root = mod_tree.getroot()
    mod_ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    mod_sheets_elem = mod_root.find("ns:sheets", mod_ns)
    mod_sheet_mapping = {}
    if mod_sheets_elem is not None:
        for idx, sheet in enumerate(mod_sheets_elem.findall("ns:sheet", mod_ns), start=1):
            sheet_name = sheet.get("name")
            sheet_file = f"sheet{idx}.xml"
            mod_sheet_mapping[sheet_name] = sheet_file
    temp_original = os.path.join(file_dir, base_name)
    worksheets_original = os.path.join(temp_original, "xl", "worksheets")
    worksheets_modified = os.path.join(temp_modified, "xl", "worksheets")
    for sheet_name, mod_sheet_file in mod_sheet_mapping.items():
        if sheet_name in sheets_mapping:
            src_sheet = os.path.join(worksheets_modified, mod_sheet_file)
            dest_sheet = os.path.join(worksheets_original, sheets_mapping[sheet_name])
            if os.path.exists(src_sheet):
                shutil.copy(src_sheet, dest_sheet)
    integrated_xlsx_file = os.path.join(file_dir, modified_xlsx_file)
    with zipfile.ZipFile(integrated_xlsx_file, "w", zipfile.ZIP_DEFLATED) as zip_out:
        for folder, _, files in os.walk(temp_original):
            for file in files:
                file_path = os.path.join(folder, file)
                arcname = os.path.relpath(file_path, temp_original)
                zip_out.write(file_path, arcname)
    shutil.rmtree(temp_modified)

def get_df_idx(df, idx_cols, idx_name='idx'):
    df_idx = df.reset_index()
    df_idx.rename({'index': idx_name}, axis=1, inplace=True)
    df_idx = df_idx.drop_duplicates(idx_cols, keep='first')
    df_idx = df_idx[idx_cols + [idx_name]]
    return df_idx

def set_hyperlink(df, sheet_name, col_name, idx_name, typ='str'):
    df = df.copy()
    idx = ~df[idx_name].isnull()
    quote = '\"' if typ == 'str' else ''
    df.loc[idx, col_name] = f"=HYPERLINK(\"#'{sheet_name}'!A" + (df.loc[idx, idx_name] + 2).astype(int).astype(str) + f"\",{quote}" + (df.loc[idx, col_name]).astype(str) + f"{quote})"
    return df

def fill_yield_report(df, wb, sheet_name='', search_parms=None):
    if search_parms is None:
        search_parms = {
            "column": 'date_from',
            "row": 'xl_field',
            "offset": {
                "Category": {"scheduled": (0, 0), "real": (1, 0)},
                "Shift": {"1s": (0, 0), "2s": (0, 2)},
            }
        }
    ws = wb[sheet_name]
    df_data_coord = df[[search_parms['column'], search_parms['row']]].drop_duplicates()
    df_data_coord['row'] = ''
    df_data_coord['column'] = ''
    coord_y = get_locations(df_data_coord, ws, search_parms['row'])
    coord_x = get_locations(df_data_coord, ws, search_parms['column'])
    for _, row in df.iterrows():
        if not (coord_x.get(row[search_parms['column']]) and coord_y.get(row[search_parms['row']])):
            continue
        cell = ws.cell(ws[coord_y[row[search_parms['row']]]].row, ws[coord_x[row[search_parms['column']]]].column)
        for key in search_parms['offset'].keys():
            if key not in df.columns:
                continue
            offset = search_parms['offset'][key][row[key]]
            cell = cell.offset(offset[0], offset[1])
        cell.value = row['Total Tested']
    return wb

# Persistence / State Management

def save_state_pickle(state, filename='folder_state_planner.pkl'):
    with open(filename, 'wb') as f:
        pickle.dump(state, f)

def load_state_pickle(filename='folder_state_planner.pkl'):
    try:
        with open(filename, 'rb') as f:
            return pickle.load(f)
    except FileNotFoundError:
        return {"folder_input": None, "folder_output": None, "selections": {}}



# DataFrame Management & Data Processing

def format_dates(df, date_cols=[], type='iso'):
    for col in date_cols:
        if col not in df.columns:
            continue
        if type == 'iso':
            df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
        else:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    return df

def append_df_to_df(df_new=pd.DataFrame(), df_old=pd.DataFrame(), table='', keys=[], date_cols=[]):
    if df_new.empty:
        return
    if not keys:
        keys = df_new.columns
    df_new = df_new.copy()
    df_new = format_dates(df_new, date_cols)
    df_old = format_dates(df_old, date_cols)
    df_old['composite_key'] = list(zip(*(df_old[col] for col in keys)))
    df_new['composite_key'] = list(zip(*(df_new[col] for col in keys)))
    df_new = df_new[~df_new['composite_key'].isin(df_old['composite_key'])]
    df_old = pd.concat([df_old, df_new])
    df_old.drop(columns=['composite_key'], inplace=True)
    df_old.reset_index(inplace=True, drop=True)
    df_grp = df_old.groupby(keys).count()
    df_grp = df_grp[df_grp[df_grp.columns[0]] > 1]
    if len(df_grp) > 0:
        df_grp.reset_index(inplace=True)
        show_popup_message(f"Hay duplicados en el archivo {table} para la llave {keys}:", df_grp[keys].sort_values(keys))
        raise SystemExit()
    return df_old

def update_dataframe(df_original, df_to_integrate, key_cols=[], exceptions=[]):
    df1_indexed = df_original.set_index(key_cols)
    df2_indexed = df_to_integrate.set_index(key_cols)
    common_cols = df1_indexed.columns.intersection(df2_indexed.columns)
    common_cols = [col for col in common_cols if col not in exceptions]
    df1_indexed.update(df2_indexed[common_cols])
    new_rows = df2_indexed.loc[~df2_indexed.index.isin(df1_indexed.index)]
    df_original = pd.concat([df1_indexed, new_rows]).reset_index()
    return df_original

def duplicate_df_cols(df, suffix, cols):
    subset = df[cols].copy()
    subset = subset.rename(columns=lambda x: x + suffix)
    cols_suffixed = [f"{x}{suffix}" for x in cols]
    df = df.drop(columns=[col for col in cols_suffixed if col in df.columns], errors='ignore')
    df = pd.concat([df, subset], axis=1)
    return {'df': df, 'cols_suffixed': cols_suffixed}

def move_columns_to_front(df, columns):
    columns = [col for col in columns if col in df.columns]
    remaining_columns = [col for col in df.columns if col not in columns]
    return df[columns + remaining_columns]

def check_duplicated_columns(df):
    duplicated_cols = df.columns[df.columns.duplicated()].to_list()
    if len(duplicated_cols) == 0:
        return
    show_popup_message(f"Columnas duplicadas: {duplicated_cols}")
    raise SystemExit()

def assign_quantities(df_pos, df_to_assign, additional_fields=[]):
    df_pos = df_pos.copy()
    df_to_assign = df_to_assign.copy()
    df_pos['Assigned'] = 0
    df_pos['quantity'] = pd.to_numeric(df_pos['quantity']).astype(float)
    df_assignments = []
    for ln_index, ln_row in df_to_assign[['po', 'modelo']].drop_duplicates().iterrows():
        df_assign_subset = df_to_assign[(df_to_assign['po'] == ln_row['po']) & (df_to_assign['modelo'] == ln_row['modelo'])].copy()
        df_assign_subset['Remaining'] = df_assign_subset['quantity']
        for wo_index, subset_row in df_assign_subset.iterrows():
            remaining_produced = subset_row['Remaining']
            for index, row in df_pos[(df_pos['po'] == ln_row['po']) & (df_pos['modelo'] == ln_row['modelo'])].iterrows():
                if remaining_produced <= 0:
                    break
                available_space = row['quantity'] - df_pos.at[index, 'Assigned']
                assignable = min(available_space, remaining_produced)
                if assignable > 0:
                    df_pos.at[index, 'Assigned'] += assignable
                    remaining_produced -= assignable
                    line_assignment = {
                        'po': row['po'],
                        'modelo': row['modelo'],
                        'AvailQuantity': subset_row['quantity'],
                        'LineNumber': row['LineNumber'],
                        'Assigned_Quantity': assignable
                    }
                    if additional_fields:
                        for field in additional_fields:
                            if field in subset_row:
                                line_assignment[field] = subset_row[field]
                    df_assignments.append(line_assignment)
    df_assignments = pd.DataFrame(df_assignments)
    assigned = {'df_pos': df_pos, 'df_assignments': df_assignments}
    return assigned

def rename_columns(df, df_col_rel, table_from='std_name', sheet_from='only', table_to='std_name', sheet_to='only'):
    required_cols = {'table', 'sheet', 'column_name', 'std_name'}
    missing_cols = required_cols - set(df_col_rel.columns)
    if missing_cols:
        raise ValueError(f"df_col_rel is missing required columns: {missing_cols}")
    mapping = {}
    if table_to == 'std_name':
        filtered = df_col_rel[(df_col_rel['table'] == table_from) & (df_col_rel['sheet'] == sheet_from)]
        if filtered.empty:
            raise ValueError("No matching mapping found for given table_from and sheet_from.")
        mapping = filtered.set_index('column_name')['std_name'].to_dict()
    elif table_from == 'std_name':
        filtered = df_col_rel[(df_col_rel['table'] == table_to) & (df_col_rel['sheet'] == sheet_to)]
        if filtered.empty:
            raise ValueError("No matching mapping found for given table_to and sheet_to.")
        mapping = filtered.set_index('std_name')['column_name'].to_dict()
    else:
        filtered_from = df_col_rel[(df_col_rel['table'] == table_from) & (df_col_rel['sheet'] == sheet_from)]
        if filtered_from.empty:
            raise ValueError("No matching mapping found for given table_from and sheet_from.")
        mapping_from = filtered_from.set_index('column_name')['std_name'].to_dict()
        df = df.rename(columns=mapping_from)
        filtered_to = df_col_rel[(df_col_rel['table'] == table_to) & (df_col_rel['sheet'] == sheet_to)]
        if filtered_to.empty:
            raise ValueError("No matching mapping found for given table_to and sheet_to.")
        mapping = filtered_to.set_index('std_name')['column_name'].to_dict()
    return df.rename(columns=mapping)

def check_mandatory_cols(cols, selector_name, raise_error=True):
    missing_columns = [col for col in mandatory_cols[selector_name] if col not in cols]
    if len(missing_columns) > 0:
        show_popup_message(f"No se encontraron las siguientes columnas en el archivo {selector_name}: {missing_columns}")
        if raise_error:
            raise SystemExit()
        return False
    return True

def set_family(df, column, dest_col='Family'):
    df[column].fillna('', inplace=True)
    for key in families.keys():
        df.loc[df[column].str.startswith(tuple(families[key])), dest_col] = key
    df[dest_col].fillna('Other', inplace=True)
    return df

def map_yield_report(df):
    df_yield = df.copy()
    df_yield.columns = df_yield.iloc[8].fillna('NA').to_list()
    df['Yield Reports'].fillna('', inplace=True)
    date_from = df[df['Yield Reports'].str.contains('From')].iloc[0]['Yield Reports']
    date_to = df[df['Yield Reports'].str.contains('From')].iloc[0]['Unnamed: 5']
    df_yield['date_from'] = date_from
    df_yield['date_to'] = date_to
    df_yield = df_yield[(~df_yield['Part Number'].isna()) & (df_yield['Part Number'] != 'Part Number')]
    df_yield['date_from'] = df_yield['date_from'].str.replace('From: ', '')
    df_yield['date_to'] = df_yield['date_to'].str.replace('To: ', '')
    df_yield['date_from'] = pd.to_datetime(df_yield['date_from'])
    df_yield['date_to'] = pd.to_datetime(df_yield['date_to'])
    df_yield.drop('NA', axis=1, inplace=True)
    df_yield.fillna(0, inplace=True)
    df_yield['month'] = df_yield['date_from'].dt.strftime('%Y-%m')
    return df_yield

def merge_additional_fields(df=pd.DataFrame(), df_edi=pd.DataFrame(), fields=[], sort_fields=[], key=[]):
    if len(df) == 0:
        df = pd.DataFrame(columns=fields)
    df = df.sort_values(sort_fields)
    df.drop_duplicates(key, keep='last', inplace=True)
    df_edi = df_edi.merge(df[fields], how='left', on=key)
    return df_edi

def apply_grouping(df, key_cols=['po'], list_cols=[], qty_cols=[], date_cols=[], prefix=''):
    df[list_cols] = df[list_cols].astype(str)
    df = df[key_cols + list_cols + date_cols + qty_cols]
    result_dict = duplicate_df_cols(df, '_first', date_cols)
    df = result_dict['df']
    date_cols_first = result_dict['cols_suffixed']
    result_dict = duplicate_df_cols(df, '_last', date_cols)
    df = result_dict['df']
    date_cols_last = result_dict['cols_suffixed']
    df[qty_cols] = df[qty_cols].astype(float)
    list_col_agg_dict = {str(x): lambda x: ','.join(x.drop_duplicates().sort_values().astype(str)) for x in list_cols}
    qty_col_agg_dict = {str(x): 'sum' for x in qty_cols}
    date_col_first_agg_dict = {str(x): lambda x: pd.to_datetime(x, errors='coerce').min() for x in date_cols_first}
    date_col_last_agg_dict = {str(x): lambda x: pd.to_datetime(x, errors='coerce').max() for x in date_cols_last}
    df = df.groupby(key_cols).agg({**date_col_first_agg_dict, **date_col_last_agg_dict, **list_col_agg_dict, **qty_col_agg_dict})
    df = df.rename(columns={col: f"{prefix}{col}" for col in df.columns if col not in key_cols})
    df.reset_index(inplace=True)
    ns = SimpleNamespace(**{"df": df,
                            "date_cols": [f"{prefix}{col}" for col in date_cols_first] + [f"{prefix}{col}" for col in date_cols_last],
                            "list_cols": [f"{prefix}{col}" for col in list_cols],
                            "qty_cols": [f"{prefix}{col}" for col in qty_cols]})
    return ns

def apply_dynamic_rules(df, rules_df, target_col='status'):
    df = df.copy()
    rules_df = rules_df.sort_values(['priority', 'group'])
    grouped_rules = rules_df.groupby(['priority', 'group'])
    for group_name, group_data in grouped_rules:
        possible_results = group_data['result'].unique()
        if len(possible_results) > 1:
            raise ValueError(f"Multiple 'result' values in group '{group_name}': {possible_results}")
        group_result = possible_results[0]
        query_expr, var_dict = build_group_query(group_data)
        selected = df.query(query_expr, local_dict=var_dict)
        df.loc[selected.index, target_col] = group_result
    return df

def build_group_query(group_data):
    op_map = {'=': '==', '!=': '!=', '<': '<', '<=': '<=', '>': '>', '>=': '>='}
    var_dict = {}
    var_counter = 0
    rules_dict = {}
    for _, row in group_data.iterrows():
        parsed_val = parse_value(row['options'])
        var_name = f"_v{var_counter}"
        var_counter += 1
        var_dict[var_name] = parsed_val
        operator = op_map.get(row['test'])
        if operator is None:
            raise ValueError(f"Unsupported test operator: {row['test']}")
        expr_str = f"`{row['column']}` {operator} @{var_name}"
        cond = str(row['condition']).strip().lower() if pd.notnull(row['condition']) else ""
        key = row['condition_group']
        if key in rules_dict:
            rules_dict[key].append((expr_str, cond))
        else:
            rules_dict[key] = [(expr_str, cond)]
    group_expressions = []
    for group_key, rule_list in rules_dict.items():
        group_expr = rule_list[0][0]
        for i in range(1, len(rule_list)):
            link_op = rule_list[i-1][1] if rule_list[i-1][1] else "and"
            group_expr += f" {link_op} " + rule_list[i][0]
        if len(rule_list) > 1:
            group_expr = f"({group_expr})"
        group_expressions.append(group_expr)
    full_expr = " and ".join(group_expressions)
    return full_expr, var_dict

def parse_value(value):
    if isinstance(value, str):
        val_lower = value.lower().strip()
        if val_lower == 'no date':
            return pd.Timestamp('2099-12-31')
        if val_lower.startswith('today +'):
            remainder = val_lower[len('today +'):]
            parts = remainder.split()
            try:
                offset_n = int(parts[0])
            except (ValueError, IndexError):
                raise ValueError(f"Cannot parse numeric offset from '{value}'")
            if len(parts) > 1 and 'workday' in parts[1]:
                return pd.Timestamp.now().normalize() + offset_n * BDay()
            else:
                return pd.Timestamp.now().normalize() + pd.Timedelta(days=offset_n)
        else:
            try:
                return float(value)
            except ValueError:
                return value
    else:
        return value

filters=['Master Doblado','Lista de ordenes','Routing']

state = load_state_pickle()
if 'folder_output' not in state:
    state['folder_output'] = ''

if state['folder_output']:
    initialdir = state['folder_output']
    set_paths(initialdir)
else:
    initialdir = 'Not selected'





folder_output_label = widgets.Label(value=initialdir)
folder_output_button = widgets.Button(description="Folder de salidas:")
folder_output_button.on_click(on_output_button_click)


# Create an array of button and label widgets
file_selectors = []
for filter_name in filters:
    # Create button and label
    button = widgets.Button(description=f"{filter_name}:")
    if state.get(filter_name):
        label = widgets.Label(value=f" {state.get(filter_name)}")
    else:
        label = widgets.Label(value=f" Not selected")
    
    # Define the button click event
    def on_button_click(filter_name=filter_name, label=label):
        show_popup_message('message')
        if (filter_name in state.keys()) and (state[filter_name]):
            initialdir=os.path.dirname(state[filter_name][0])
        else:
            initialdir='/'
        selected_dir = open_file_selection(initialdir=initialdir,filter_name=filter_name)  # Adjust initialdir as needed
        if selected_dir:
            label.value = f" {selected_dir[0]}"
            state[filter_name]=selected_dir[0]
            
            
        else:
            label.value = f" Not selected"
            state[filter_name]=f" Not selected"
        save_state_pickle(state)

    # Attach the event to the button
    button.on_click(lambda b, f=filter_name, l=label: on_button_click(f, l))
    
    # Add the button and label as a horizontal box
    file_selectors.append(widgets.HBox([button, label]))

initial_date=date.today()
if 'initial_date' in state.keys():
    initial_date=state['initial_date']
datepicker_init_date = widgets.DatePicker(
    description='Fecha Inicial de Programacion',
    disabled=False,
    value=initial_date,
    layout=widgets.Layout(width='350px'),
    style={'description_width': '170px'}
)
datepicker_init_date.observe(on_date_change, names='value')

ui = widgets.VBox([
    widgets.HBox([folder_output_button, folder_output_label])  ,
    datepicker_init_date 
] + file_selectors)
display(ui)
set_col_rel(output_paths)
# %% [markdown]
# ## Crear plan

# %%
# Load files
if not os.path.exists(output_paths['path_xl_format']):
    show_popup_message("No se encuentra el archivo: columns and formatting.xlsx")
    raise SystemExit()

if folder_output_label.value=='Not selected':
    show_popup_message("Seleccione el folder de salida")
    raise SystemExit()
path_master_doblado=get_path(file_selectors,'Master Doblado')
df_master_doblado=load_excel_with_header_key(path_master_doblado,sheet_name='00. Formato para Master de WC',key_text='PN')
df_master_doblado=rename_columns(df_master_doblado,df_col_rel,table_from='Master Doblado',sheet_from='00. Formato para Master de WC')

path_routing=get_path(file_selectors,'Routing')
df_routing=load_excel_with_header_key(path_routing,sheet_name='Operations',key_text='Routing')
df_routing=rename_columns(df_routing,df_col_rel,table_from='Routing',sheet_from='Operations')

path_order_list=get_path(file_selectors,'Lista de ordenes')
df_order_list=load_excel_with_header_key(path_order_list,key_text='Priority')
df_order_list=rename_columns(df_order_list,df_col_rel,table_from='Lista de ordenes')
# %
# Initialize machine status dictionary with two shifts per day
machine_status = {}

assignments = []
df_order_list.sort_values('priority',inplace=True)
shifts={'PRIMER TURNO': 8.0, 'SEGUNDO TURNO': 7.0}
for idx, order in df_order_list.iterrows():
    pn = order['pn']
    qty = order['pzas_x_hacer']
    wo=order['wo']
    pty=order['priority']
    # Get processing times for this part number
    pn_info = df_routing[df_routing['pn'] == pn]
    if pn_info.empty:
        continue
    run_time = pn_info.iloc[0]['run_time']
    setup_time = pn_info.iloc[0]['setup_time']
    
    # Get available machines for this part number from machine_options
    pn_machines = df_master_doblado[df_master_doblado['pn'] == pn]
    if pn_machines.empty:
        continue
    row = pn_machines.iloc[0]
    machines=[]
    for key, item in row.items():
        if ('maq_opc' in key) & (item is not None) & (item is not ''):
            machines.append(item)

    # Initialize machine status for these machines if not already done
    for m in machines:
        if m not in machine_status:
            machine_status[m] = {'day': 1, 'avail': copy(shifts), 'last_pn': None}
    if len(machines)==0:
        show_popup_message(f"Favor de asignar maquina al PN:{pn}")
        raise SystemExit()
    # Process the order until all pieces are assigned
    while qty > 0:
        assigned = False
        # Try each machine option in order and then each shift
        for m in machines:
            for shift in shifts.keys():
                available_time = machine_status[m]['avail'][shift]
                # If switching to a new part number, include setup time
                if machine_status[m]['last_pn'] != pn:
                    if run_time == 0:
                        pieces_to_assign = qty
                        time_used = setup_time
                    else:
                        if available_time < (setup_time + run_time):
                            continue
                        pieces_possible = 1 + int((available_time - (setup_time + run_time)) / run_time)
                        pieces_possible = max(pieces_possible, 1)
                        pieces_to_assign = min(pieces_possible, qty)
                        time_used = setup_time + pieces_to_assign * run_time
                else:
                    # Same part continues; no setup time required
                    if run_time == 0:
                        pieces_to_assign = qty
                        time_used = 0
                    else:
                        if available_time < run_time:
                            continue
                        pieces_possible = int(available_time // run_time)
                        pieces_to_assign = min(pieces_possible, qty)
                        time_used = pieces_to_assign * run_time
                        
                assignments.append({
                    'day': machine_status[m]['day'],
                    'machine': m,
                    'shift': shift,
                    'pn': pn,
                    'wo':wo,
                    'priority':pty,
                    'pzas_x_hacer': pieces_to_assign,
                    'time_used': time_used
                })
                machine_status[m]['avail'][shift] -= time_used
                machine_status[m]['last_pn'] = pn
                qty -= pieces_to_assign
                assigned = True
                if qty == 0:
                    break
            if qty == 0:
                break
        # If none of the machines/shifts can process any piece, roll over to the next day
        if not assigned:
            for m in machines:
                machine_status[m]['day'] += 1
                machine_status[m]['avail'] = copy(shifts)
                machine_status[m]['last_pn'] = None

# Generate the planning report DataFrame and print it
report = pd.DataFrame(assignments)
max_day = report['day'].max()
workdays = pd.bdate_range(start=pd.to_datetime(datepicker_init_date.value), periods=max_day)
day_to_date = {day: workdays[day - 1] for day in range(1, max_day + 1)}
report['date'] = report['day'].map(day_to_date)
report.sort_values(['date','machine','shift'])
report.sort_values(['date','machine','shift','priority','wo'],inplace=True)
report.to_excel(output_paths['path_plan'],index=False)
