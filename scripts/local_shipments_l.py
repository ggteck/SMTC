"""
# Seguimiento a embarques
- V32. 2025-04-14
    - Mejoras de eficiencia
- V31. 2025-04-12
    - Cambiar a streamlit
- V30. 2025-04-02
    - Conservar fechas del OOR
- V29. 2025-04-01
    - Calculo de TAT Category
- V28. 2025-03-31
    - Correccion para columnas del Tracker, quitar ceros en algunas col, eliminar _c1 y _c2 de status
- V27. 2025-03-30
    - Cambios a la integracion de gating part report, actualizar latest commit date
- V26. 2025-03-26
    - Correcciones, se verifica que se cierren los archivos al inicio, se muestran todos los duplicados
- V25. 2025-03-25
    - Se cambia la manera de actualizar elp master log
- V24. 2025-03-23
    - Setup de Status, integracion de Gating parts report
- V23. 2025-03-19
    - Correcciones, los campos no actualizables se definen en el archivo de formatos
- V22. 2025-03-17
    - Sistema de actualizacion de status
- V21. 2025-03-13
    - Correccion en mail, el filtro de fecha
- V20. 2025-03-12
    - Se eliminan hojas no necesarias del archivo OOR
    - Correcciones y mejoras de eficiencia
- V20. 2025-03-10
    - Proceso de definicion de Status
- V19. 2025-03-07
    - Se conserva el OOR original con formulas y datos generados por el usuario
- V18. 2025-03-03
    - Correccion a la integracion de inventorystagebackup
- V17. 2025-02-26
    - Se agregan cajas separadas del inventorystagebackup
    - Se buscan las columnas en el OOR Old
- V16. 2025-02-24
    - Cambio Cancelada por Cancelled
- V15. 2025-02-24
    - Los resultados de yield reports se agregan a la hoja Trov Daily Status NEW
- V14. 2025-02-23
    - Correccion menor
- V13. 2025-02-19
    - Se concatenan las work orders
    - Se integran fechas de llegada al EDI a partir de un OOR antiguo 
- V12. 2025-02-19
    - Se utiliza ELP Master log como fuente de datos
    - Se agregan columnas y calculos al OOR
- V11. 2025-02-13
    - Revisar columnas de elp master log
- V10. 2025-02-09
    - Correccion en links
    - Se agrega fecha al OOR
- V9. 2025-02-08
    - Correccion al proceso de integracion de envios a ELP
    - El Shipment transactions ya no es acumulable debido a que no se pueden eliminar duplicados
    - Se agrega proceso de integracion de Yield reports (falta determinar como visualizarlos)
    - Se agrega inicio para panel
- V8. 2025-02-06
    - El archivo korrus_Data no es obligatorio
    - Drop Zone es NULL en lugar de blanco
- V7. 2025-02-04
    - Mejoras en el manejo de errores
- V6. 2025-02-03
    - Se integra el archivo de ELP master log si esta seleccionado
    - Correcciones para la primera ejecicion del script
- V5. 2025-02-01
    - Se elimino fecha de reparacion    
- V4. 2025-01-31
    - Correcciones para evitar errores por dataframes vacios
    - Se ajustan las columnas de embarques al paso para que sean los del reporte consolidado actual
- V3. 2025-01-28
    - Se guarda el ancho de columnas
    - Se cambia el orden de algunas columnas
- V2. 2025-01-28
    - Se consolidan los archivos Master Edi, Shipped to Cust y Shipped to ELP
    - Se genera el reporte OOR
- V1. Version inicial
"""

import streamlit as st
from pandas.tseries.offsets import BDay
import os, shutil
import pickle
import pandas as pd
from openpyxl import load_workbook 
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import TableStyleInfo, Table
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import coordinate_from_string, get_column_letter, coordinate_from_string, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from types import SimpleNamespace
from datetime import date, timedelta, datetime
from tkinter import Tk, filedialog as fd
import win32com.client
import pythoncom
from copy import copy
import zipfile
import xml.etree.ElementTree as ET
import warnings
warnings.filterwarnings("ignore")
# ----------------------------------------------------------------
# Mandatory columns dictionary for check_mandatory_cols (for Korrus)
# ----------------------------------------------------------------
families={'TROV':['L','CS-L','CC-L'],
          'RISE':['F','CS-F'],
          'Accesories':['RISE','M','CBL','LENS','WMA'],
          'DZ':['AssignedDropZone']}

mandatory_cols={
    'Korrus':[
        'PurchaseOrder',
        'PODate',
        'REF02_ClearTextClause',
        'REF02_CustomerOrderNumber',
        'REF02_CarrierAccount',
        'TermsTypeCode',
        'TermsBasisDateCode',
        'description',
        'Routing',
        'JobName',
        'JobNameNotes',
        'ShipmentMarkings',
        'ShipmentMarkingsNotes',
        'ShipToParty',
        'ShipToAddress1',
        'ShipToAddress2',
        'ShipToCity',
        'ShipToState',
        'ShipToPostalCode',
        'ShipToCountryCode',
        'BillToParty',
        'Address1',
        'Address2',
        'BillToCity',
        'BillToState',
        'BillToPostalCode',
        'BillToCountryCode',
        'Supplier',
        'SupplierAddress1',
        'SupplierAddress2',
        'SupplierCity',
        'SupplierState',
        'SupplierPostalCode',
        'SupplierCountryCode',
        'LineNumber',
        'Quantity',
        'Uom',
        'price',
        'ProductServiceID',
        'StorageLocation',
        'AssignedDropZone'
        ],
    'Tracker':[
        'PO cliente',
        'Modelo',
        'WO\n QTY',
        'START DATE',
        'REPROGRAMACION',
        'FINISH DATE',
        'SHIP DATE'
        ],
    'Shipment transactions':[
        'Part No.',
        'Customer PO#',
        'Date Shipped',
        'Qty. Shipped'
    ],
    'InventoryStageBakup':[
        'Producto',
        'PO',
        'Box Id',
        'Cantidad'
    ],
    'ELP Master':[
        'CUU ship Date', 
        'PN', 
        'PO', 
        'DZ', 
        'BOX ID', 
        'Unit\nQTY'],
    'OOR':[
        'PurchaseOrder', 
        'ProductServiceID', 
        'LineNumber', 
        'EDI Received']       
                }

mandatory_selectors=['Tracker','ELP Master','OOR']

sheets_to_keep=['OOR',
 'Trov Daily Status',
 'Rise Daily Status',
 'WorkOrder Detail',
 'Shipped to Elp',
 'Shipped to Cust',
 'SMTC Daily Tracker']

# ----------------------------------------------------------------
# File / Directory Utilities
# ----------------------------------------------------------------

# Decorator to handle the permission error
def handle_permission_error_with_popup(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except PermissionError as e:
            if e.errno == 13:  # Permission denied error
                st.info(f"Error: {e}\nFavor de cerrar el archivo.")
    return wrapper

def open_file_selection(initialdir='', filter_name='*.*'):
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    filetypes = (
        (filter_name, f"*{filter_name.split(' ')[0]}*.*"),
        ('All files', '*.*'),
        ('Excel', '*.xlsx'),
        ('CSV', '*.doc')
    )
    files = fd.askopenfilenames(filetypes=filetypes, initialdir=initialdir)
    root.destroy()
    return files

def is_file_open(filepath):
    # Check if the file exists
    if not os.path.exists(filepath):
        return False  # File does not exist, so treat it as "open" for your logic

    try:
        # Try to open the file in write mode
        with open(filepath, 'a'):
            pass
        return False  # File is not open (no exception raised)
    except PermissionError:
        return True 
    
def select_directory(initialdir):
    """
    Opens a directory selection dialog and returns the selected directory path.
    
    :param initialdir: The initial directory to open in the dialog.
    :return: The selected directory path as a string, or an empty string if no directory is selected.
    """
    # Create a hidden root window
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)  # Ensure the dialog appears on top

    # Open the directory selection dialog
    directory = fd.askdirectory(initialdir=initialdir)

    # Destroy the root window after use
    root.destroy()
    return directory

def save_state_pickle(state, filename='folder_state_local_shipments.pkl'):
    with open(filename, 'wb') as f:
        pickle.dump(state, f)

def load_state_pickle(filename='folder_state_local_shipments.pkl'):
    try:
        with open(filename, 'rb') as f:
            return pickle.load(f)
    except FileNotFoundError:
        return {
            "folder_output": None,
            "selections": {},
            "fecha_mail": None,
            "fecha_shipments_elp": None,
            "fecha_freeze": None,
            "outlook_folder": None
        }

def set_paths(path):
    global output_paths
    output_paths = {}
    output_paths['path_attachments'] = os.path.join(path, "attachments")
    output_paths['path_attachments_done'] = os.path.join(output_paths['path_attachments'], "Done")
    output_paths['path_korrus_list'] = os.path.join(path, "Korrus_list.xlsx")
    output_paths['path_korrus_data'] = os.path.join(path, "Korrus_data.xlsx")
    output_paths['path_xl_format']=os.path.join(path,'columns and formatting.xlsx')
    output_paths['path_ship_cust']=os.path.join(path,'Shipped to Cust.xlsx')
    set_col_rel(output_paths)

def get_path(state, selector):
    path = state["selections"].get(selector, "Not selected")
    if not path or path == "Not selected":
        st.info(f"{selector} no seleccionado.")
    return path

# ----------------------------------------------------------------
# Dataframe management
# ----------------------------------------------------------------
def rename_columns(df, df_col_rel, table_from='std_name', sheet_from='only', table_to='std_name', sheet_to='only'):
    """
    Renames columns of a DataFrame to a standard name or between tables.

    Parameters:
    -----------
    df : pd.DataFrame
        The DataFrame whose columns need renaming.
    df_col_rel : pd.DataFrame
        DataFrame containing the column relationships with the columns:
        ['table', 'sheet', 'column_name', 'std_name'].
    table_from : str
        The source selector to map from.
    sheet_from : str
        The source sheet name.
    selector_to : str
        The target selector to map to.
    sheet_to : str
        The target sheet name.

    Returns:
    --------
    pd.DataFrame
        DataFrame with renamed columns.

    Raises:
    -------
    ValueError:
        If required columns are missing or if no matching mapping is found.
    """
    # Validate that df_col_rel contains all required columns
    required_cols = {'table', 'sheet', 'column_name', 'std_name'}
    missing_cols = required_cols - set(df_col_rel.columns)
    if missing_cols:
        raise ValueError(f"df_col_rel is missing required columns: {missing_cols}")
    
    mapping = {}

    if table_to == 'std_name':
        filtered = df_col_rel[(df_col_rel['table'] == table_from) & 
                              (df_col_rel['sheet'] == sheet_from)]
        if filtered.empty:
            raise ValueError("No matching mapping found for given table_from and sheet_from.")
        mapping = filtered.set_index('column_name')['std_name'].to_dict()
    
    elif table_from == 'std_name':
        filtered = df_col_rel[(df_col_rel['table'] == table_to) & 
                              (df_col_rel['sheet'] == sheet_to)]
        if filtered.empty:
            raise ValueError("No matching mapping found for given table_to and sheet_to.")
        mapping = filtered.set_index('std_name')['column_name'].to_dict()
    
    else:
        filtered_from = df_col_rel[(df_col_rel['table'] == table_from) & 
                                   (df_col_rel['sheet'] == sheet_from)]
        if filtered_from.empty:
            raise ValueError("No matching mapping found for given table_from and sheet_from.")
        mapping_from = filtered_from.set_index('column_name')['std_name'].to_dict()
        # First renaming step
        df = df.rename(columns=mapping_from)

        filtered_to = df_col_rel[(df_col_rel['table'] == table_to) & 
                                 (df_col_rel['sheet'] == sheet_to)]
        if filtered_to.empty:
            raise ValueError("No matching mapping found for given table_to and sheet_to.")
        mapping = filtered_to.set_index('std_name')['column_name'].to_dict()

    # Rename using the final mapping and return a new DataFrame
    return df.rename(columns=mapping) 

def format_dates(df, date_cols=[],type='iso'):
    for col in date_cols:
        if not col in df.columns:
            continue
        if type=='iso':
            df[col]=pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
        else:
            df[col]=pd.to_datetime(df[col], errors='coerce')
    return df

def move_columns_to_front(df, columns):
    """
    Move specified columns to the beginning of the DataFrame.
    """
    columns = [col for col in columns if col in df.columns]
    remaining_columns = [col for col in df.columns if col not in columns]
    return df[columns + remaining_columns]

def append_df_to_df(df_new=pd.DataFrame(),df_old=pd.DataFrame(),table='',keys=[],date_cols=[]):
    """
    Appends a dataframe to an existing Dataframe
    keys: Fields that should not be duplicated, the old records are kept
    table: Table which is source of the dataframe, just to show it in the error
    date_cols: Columns transformed to date
    """
    if df_new.empty:
        return
    if not keys:
        keys=df_new.columns
    df_new=df_new.copy()
    df_new=format_dates(df_new,date_cols)
    df_old=format_dates(df_old,date_cols)
    df_old['composite_key'] = list(zip(*(df_old[col] for col in keys)))
    df_new['composite_key'] = list(zip(*(df_new[col] for col in keys)))
    df_new=df_new[~df_new['composite_key'].isin(df_old['composite_key'])]
    df_old=pd.concat([df_old,df_new])  
    df_old.drop(columns=['composite_key'],inplace=True)
    df_old.reset_index(inplace=True,drop=True)
    df_grp=df_old.groupby(keys).count()
    df_grp=df_grp[df_grp[df_grp.columns[0]]>1]
    if len(df_grp)>0:
        df_grp.reset_index(inplace=True)
        st.error(f"Hay duplicados en el archivo {table} para la llave {keys}:",df_grp[keys].sort_values(keys))
        st.stop()
    return df_old

def update_dataframe(df_original,df_to_integrate,key_cols=[],exceptions=[]):
    """
    Update a dataframe with the values of another dataframe
    key_cols: an array of columns that will be used to match the rows between the dataframes
    exceptions: columns not to update, we keep values from old df
    data from the secod dataframe is appended if not found in the first one
    """
    df1_indexed = df_original.set_index(key_cols)
    df2_indexed = df_to_integrate.set_index(key_cols)

    # Update matching rows for common columns
    common_cols = df1_indexed.columns.intersection(df2_indexed.columns)
    common_cols = [col for col in common_cols if col not in exceptions]
    df1_indexed.update(df2_indexed[common_cols])

    # Append rows from df2 that are not in df1
    new_rows = df2_indexed.loc[~df2_indexed.index.isin(df1_indexed.index)]
    df_original = pd.concat([df1_indexed, new_rows]).reset_index()
    return df_original

def set_family(df,column,dest_col='Family'):
    """
    column: name of the column that contains the Model
    """
    df[column].fillna('',inplace=True)
    for key in families.keys():
        df.loc[df[column].str.startswith(tuple(families[key])), dest_col] = key
    df[dest_col].fillna('Other', inplace=True)
    return df

def check_duplicated_columns(df):
    duplicated = df.columns[df.columns.duplicated()].tolist()
    if duplicated:
        st.error(f"Columnas duplicadas: {duplicated}")
        st.stop()

def append_to_sheet(ws_dict, ws):
    """
    This will append the data that contains no value in the first _cell column
    It is important that the order of the header is preserved
    Data is appended at the end of the ws
    """
    df=ws_dict['df']
    cell_cols = [col for col in df.columns if ((not col is None) and (col.endswith('_cell')))]
    data_cols = list(ws_dict["header"].keys())
    df_to_append=df[df[cell_cols[0]].isnull()][data_cols]
    for i, r in enumerate(dataframe_to_rows(df_to_append, index=False, header=False)):
        ws.append(r)
    return ws

def update_column(ws_dict,ws,column=None):
    """
    Updates a column of the worksheet with the values on the dataframe
    """
    df=ws_dict['df'].copy()
    if not column:
        return ws
    df.dropna(subset=[f'{column}_cell'],inplace=True)
    for index,row in df[[column,f'{column}_cell']].iterrows():
        ws[row[f'{column}_cell']].value=row[column]
    return ws

def save_df(df, filepath, sheet_name='Sheet', index=False):
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index)

def check_mandatory_cols(cols,selector_name, raise_error=True):
    missing_columns = [col for col in mandatory_cols[selector_name] if col not in cols]
    if len(missing_columns)>0:
        st.error(f"No se encontraron las siguientes columnas en el archivo {selector_name}: {missing_columns}")
        if raise_error:
            st.stop()
        return False
    return True

def set_col_rel(output_paths):
    global df_col_rel,df_columns
    df_columns=read_excel(output_paths['path_xl_format'],sheet_name='column_format')
    df_col_rel=df_columns[~df_columns['std_name'].isnull()].copy()
# ----------------------------------------------------------------
# Excel Utility Functions
# ----------------------------------------------------------------
def load_excel_with_header_key(file_path, sheet_name=0, key_text='', dtype=None, **kwargs):
    df = read_excel(file_path,sheet_name=sheet_name, keep_default_na=False,dtype=dtype)
    header_row = None
    for col in df.columns:
        for i, cell in df[col].items():
            if pd.notna(cell) and key_text in str(cell):
                header_row = i
                break
        if header_row is not None:
            break
    if header_row is None:
        raise ValueError(f"Key text '{key_text}' not found in the sheet.")
    df.columns = df.iloc[header_row]
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    return df

def duplicate_df_cols(df, suffix, cols):
    subset = df[cols].copy()
    subset = subset.rename(columns=lambda x: x + suffix)
    cols_suffixed = [f"{x}{suffix}" for x in cols]
    # Remove any existing columns with the suffixed names before adding the new ones
    df = df.drop(columns=[col for col in cols_suffixed if col in df.columns], errors='ignore')
    df = pd.concat([df, subset], axis=1)
    return {'df': df, 
            'cols_suffixed': cols_suffixed}

def find_cell_by_text(ws, text):
    """
    Find the cell containing the specified date in the worksheet.
    
    :param ws: The worksheet object from openpyxl
    :param date_value: The date value to search for (datetime object or string)
    :return: The cell address (e.g., 'B2') or None if not found
    """
    # for row in ws.iter_rows():
    for row in ws[ws.calculate_dimension()]:
        for cell in row:
            if cell.value == text:
                return cell.coordinate  # Return cell address (e.g., 'B2')
    return None 

def read_excel(path=None,sheet_name=0,header=0,keep_default_na=True,dtype=None):
    with pd.ExcelFile(path) as xls:
        df=pd.read_excel(path,sheet_name=sheet_name,header=header,keep_default_na=keep_default_na,dtype=dtype)
    return df

def close_xl_if_open(path):
    if is_file_open(path):
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            workbook = excel.Workbooks(path)
            workbook.Save()
            workbook.Close()
        except:
            st.error(f"Cerrar el archivo: {path}")
            st.stop()   

def get_locations(df,ws,column):
    """
    Obtains all the cell locations in a worksheet (ws) for each text in a column of a dataframe
    Pass a dataframe with 
    """
    coord={}
    for text in df[column].drop_duplicates():
        coord[text]=find_cell_by_text(ws,text)
        coord
    return coord

def get_worksheet_df(ws, key_text=None,data_only=False):
    """
    This is similar to pd.read_excel, except that the dataframe includes data for the cell coordinates and properties.
    It takes time to read so it is used only for wb where we need to preserve data, cell format and structure.
    Returns a dictionary with a dataframe (including cell properties), header properties, formula properties
    """
    if key_text:
        header_row_idx = None
        for col in ws.iter_cols(min_row=ws.min_row, max_row=ws.max_row):
            for cell in col:
                if cell.value is not None and key_text in str(cell.value):
                    header_row_idx = cell.row
                    break
            if header_row_idx is not None:
                break

        if header_row_idx is None:
            st.error(f"No se encontró el header: {key_text}")
            st.stop()     
    else:
        header_row_idx=1
    header = [cell for cell in ws[header_row_idx]]
    num_cols = len(header)
    
    formula_dict = {}
    if not data_only:
        first_data_rows = list(ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 1, max_col=num_cols))
        if first_data_rows:
            first_data_row = first_data_rows[0]
            for i, cell in enumerate(first_data_row):
                col_name = header[i].value
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_dict[col_name] = cell.value


    data_rows = []
    cell_properties_dict={}
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, max_col=num_cols):
        row_dict = {}
        for i, cell in enumerate(row):
            col_name = header[i].value
            row_dict[col_name] = cell.value
            
            if data_only:
                row_dict[f"{col_name}_cell"] = cell.coordinate
            else:
                if (cell.fill.patternType) or ((cell.font.color) and (str(cell.font.color.rgb) !="Values must be of type <class 'str'>")):
                    cell_properties=get_cell_properties(cell)
                    cell_properties_dict[cell.row]={cell.coordinate:cell_properties}
                else:
                    cell_properties=None

                row_dict[f"{col_name}_cell"] = cell_properties
        data_rows.append(row_dict)
    df = pd.DataFrame(data_rows)
    
    header_dict = {}
    for cell in header:
        header_dict[cell.value] = {"coordinate": cell.coordinate,
                                   "properties": get_cell_properties(cell)}
    if len(df)==0:
        df=pd.DataFrame(columns=header_dict.keys())
    worksheed_df = {
        "df": df,
        "header": header_dict,
        "formulas": formula_dict,
        "properties":cell_properties_dict
    }
    return worksheed_df

def update_sheet(ws_dict, ws, apply_formats=True):
    """
    Updates the worksheet with the data and formats in the dataframe from the dictionariy
    """
    df = ws_dict["df"]
    header_info = ws_dict["header"]
    formulas = ws_dict.get("formulas", {})
    properties_info = ws_dict.get("properties", {})

    # Get header row number from first header coordinate
    first_header = next(iter(header_info.values()))
    _, header_row = coordinate_from_string(first_header["coordinate"])
    # Delete all rows below the header
    if ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)

    # Reorder dataframe columns:
    data_columns = [col for col in df.columns if col is not None and not col.endswith('_cell')]
    format_columns = [col for col in df.columns if col is not None and col.endswith('_cell')]
    header_columns = sorted(
        [col for col in header_info.keys() if col in data_columns],
        key=lambda x: column_index_from_string(coordinate_from_string(header_info[x]['coordinate'])[0])
    )
    remaining_columns = [col for col in data_columns if col not in header_columns]
    ordered_columns = header_columns + remaining_columns
    df_ordered=df.copy()
    df_ordered.reset_index(drop=True,inplace=True)
    df_ordered = df_ordered[ordered_columns]
    df_formats=df[format_columns]
    # Write dataframe starting below the header
    for col_index, header in enumerate(ordered_columns, start=1):
        ws.cell(row=header_row, column=col_index, value=header)
        
    start_data_row = header_row + 1
    for i, r in enumerate(dataframe_to_rows(df_ordered, index=False, header=False)):
        ws.append(r)
        if not apply_formats:
            continue
        formats=df_formats.loc[i,format_columns].dropna()
        for col in formats.keys():
            current_row = start_data_row + i
            cell = ws.cell(row=current_row, column=header_columns.index(col[0:-5])+1)
            format_cell(cell,formats[col])
    
    # Write formulas and copy them down using Translator
    for col, formula in formulas.items():
        origin_cell=ws.cell(start_data_row,ordered_columns.index(col)+1)
        origin_cell.value=formula
        last_cell=origin_cell.offset(i,0)
        cell_range=ws[origin_cell.offset(1,0).coordinate:last_cell.coordinate]
        for cell in cell_range:
            cell[0].value=Translator(formula, origin=origin_cell.coordinate).translate_formula(cell[0].coordinate)
    return ws

def create_new_sheet(wb,sheet_name='Sheet1',df=pd.DataFrame()):
    """
    Creates a new sheet in a workbook and appends a dataframe's data
    """
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws=wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    return wb

def get_cell_properties(cell):
    properties = {}
    fill = cell.fill
    properties["background_color"] = fill
    properties["font"]=cell.font
    properties["alignment"] = cell.alignment
    return properties

def get_xl_formatting(table_name=None):
    """
    Returns the formatting for the excel file defined in columns and formatting.xlsx
    return example
    {'oor':
    {
        'columna':{
            'head':{properties},
            'data':{properties}
        }
    }
    }
    """
    wb = load_workbook(output_paths['path_xl_format'])
    ws = wb['column_format']
    col_info=get_column_info(ws,'column_name')
    cell_properties={}
    for cell in col_info['data_range']:
        head_properties=get_cell_properties(cell)
        data_properties=get_cell_properties(cell.offset(0, 1))
        if cell.offset(0, -1).value not in cell_properties:
            cell_properties[cell.offset(0, -1).value]={} #Table name
        if cell.value not in cell_properties[cell.offset(0, -1).value]:
            cell_properties[cell.offset(0, -1).value][cell.value]= {} #Column name
        cell_properties[cell.offset(0, -1).value][cell.value]['head_properties']=head_properties
        cell_properties[cell.offset(0, -1).value][cell.value]['data_properties']=data_properties
    ws = wb['special_format']
    col_info=get_column_info(ws,'format_name')
    cell_properties['special_format']={}
    for cell in col_info['data_range']:
        cell_properties['special_format'][cell.value]=get_cell_properties(cell)

    if table_name:
        cell_properties=cell_properties[table_name]
    return cell_properties  

def get_column_info(ws, col_name, raise_error=True):
    """
    Returns a dictionary with the column cell and the data range for the column.
    """
    col_cell=find_cell_by_text(ws, col_name)
    if not col_cell:
        if not raise_error:
            return False
        st.error(f"Column '{col_name}' not found in the sheet.")
        st.stop()
    col_cell=ws[col_cell]
    last_cell=ws[ws.calculate_dimension()][-1][-1].coordinate
    data_range=ws[col_cell.offset(1,0).coordinate:ws.cell(ws[last_cell].row,col_cell.offset(1,0).column).coordinate]
    data_range_list=[cell[0] for cell in data_range]
    return {'data_range':data_range_list,
            'col_cell':col_cell}

def format_cell(cell,properties):
    "Apply properties based on the dict of prooperties"
    cell.fill=copy(properties['background_color'])
    cell.font = copy(properties["font"])
    cell.alignment = copy(properties["alignment"])

def format_on_change(zip_cols, ws, start_row=1, format1=None, format2=None):
    """
    Formats rows in a worksheet based on changes in values across zipped columns.

    Parameters:
    - zip_cols: zip of multiple columns (e.g., zip(col_info1, col_info2, ...)).
    - worksheet: The active worksheet to apply formatting.
    - start_row: Starting row number (default is 1).
    - format1: First alternating format (default yellow).
    - format2: Second alternating format (default light blue).
    """
    # Default styles
    if format1 is None:
        format1 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    if format2 is None:
        format2 = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue

    # Initialize tracking variables
    previous_values = None
    current_format = format1

    # Iterate through the zipped columns and their row numbers
    for row_idx, cells_rows in enumerate(zip_cols, start=start_row):
        # Check if the current row values differ from the previous row
        combined_values=""
        for cell in cells_rows:
            combined_values=combined_values+cell.value
        if previous_values != combined_values:
            # Alternate the format
            current_format = format1 if current_format == format2 else format2
            previous_values = combined_values

        # Apply the format to all cells in the current row from the zipped columns
        for col_idx, cell in enumerate(cells_rows):
            # cell = ws.cell(row=row_idx, column=col_idx + 1)
            format_cell(cell, current_format)

def format_xl_dates(wb,sheet_name='',date_columns=[]):
    """
    Set format to a date column in an excel worksheet
    """
    ws=wb[sheet_name]
    for col in date_columns:
        col_info=get_column_info(ws,col)
        for cell in col_info['data_range']:
            cell.number_format = 'mm/dd/yyyy'  
    return wb

def get_col_sizes(wb):
    """
    Obtains a dictionary with the column sizes of each sheet in a workbook
    """
    col_sizes={}
    for sheet in wb.sheetnames:
        col_sizes[sheet]=wb[sheet].column_dimensions
    return col_sizes

def get_table_styles(wb):
    """
    Obtains a dictionary with the table styles of each sheet in a workbook.

    Parameters:
    -----------
    wb : openpyxl.Workbook
        The loaded workbook object.

    Returns:
    --------
    dict
        A dictionary where the keys are sheet names and the values are lists of 
        (table_name, table_style) tuples for each table in the sheet.
    """
    table_styles = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        tables = ws._tables  # List of Table objects in the sheet

        table_info=None
        for table in tables:
            
            if hasattr(tables[table], "displayName"):  # Ensure it's an openpyxl Table object
                style_name = tables[table].tableStyleInfo.name if tables[table].tableStyleInfo else "No Style"
                table_info=style_name
                break
        if table_info:
            table_styles[sheet_name] = table_info  # Store tables for the sheet

    return table_styles


def apply_col_sizes(wb,col_sizes):
    """
    Applies the column sizes obtained by function get_col_sizes
    """
    for sheet in col_sizes.keys():
        if sheet not in wb.sheetnames:
            continue
        ws=wb[sheet]
        for column_letter in col_sizes[sheet].keys():
            if column_letter==0:
                continue
            ws.column_dimensions[column_letter].width=col_sizes[sheet][column_letter].width
    return wb
        

def apply_table_styles(wb, table_styles):
    """
    Applies table styles to the first table in each sheet based on a given dictionary.

    Parameters:
    -----------
    wb : openpyxl.Workbook
        The workbook object where the styles will be applied.
    
    table_styles : dict
        A dictionary where keys are sheet names and values are the desired table style names.
        Example: {'EDI Master': 'TableStyleMedium2'}
    
    Returns:
    --------
    None
    """
    for sheet_name, style_name in table_styles.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws._tables:
                continue
            table_range = ws.calculate_dimension()
            table_name = f"{sheet_name}_Table".replace(" ", "")
            table = Table(displayName=table_name, ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name=style_name,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
    return wb

@handle_permission_error_with_popup
def save_wb(wb, filepath):
    wb.save(filepath)
    wb.close()

def extract_selected_sheets(xlsx_file, sheets_to_keep, keep_original=True):
    """
    Extracts only the specified sheets from an XLSX file, removing all others,
    and saves a new file with "_reduced.xlsx" suffix. The removed sheets are
    stored separately for future reintegration. Also, makes a copy of the original
    extracted folder (temp_xlsx_original) to be used later in integration.
    
    Parameters:
      xlsx_file (str): Path to the original XLSX file.
      sheets_to_keep (list): List of sheet names to keep.
    
    Returns:
      dict: Relation between sheets and sheet names human readable
    """

    current_sheets=pd.ExcelFile(xlsx_file).sheet_names
    if set(sheets_to_keep) == set(current_sheets):
        return False

    # Get file directory and base name
    file_dir = os.path.dirname(xlsx_file)
    file_name = os.path.basename(xlsx_file).replace(".xlsx", "")
    # Temporary folder for extracted contents (working folder)
    temp_dir = os.path.join(file_dir, "temp_xlsx")
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir, exist_ok=True)
    # Step 1: Extract XLSX contents to temp directory
    with zipfile.ZipFile(xlsx_file, "r") as zip_ref:
        zip_ref.extractall(temp_dir)
    
    # Make a copy of the original extracted folder for later integration
    original_temp_dir = os.path.join(file_dir, file_name)
    if os.path.exists(original_temp_dir):
        shutil.rmtree(original_temp_dir.strip())
    shutil.copytree(temp_dir, original_temp_dir.strip())
    
    # Step 2: Parse xl/workbook.xml to get sheet name -> sheetX.xml mapping
    workbook_xml_path = os.path.join(temp_dir, "xl", "workbook.xml")
    tree = ET.parse(workbook_xml_path)
    root = tree.getroot()
    ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    sheets_elem = root.find("ns:sheets", ns)
    sheet_mapping = {}
    
    if sheets_elem is not None:
        for idx, sheet in enumerate(sheets_elem.findall("ns:sheet", ns), start=1):
            sheet_name = sheet.get("name")
            sheet_mapping[sheet_name] = f"sheet{idx}.xml"
    
    # Step 3: Separate sheets: move unwanted sheets to backup folder
    worksheets_dir = os.path.join(temp_dir, "xl", "worksheets")
    for sheet_name, sheet_xml_file in sheet_mapping.items():
        sheet_xml_path = os.path.join(worksheets_dir, sheet_xml_file)
        if sheet_name not in sheets_to_keep:
            if os.path.exists(sheet_xml_path):
                os.remove(sheet_xml_path)
    
    # Step 4: Repackage only the remaining sheets into the reduced XLSX file
    reduced_xlsx_file = os.path.join(file_dir, xlsx_file)
    with zipfile.ZipFile(reduced_xlsx_file, "w", zipfile.ZIP_DEFLATED) as new_zip:
        for root_dir, _, files in os.walk(temp_dir):
            for file in files:
                file_path = os.path.join(root_dir, file)
                arcname = os.path.relpath(file_path, temp_dir)
                new_zip.write(file_path, arcname)
    
    # Optionally, you can leave temp_dir intact or remove it here.
    if not keep_original:
        shutil.rmtree(original_temp_dir.strip())
    shutil.rmtree(temp_dir)
    return sheet_mapping

def integrate_modified_file_with_backup(modified_xlsx_file, sheets_mapping):
    """
    Integrates a modified _reduced.xlsx file with the original workbook.
    
    The function:
      1. Extracts the modified file into a temporary folder (temp_xlsx_reduced).
      2. Parses its xl/workbook.xml to build a mapping of human-readable sheet names to their sheet XML filenames.
      3. For each sheet in this mapping that is also present in the provided sheets_mapping (which maps the human-readable name 
         to the original sheet XML filename), it copies the corresponding modified sheet file from the modified folder to the 
         original folder (temp_xlsx_original) using the filename defined in sheets_mapping.
      4. Repackages the original folder into an integrated XLSX file.
    
    Parameters:
      modified_xlsx_file (str): Path to the modified _reduced.xlsx file.
      sheets_mapping (dict): Dictionary mapping human-readable sheet names to their original sheet XML filenames.
    
    Returns:
      str: Path to the integrated XLSX file.
    """
    file_dir = os.path.dirname(modified_xlsx_file)
    base_name = os.path.basename(modified_xlsx_file).replace(".xlsx", "")
    
    # Extract modified file into a temporary reduced folder
    temp_modified = os.path.join(file_dir, "temp_xlsx_reduced")
    if os.path.exists(temp_modified):
        shutil.rmtree(temp_modified)
    os.makedirs(temp_modified, exist_ok=True)
    with zipfile.ZipFile(modified_xlsx_file, "r") as zip_ref:
        zip_ref.extractall(temp_modified)
    
    # Parse the modified file's workbook.xml to build a mapping of human-readable sheet names to sheet XML filenames
    mod_workbook_xml_path = os.path.join(temp_modified, "xl", "workbook.xml")
    mod_tree = ET.parse(mod_workbook_xml_path)
    mod_root = mod_tree.getroot()
    mod_ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    mod_sheets_elem = mod_root.find("ns:sheets", mod_ns)
    mod_sheet_mapping = {}
    if mod_sheets_elem is not None:
        for idx, sheet in enumerate(mod_sheets_elem.findall("ns:sheet", mod_ns), start=1):
            sheet_name = sheet.get("name")
            sheet_file = f"sheet{idx}.xml"
            mod_sheet_mapping[sheet_name] = sheet_file
    
    # Define the original temp folder (created during extraction in extract_selected_sheets)
    temp_original = os.path.join(file_dir, base_name)
    worksheets_original = os.path.join(temp_original, "xl", "worksheets")
    worksheets_modified = os.path.join(temp_modified, "xl", "worksheets")
    
    # For each sheet in the modified file that is in the provided sheets_mapping,
    # copy the modified sheet XML file (using the modified file's naming) into the original folder
    # and rename it to match the original sheet XML filename from sheets_mapping.
    for sheet_name, mod_sheet_file in mod_sheet_mapping.items():
        if sheet_name in sheets_mapping:
            src_sheet = os.path.join(worksheets_modified, mod_sheet_file)
            dest_sheet = os.path.join(worksheets_original, sheets_mapping[sheet_name])
            if os.path.exists(src_sheet):
                shutil.copy(src_sheet, dest_sheet)

    # Repackage the integrated workbook from the original temp folder
    integrated_xlsx_file = os.path.join(file_dir, modified_xlsx_file)
    with zipfile.ZipFile(integrated_xlsx_file, "w", zipfile.ZIP_DEFLATED) as zip_out:
        for folder, _, files in os.walk(temp_original):
            for file in files:
                file_path = os.path.join(folder, file)
                arcname = os.path.relpath(file_path, temp_original)
                zip_out.write(file_path, arcname)
    
    # Cleanup the temporary modified folder
    shutil.rmtree(temp_modified)
# ----------------------------------------------------------------
# Hyperlink functions
# ----------------------------------------------------------------
# Get indexes to use in hyperlinks
def get_df_idx(df,idx_cols,idx_name='idx'):
    """
    Returns a dataframe indicating the row index where the data of a group of values start so it can be used for hyperlinks
    idx_cols=The key columns, duplicates are removed but the number of initial row is preserved
    idx_name=Name for the column containing the row numbers
    """
    df_idx=df.reset_index()
    df_idx.rename({'index':idx_name},axis=1,inplace=True)
    df_idx=df_idx.drop_duplicates(idx_cols,keep='first')
    df_idx=df_idx[idx_cols+[idx_name]]
    return df_idx

# Generate hyperlinks
def set_hyperlink(df,sheet_name,col_name,idx_name,typ='str'):
    """
    sheet_name: Destination sheet of the hyperlink
    col_name: Column name where the hyperlink will be inserted
    idx_name: Column name where the index is located
    """
    df=df.copy()
    idx=~df[idx_name].isnull()
    if typ=='str':
        quote='\"'
    else:
        quote=''
    df.loc[idx,col_name]=f"=HYPERLINK(\"#'{sheet_name}'!A"+(df.loc[idx,idx_name]+2).astype(int).astype(str)+f"\",{quote}"+(df.loc[idx,col_name]).astype(str)+f"{quote})"
    return df


# ----------------------------------------------------------------
# Manage File Selector (Streamlit)
# ----------------------------------------------------------------
def manage_file_selector(selector_key, display_label, state):
    if not state["selections"].get(selector_key):
        if st.button(f"{display_label} File", key=f"select_{selector_key}"):
            files = open_file_selection(initialdir=state["folder_output"] or os.getcwd())
            if files:
                state["selections"][selector_key] = files[0]
                save_state_pickle(state)
                st.rerun()
        st.info(f"{display_label} no seleccionado.")
    else:
        st.success(f"Selected {display_label} File: {state['selections'][selector_key]}")
        if st.button(f"Change {display_label} File", key=f"change_{selector_key}"):
            state["selections"][selector_key] = ""
            save_state_pickle(state)
            st.rerun()

def verify_selections(file_selectors):
    not_selected = []
    for file_selector in file_selectors:
        selector = file_selector.children[0].description[:-1]
        selected = file_selector.children[1].value.strip()
        if not os.path.exists(selected):
            file_selector.children[1].value = 'Not selected'
            selected = 'Not selected'
        # Assuming mandatory_selectors is defined globally
        if selector not in mandatory_selectors:
            continue
        if selected == 'Not selected':
            not_selected.append(selector)
    if len(not_selected) > 0:
        st.error(f"Favor de seleccionar los siguientes archivos: {not_selected}")
        st.stop()


# ----------------------------------------------------------------
# Dummy functions for the 5 process buttons
# ----------------------------------------------------------------
def generar_reportes():    
    # ## Generar reportes
    # - Hay tres reportes EDI Master, Shipped to Cust, Shipped to ELP
    # - Si hay archivos seleccionados se integran a estos reportes


    # ### Consolidar
    # - Korrus_data --> EDI Master
    # - InventoryStage --> Shipment to ELP
    # - Shipment transactions: Standalone file
    #
    # Consolidar reportes 
    msg_reportes=st.empty()
    msg_reportes=st.info("Generando reportes")
    path_ship_elp=get_path(state,'ELP Master')
    close_xl_if_open(path_ship_elp)
    path_oor_old=get_path(state,'OOR')
    close_xl_if_open(path_oor_old)
    wb_elp=load_workbook(path_ship_elp)
    ws_edi=wb_elp['EDI Master']
    ws_dict_edi=get_worksheet_df(ws_edi,key_text='PO',data_only=True)
    df_edi=ws_dict_edi['df']
    df_edi['ProductService ID']=df_edi['ProductService ID'].str.upper()
    df_edi['Quantity']=df_edi['Quantity'].astype(float)
    ws_ship_elp=wb_elp['Shipment to ELP']
    ws_dict_ship_elp=get_worksheet_df(ws_ship_elp,key_text='CUU ship Date',data_only=True)
    df_ship_elp=ws_dict_ship_elp['df']


    if not os.path.exists(output_paths['path_xl_format']):
        st.error("No se encuentra el archivo: columns and formatting.xlsx")
        st.stop()
    # Demanda del cliente
    if not os.path.exists(output_paths['path_korrus_data']):
        df=pd.DataFrame(columns=[mandatory_cols['Korrus']+['origin_file']])
        save_df(df,filepath=output_paths['path_korrus_data'],sheet_name='Korrus Data',index=False)

    df_korrus_data_new=pd.read_excel(output_paths['path_korrus_data'])
    if (len(df_korrus_data_new)>0):
        msg_reportes.info("Integrando Korrus data")
        df_korrus_data_new=df_korrus_data_new.loc[:, ~df_korrus_data_new.columns.str.startswith('Unnamed:')]
        df_korrus_data_new=df_korrus_data_new[~df_korrus_data_new['PurchaseOrder'].str.contains('---')]
        df_korrus_data_new['PODate']=pd.to_datetime(df_korrus_data_new['PODate'],format='mixed', errors='coerce')
        df_korrus_data_new['LineNumber']=df_korrus_data_new['LineNumber'].astype(int)
        df_korrus_data_new['AssignedDropZone'].fillna('NULL',inplace=True)
        df_korrus_data_new['ProductServiceID']=df_korrus_data_new['ProductServiceID'].str.upper()

        key_korrus=['PurchaseOrder','LineNumber','PODate','ProductServiceID','AssignedDropZone']
        df_korrus_data_new.drop_duplicates(subset=key_korrus,keep='last',inplace=True)
        
        # Eliminar registros no requeridos
        df_korrus_data_new['EDI Received']=df_korrus_data_new['origin_file'].str.extract(r'(\d{4}-\d{2}-\d{2})')
        df_korrus_data_new['EDI Received']=pd.to_datetime(df_korrus_data_new['EDI Received'],format='mixed', errors='coerce')
        df_korrus_data_new.drop('origin_file',axis=1,inplace=True)
        df_korrus_data_new.drop_duplicates(inplace=True)
        df_korrus_data_new['PODate']=pd.to_datetime(df_korrus_data_new['PODate'])
        df_korrus_data_new.rename({
            'PurchaseOrder':'PO',
            'ProductServiceID':'ProductService ID'
            },axis=1,inplace=True)
        df_korrus_data_new['CONCAT (PN+PO)']=df_korrus_data_new['ProductService ID']+df_korrus_data_new['PO']
        df_korrus_data_new['CONCAT (PN+PO+DZ)']=df_korrus_data_new['ProductService ID']+df_korrus_data_new['PO']+df_korrus_data_new['AssignedDropZone']
        edi_keys=['PO','LineNumber','ProductService ID','AssignedDropZone']
        df_edi=append_df_to_df(df_new=df_korrus_data_new,df_old=df_edi,table='EDI Master',keys=edi_keys)


    # Shipment transactions, lo embarcado al cliente
    path_ship_cust_new=get_path(state,'Shipment transactions')
    if path_ship_cust_new!='Not selected':
        msg_reportes.info("Integrando Shipments transactions")
        df_ship_cust_new=pd.read_excel(path_ship_cust_new)
        check_mandatory_cols(df_ship_cust_new.columns,'Shipment transactions')
        df_ship_cust_new=df_ship_cust_new[~df_ship_cust_new['Customer PO#'].isna()]
        save_df(df_ship_cust_new,output_paths['path_ship_cust'],sheet_name='Shipped to Cust',index=False)

    # InventoryStage, lo que se embarco a ELP 
    path_ship_elp_new=get_path(state,'InventoryStageBakup')
    if path_ship_elp_new!='Not selected':
        msg_reportes.info("Integrando InventoryStageBackup")
        df_ship_elp_new=read_excel(path_ship_elp_new)
        check_mandatory_cols(df_ship_elp_new.columns,'InventoryStageBakup')
        df_ship_elp_new=df_ship_elp_new[df_ship_elp_new['Cliente']!='Total']
        df_ship_elp_new=df_ship_elp_new[['Cliente','PO','Producto','Box Id','Cantidad']].ffill()
        df_ship_elp_new=df_ship_elp_new[~df_ship_elp_new['PO'].isna()]
        df=df_ship_elp_new['PO'].str.split('|', expand=True)
        df_ship_elp_new['DZ']=''
        if len(df.columns)>1:
            df_ship_elp_new['DZ']=df[1].str.strip()
            df_ship_elp_new['PO']=df[0].str.strip()
        df_ship_elp_new['Producto']=df_ship_elp_new['Producto'].str.upper()
        df_ship_elp_new['CUU ship Date']=state["fecha_shipments_elp"]
        df_ship_elp_new['CUU ship Date']=pd.to_datetime(df_ship_elp_new['CUU ship Date']).dt.strftime('%m/%d/%Y')
        df_ship_elp_new['BOX qty']=1
        df_ship_elp_new=df_ship_elp_new.groupby(['PO','Producto','DZ','Box Id']).agg({
            'Cantidad': 'sum',
            'BOX qty':'count',
            'CUU ship Date':'last'
        }).reset_index()

        df_ship_elp_new=rename_columns(df_ship_elp_new,df_col_rel,table_from='InventoryStageBakup',table_to='ELP Master',sheet_to='Shipment to ELP')
        df_ship_elp['DZ'].fillna('NULL',inplace=True)
        df_ship_elp_new['DZ'].fillna('NULL',inplace=True)
        df_ship_elp_new.loc[df_ship_elp_new['DZ']=='NA','DZ']='NULL'
        df_ship_elp=append_df_to_df(df_new=df_ship_elp_new,df_old=df_ship_elp,table='Shipment to ELP',keys=['PO','PN','BOX ID','DZ'])
        df_ship_elp['Family'].fillna('',inplace=True)
        df_ship_elp['CUU ship Date']=pd.to_datetime(df_ship_elp['CUU ship Date'], errors='coerce').dt.strftime('%m/%d/%Y')
        df_ship_elp=set_family(df_ship_elp,column='PN',dest_col='Family')

    # Ordenes Canceladas   
    msg_reportes.info("Integrando Ordenes canceladas")
    df_cancelled=read_excel(path_ship_elp,sheet_name='Cancelled Orders')
    df_cancelled=rename_columns(df_cancelled,df_col_rel,table_from='ELP Master',sheet_from='Cancelled Orders',table_to='ELP Master',sheet_to='EDI Master')
    df_cancelled=df_cancelled[['PO','ProductService ID','LineNumber']].drop_duplicates()
    df_cancelled['ProductService ID']=df_cancelled['ProductService ID'].str.upper()
    df_cancelled['status_cancelled']=True
    df_edi=df_edi.merge(df_cancelled,how='left',on=['PO','ProductService ID','LineNumber'])
    df_edi['Order/Line cancelled?']=''
    df_edi.loc[df_edi['status_cancelled']==True,'Order/Line cancelled?']='Cancelled'
    df_edi.drop(columns='status_cancelled',inplace=True)
    ws_dict_edi['df']=df_edi
    ws_edi=append_to_sheet(ws_dict_edi,ws_edi)
    ws_edi=update_column(ws_dict_edi,ws_edi,column='Order/Line cancelled?')
    ws_edi=update_column(ws_dict_edi,ws_edi,column='EDI Received')
    ws_dict_ship_elp['df']=df_ship_elp
    ws_ship_elp=append_to_sheet(ws_dict_ship_elp,ws_ship_elp)
    date_cols=df_columns[(df_columns['sheet']=='EDI Master')&(df_columns['data_type']=='date')]['column_name'].to_list()
    wb_elp=format_xl_dates(wb_elp,sheet_name='EDI Master',date_columns=date_cols)
    date_cols=df_columns[(df_columns['sheet']=='Shipment to ELP')&(df_columns['data_type']=='date')]['column_name'].to_list()
    wb_elp=format_xl_dates(wb_elp,sheet_name='Shipment to ELP',date_columns=date_cols)
    msg_reportes.info("Guardando Elp Master")

    save_wb(wb_elp,path_ship_elp)


    # Reporte de work orders, que se encuentra en proceso de produccion
    msg_reportes.info("Integrando tracker")
    path_tracker=get_path(state,'Tracker')
    xls = pd.ExcelFile(path_tracker)
    wo_sheets=[sheet for sheet in xls.sheet_names if (('Plan de produccion' in sheet) or ('TERMINADAS' in sheet))]
    df_wo=pd.DataFrame()
    for sheet in wo_sheets:
        df_wo=pd.concat([df_wo,pd.read_excel(path_tracker,sheet_name=sheet)])
        df_wo.dropna(subset=['PO cliente','Modelo'],inplace=True)
    check_mandatory_cols(df_wo.columns,'Tracker')
    df_wo=rename_columns(df_wo,df_col_rel,table_from='Tracker',sheet_from='Plan de produccion')
    df_wo['modelo']=df_wo['modelo'].str.upper()
    df_wo=format_dates(df_wo,['Date','START DATE', 'FINISH DATE', 'reprogrammed_cuu','SHIP DATE'])
    df_wo.rename(columns={"wo_qty":"quantity"},inplace=True)
    df_wo=move_columns_to_front(df_wo,['po','modelo','wo','quantity','START DATE', 'FINISH DATE', 'reprogrammed_cuu','estimated_move_date_cuu'])
    #% Edi
    df_edi=rename_columns(df_edi,df_col_rel,table_from='ELP Master',sheet_from='EDI Master')


    # Procesar envios al cliente eliminando cantidades negativas. Se conserva la fecha mas nueva de envio.

    if not os.path.exists(output_paths['path_ship_cust']):
        st.error('No hay envios al paso, integre al menos un Shipment Transactions')
        st.stop()
    df_ship_cust=read_excel(path=output_paths['path_ship_cust'])
    df_ship_cust=rename_columns(df_ship_cust,df_col_rel,table_from='Shipment transactions')
    df_ship_cust['modelo']=df_ship_cust['modelo'].str.upper()
    df_ship_cust_grp=df_ship_cust.groupby(['po','modelo']).sum('Qty. Shipped').reset_index()
    df_ship_cust_grp=df_ship_cust_grp[['po','modelo','quantity']]
    df_ship_cust_dates=df_ship_cust.sort_values(['po','modelo','shipment_date_cust']).drop_duplicates(subset=['po','modelo'],keep='last')
    df_ship_cust_dates.drop('quantity',axis=1,inplace=True)
    df_ship_cust_dates=df_ship_cust_dates.merge(df_ship_cust_grp,on=['po','modelo'])
    df_ship_cust_dates=df_ship_cust_dates[['po','modelo','quantity','shipment_date_cust']]


    # Envios a ELP
    df_ship_elp=rename_columns(df_ship_elp,df_col_rel,table_from='ELP Master',sheet_from='Shipment to ELP')
    df_ship_elp['modelo']=df_ship_elp['modelo'].str.upper()
    df_ship_elp['po']=df_ship_elp['po'].str.strip()
    df_ship_elp['modelo']=df_ship_elp['modelo'].str.strip()
    df_ship_elp['quantity'].fillna(0,inplace=True)
    df_ship_elp_grp=df_ship_elp.groupby(['po','modelo','shipment_date_elp']).sum('quantity').reset_index()
    df_ship_elp_grp=df_ship_elp_grp[['po','modelo','quantity','shipment_date_elp']]


    # Merge EDI con work orders y embarques

    df_edi=df_edi[df_edi['po_date']>pd.to_datetime(state["fecha_freeze"])]
    df_edi_po_dates=df_edi[['po','modelo','po_date']].copy()
    df_edi_po_dates.sort_values(['po_date'],inplace=True)
    df_edi_po_dates.drop_duplicates(['po','modelo'],keep='last',inplace=True)
    df_wo=df_wo.merge(df_edi_po_dates,how='left',on=['po','modelo'])
    df_wo=df_wo[df_wo['po_date']>pd.to_datetime(state["fecha_freeze"])]
    df_ship_cust_dates=df_ship_cust_dates.merge(df_edi_po_dates,how='left',on=['po','modelo'])
    df_ship_cust_dates=df_ship_cust_dates[df_ship_cust_dates['po_date']>pd.to_datetime(state["fecha_freeze"])]
    df_ship_elp=df_ship_elp.merge(df_edi_po_dates,how='left',on=['po','modelo'])
    df_ship_elp=df_ship_elp[df_ship_elp['po_date']>pd.to_datetime(state["fecha_freeze"])]

    msg_reportes.info("Asignando ordenes")
    assignments_wo=assign_quantities(df_pos=df_edi,df_to_assign=df_wo,additional_fields=['WO','START DATE','FINISH DATE','reprogrammed_cuu','estimated_move_date_cuu'])
    df_edi_combined=assignments_wo['df_pos']
    df_edi_combined.rename({'Assigned':'WO Qty'},axis=1,inplace=True)
    assignments_shp_cust=assign_quantities(df_pos=df_edi_combined,df_to_assign=df_ship_cust_dates,additional_fields=['shipment_date_cust'])
    df_edi_combined=assignments_shp_cust['df_pos']
    df_edi_combined.rename({'Assigned':'Shipped to Cust'},axis=1,inplace=True)
    assignments_shp_elp=assign_quantities(df_pos=df_edi_combined,df_to_assign=df_ship_elp,additional_fields=['shipment_date_elp'])
    df_edi_combined=assignments_shp_elp['df_pos']
    df_edi_combined.rename({'Assigned':'Shipped to Elp'},axis=1,inplace=True)

    # ### OOR Report

    # Add additional fields to EDI, only the record with last ship date, the detail is in Work orders sheet, Shipped to ELP and Shipped to cust reports

    df_assigned_wo = assignments_wo['df_assignments']
    df_edi_combined=merge_additional_fields(df_assigned_wo,
                            df_edi_combined,
                            fields=['po','modelo','LineNumber', 'START DATE', 'FINISH DATE','estimated_move_date_cuu','reprogrammed_cuu'],
                            sort_fields=['estimated_move_date_cuu'],
                            key=['po','modelo','LineNumber'])
    # Add WO field
    df_wo_qty=df_assigned_wo[['po','modelo','LineNumber','WO','AvailQuantity']]
    df_wo_qty['AvailQuantity']=df_wo_qty['AvailQuantity'].fillna(0)
    df_wo_qty['WO']=df_wo_qty['WO'].astype(int).astype(str) + ' TotQty: ' + df_wo_qty['AvailQuantity'].astype(int).astype(str) + '.'
    df_wo_qty=df_wo_qty.groupby(['po','modelo','LineNumber']).agg({
        'WO': ' '.join
    }).reset_index()
    df_edi_combined=merge_additional_fields(df_wo_qty,
                            df_edi_combined,
                            fields=['po','modelo','LineNumber', 'WO'],
                            sort_fields=['po'],
                            key=['po','modelo','LineNumber'])


    df_assigned_shp_elp=assignments_shp_elp['df_assignments']
    df_edi_combined=merge_additional_fields(df_assigned_shp_elp,
                            df_edi_combined,
                            fields=['po','modelo','LineNumber','shipment_date_elp'],
                            sort_fields=['shipment_date_elp'],
                            key=['po','modelo','LineNumber'])
    df_assigned_shp_cust=assignments_shp_cust['df_assignments']
    df_edi_combined=merge_additional_fields(df_assigned_shp_cust,
                            df_edi_combined,
                            fields=['po','modelo','LineNumber','shipment_date_cust'],
                            sort_fields=['shipment_date_cust'],
                            key=['po','modelo','LineNumber'])


    # Set production, and shipped STATUS
    df_edi_combined[['quantity','WO Qty','Shipped to Elp','Shipped to Cust']].fillna(0,inplace=True)
    df_edi_combined['quantity']=df_edi_combined['quantity'].astype(float)
    df_edi_combined['Shipped to Elp']=df_edi_combined['Shipped to Elp'].astype(float)
    df_edi_combined['Shipped to Cust']=df_edi_combined['Shipped to Cust'].astype(float)
    df_edi_combined.loc[(df_edi_combined['quantity']-df_edi_combined['Shipped to Elp']<=0),'Movement Status']='Moved'
    df_edi_combined.loc[(df_edi_combined['quantity']-df_edi_combined['Shipped to Cust']<=0),'Status']='Shipped'
    cols=['quantity','WO Qty','Shipped to Elp','Shipped to Cust']
    df_po_status_cls=df_edi_combined[['po']+cols]
    for col in cols:
        df_po_status_cls[col]=df_po_status_cls[col].astype(float)
    df_po_status_cls=df_po_status_cls.groupby(['po']).sum()
    df_po_status_cls.reset_index(inplace=True)


    df_po_status=df_po_status_cls[['po','quantity','WO Qty','Shipped to Elp','Shipped to Cust']]
    df_edi_combined.sort_values(['po','modelo','LineNumber','po_date'],inplace=True)
    df_edi_combined.reset_index(drop=True,inplace=True)
    df_edi_combined['Balance to ship']=df_edi_combined['quantity']-df_edi_combined['Shipped to Elp']

    df_po_status_cls=df_po_status_cls.loc[(df_po_status_cls['quantity']-df_po_status_cls['Shipped to Cust'])<=0,['po']]
    df_po_dates=df_assigned_shp_cust.sort_values(['po','shipment_date_cust']).drop_duplicates('po',keep='last')[['po','shipment_date_cust']]
    df_po_status_cls=df_po_status_cls.merge(df_po_dates,how='left',on=['po'])
    df_po_status_cls.rename(columns={'shipment_date_cust':'\nPO closure date'},inplace=True)
    df_edi_combined=df_edi_combined.merge(df_po_status_cls,how='left',on=['po'])
    df_edi_combined=set_family(df_edi_combined,column='modelo',dest_col='family')
    df_po_type=df_edi_combined[['po','family']].drop_duplicates()
    df_po_type.loc[df_po_type['family']=='Accesories','PO TYPE']='Accesories only'
    df_po_type.loc[(~(df_po_type['family']=='Accesories'),'PO TYPE')]='Fixtures only'
    df_po_type=df_po_type.drop_duplicates(['po','PO TYPE']).sort_values(['po','PO TYPE'])
    df_po_type=df_po_type.groupby('po', as_index=False).agg({'PO TYPE': '/'.join})
    df_po_type.loc[df_po_type['PO TYPE']=='Accesories only/Fixtures only','PO TYPE']='Both types'
    df_edi_combined=df_edi_combined.merge(df_po_type,how='left',on='po')

    df_edi_combined.loc[df_edi_combined['Order/Line cancelled?']=='Cancelled','Status']='Cancelled'

    if not 'edi_received' in df_edi_combined.columns:
        df_edi_combined['edi_received']=''



    # Read customer shipments and shipments to elp
    df_ship_cust=read_excel(output_paths['path_ship_cust'])
    df_ship_cust=rename_columns(df_ship_cust,df_col_rel,table_from='Shipment transactions')
    df_ship_cust['modelo']=df_ship_cust['modelo'].str.upper()
    df_ship_cust.sort_values(['po','modelo','shipment_date_cust'],inplace=True)
    df_ship_cust.reset_index(drop=True,inplace=True)
    df_ship_cust=format_dates(df_ship_cust,['shipment_date_cust'])
    df_ship_cust=move_columns_to_front(df_ship_cust,['po','modelo','shipment_date_cust','Quantity'])
    df_ship_elp=rename_columns(df_ship_elp,df_col_rel,table_from='ELP Master',sheet_from='Shipment to ELP')
    df_ship_elp['modelo']=df_ship_elp['modelo'].str.upper()
    df_ship_elp.sort_values(['po','modelo','shipment_date_elp'],inplace=True)
    df_ship_elp.reset_index(drop=True,inplace=True)
    df_ship_elp.loc[df_ship_elp['dz']=='','dz']='NULL'
    df_ship_elp['dz'].fillna('NULL',inplace=True)
    df_ship_elp=format_dates(df_ship_elp,['ShipmentDate'])
    df_ship_elp=move_columns_to_front(df_ship_elp,['po','modelo','ShipmentDate','Quantity'])
    # Get index for hyperlinks

    df_wo.reset_index(inplace=True,drop=True)
    df_ship_elp.reset_index(inplace=True,drop=True)
    df_ship_cust.reset_index(inplace=True,drop=True)
    df_wo_idx=get_df_idx(df_wo,idx_cols=['po','modelo'],idx_name='idx_wo')
    df_ship_elp_idx=get_df_idx(df_ship_elp,idx_cols=['po','modelo'],idx_name='idx_elp')
    df_ship_cust_idx=get_df_idx(df_ship_cust,idx_cols=['po','modelo'],idx_name='idx_cust')

    df_edi_combined=df_edi_combined.merge(df_wo_idx,how='left',on=['po','modelo'])
    df_edi_combined=df_edi_combined.merge(df_ship_elp_idx,how='left',on=['po','modelo'])
    df_edi_combined=df_edi_combined.merge(df_ship_cust_idx,how='left',on=['po','modelo'])

    # Get Hyperlinks
    df_edi_combined=set_hyperlink(df_edi_combined,sheet_name='WorkOrder Detail',col_name='WO',idx_name='idx_wo')
    df_edi_combined=set_hyperlink(df_edi_combined,sheet_name='Shipped to Cust',col_name='Shipped to Cust',idx_name='idx_cust',typ='int')
    df_edi_combined=set_hyperlink(df_edi_combined,sheet_name='Shipped to Elp',col_name='Shipped to Elp',idx_name='idx_elp',typ='int')

    df_edi_combined['WO'].fillna('',inplace=True)

    # Get prices if selected
    path_prices=get_path(state,'Prices')
    df_prices=None
    if path_prices!='Not selected':
        msg_reportes.info("Integrando precios")
        df_prices=read_excel(path_prices)
        df_prices=rename_columns(df_prices,df_col_rel,table_from='Prices',table_to='OOR Report',sheet_to='OOR')
        df_prices=df_prices[['ProductServiceID','Price']]
        df_prices.drop_duplicates(['ProductServiceID'],keep='last',inplace=True)

    # ### Actualizar OOR con datos nuevos

    st.info("Actualizando OOR")
    # Actualizar OOR
    df_oor=rename_columns(df_edi_combined,df_col_rel,table_from='EDI Combined',table_to='OOR Report',sheet_to='OOR')

    df_oor=set_family(df_oor,'ProductServiceID')
    oor_cols=df_columns[df_columns['table']=='OOR Report']['column_name'].to_list()
    # Calculo de valores por columna
    check_duplicated_columns(df_oor)
    df_oor=format_dates(df_oor,['START DATE','FINISH DATE','SHIP DATE','EDI Received','POUS Date','Actual Ship date to Customer\n','Actual Move date from CUU'],type='')
    df_oor.loc[df_oor['AssignedDropZone']=='','AssignedDropZone']='NULL'
    df_oor['AssignedDropZone'].fillna('NULL',inplace=True)
    df_oor['NP+PO+DZ']=df_oor['ProductServiceID']+df_oor['PurchaseOrder']+df_oor['AssignedDropZone'].astype(str)
    df_oor['CONCAT\nPN+PO']=df_oor['ProductServiceID']+df_oor['PurchaseOrder']

    oor_current_cols=[x for x in oor_cols if x in df_oor]

    df_oor=df_oor[oor_current_cols]
    df_oor.sort_values(['Family ','PurchaseOrder','ProductServiceID','LineNumber'],inplace=True)
    df_oor['PO Aging']='' # Formula
    df_oor['Aging Category']='' # Formula
    df_oor['Balance to move']='' # Formula
    # Filtrar segun la seleccion de fecha
    df_oor=df_oor[df_oor['POUS Date'] >= pd.to_datetime(state["fecha_freeze"])]

    #Integrar datos al OOR Anterior
    path_oor_old=get_path(state,'OOR')

    close_xl_if_open(path_oor_old)
    # sheet_rel=extract_selected_sheets(path_oor_old,sheets_to_keep,keep_original=False)
    wb_oor_old=load_workbook(path_oor_old)

    ws_oor_old=wb_oor_old['OOR']
    dict_oor_old=get_worksheet_df(ws_oor_old,'Family')
    df_oor_old=dict_oor_old['df']
    check_mandatory_cols(df_oor_old.columns,'OOR')
    # Part of the OOR that will be updated, we will remove duplicates from this part
    key_cols=['PurchaseOrder','ProductServiceID','LineNumber','AssignedDropZone']
    df_edi_rec_dates=df_oor_old[key_cols+['EDI Received']].drop_duplicates(key_cols).copy()

    df_oor_to_update=df_oor_old[df_oor_old['POUS Date'] >= pd.to_datetime(state["fecha_freeze"])]
    # This part will be kept as is, duplicates are allowed
    df_oor_old=df_oor_old[df_oor_old['POUS Date'] < pd.to_datetime(state["fecha_freeze"])]
    except_columns=df_columns[(~df_columns['user_column'].isna())&(df_columns['sheet']=='OOR')]['column_name'].to_list()
    df_oor_to_update.drop_duplicates(subset=key_cols,inplace=True)
    if not 'Comment' in df_oor.columns:
        df_oor['Comment']=''
    df_oor[['Comment','Status']]=df_oor[['Comment','Status']].fillna('')
    
    df_oor_to_update=update_dataframe(df_oor_to_update,df_oor.fillna(0),key_cols,exceptions=except_columns)
    df_oor_old=pd.concat([df_oor_old,df_oor_to_update])
    df_oor_old.reset_index(drop=True,inplace=True)
    path_oh_max=get_path(state,'OH Max')

    if path_oh_max!="Not selected":
        # Se actualiza el OH Max para todo el workbook ya que la formula toma en cuenta los embarcados y duplicados
        df_oh_max=read_excel(path_oh_max)
        df_oh_max=rename_columns(df_oh_max,df_col_rel,table_from='OH Max',sheet_from='only',table_to='OOR Report',sheet_to='OOR')
        df_oh_max=df_oh_max.groupby(['ProductServiceID']).sum('OH MAX')
        df_oh_max.reset_index(inplace=True)
        if 'OH MAX' in df_oor_old.columns:
            df_oor_old.drop(columns='OH MAX',inplace=True)
        df_oor_old=df_oor_old.merge(df_oh_max,how='left',on='ProductServiceID')
        df_oor_old['OH MAX'].fillna(0,inplace=True)
    df_oor_old=df_oor_old.merge(df_edi_rec_dates,how='left',on=key_cols,suffixes=('', '_new'))
    df_oor_old.loc[df_oor_old['EDI Received'].isna(),'EDI Received']=df_oor_old.loc[df_oor_old['EDI Received'].isna(),'EDI Received_new']
    df_oor_old.loc[df_oor_old['EDI Received']==0,'EDI Received']=df_oor_old.loc[df_oor_old['EDI Received']==0,'EDI Received_new']
    df_oor_old.drop(columns='EDI Received_new',inplace=True)
    dict_oor_old['df']=df_oor_old
    
    if len(df_prices)>0:
        # if 'Price' in df_oor_old.columns:
        #     df_oor_old.drop(columns=['Price'],inplace=True)
        df_oor_old=df_oor_old.merge(df_prices,how='left',on=['ProductServiceID'])

    ws_oor_old=update_sheet(dict_oor_old,ws_oor_old)
    # Value A1=1 is just to check later if formulas are evaluated
    ws_oor_old['A1'].value="=1"

    # Put date in oh max column

    cell=find_cell_by_text(ws_oor_old,'OH MAX')
    if cell:
        cell=ws_oor_old[cell]
        cell.offset(-1,0).value=datetime.now().date()

    # Formato excel

    wb_oor={}
    wb_oor_old=create_new_sheet(wb_oor_old,sheet_name='WorkOrder Detail',df=df_wo)
    wb_oor_old=create_new_sheet(wb_oor_old,sheet_name='Shipped to Elp',df=df_ship_elp)
    wb_oor_old=create_new_sheet(wb_oor_old,sheet_name='Shipped to Cust',df=df_ship_cust)


    #Apply autofilter and freeze panes
    wb_oor_old['WorkOrder Detail'].auto_filter.ref = wb_oor_old['WorkOrder Detail'].dimensions
    wb_oor_old['WorkOrder Detail'].freeze_panes = 'A2'
    wb_oor_old['Shipped to Elp'].auto_filter.ref = wb_oor_old['Shipped to Elp'].dimensions
    wb_oor_old['Shipped to Elp'].freeze_panes = 'A2'
    wb_oor_old['Shipped to Cust'].auto_filter.ref = wb_oor_old['Shipped to Cust'].dimensions
    wb_oor_old['Shipped to Cust'].freeze_panes = 'A2'


    # Format row when key values change
    dict_special_formats=get_xl_formatting('special_format')
    col_info1=get_column_info(ws_oor_old,'PurchaseOrder')['data_range']
    col_info2=get_column_info(ws_oor_old,'ProductServiceID')['data_range']
    format_on_change(zip(col_info1,col_info2),ws_oor_old,start_row=1,format1=dict_special_formats['on_change_a'],format2=dict_special_formats['on_change_b'])

    # Format dates
    wb_oor_old=format_xl_dates(wb_oor_old,sheet_name='OOR',date_columns=['POUS Date',
                                                        'EDI Received',
                                                        '\nPO closure date',
                                                        'Actual Move date from CUU',
                                                        'Actual Ship date to Customer\n',
                                                        'Estimated \nMOVE DATE\nCUU',
                                                        'Reprogrammed finish date CUU',
                                                        'START DATE',
                                                        'FINISH DATE',
                                                        'Actual Ship date to Customer\n'])



    ws=wb_oor_old['WorkOrder Detail']
    col_info1=get_column_info(ws,'po')['data_range']
    col_info2=get_column_info(ws,'modelo')['data_range']
    format_on_change(zip(col_info1,col_info2),ws,start_row=1,format1=dict_special_formats['on_change_a'],format2=dict_special_formats['on_change_b'])


    ws=wb_oor_old['Shipped to Elp']
    col_info1=get_column_info(ws,'po')['data_range']
    col_info2=get_column_info(ws,'modelo')['data_range']
    format_on_change(zip(col_info1,col_info2),ws,start_row=1,format1=dict_special_formats['on_change_a'],format2=dict_special_formats['on_change_b'])

    ws=wb_oor_old['Shipped to Cust']
    col_info1=get_column_info(ws,'po')['data_range']
    col_info2=get_column_info(ws,'modelo')['data_range']
    format_on_change(zip(col_info1,col_info2),ws,start_row=1,format1=dict_special_formats['on_change_a'],format2=dict_special_formats['on_change_b'])

    save_wb(wb_oor_old,path_oor_old)
    os.startfile(path_oor_old)

def fill_aar():
    st.write("Ejecutando función: AAR")
    #Yield reports
    path_oor_old=get_path(state,'OOR')
    folder_yield=state.get('folder_yield', "Not selected")

    if folder_yield=="Not selected":
        st.error(f'Favor de seleccionar los folder con Yield reports')
        st.stop()
    yield_files_lst=os.listdir(folder_yield)
    df_yield=pd.DataFrame()
    for file in yield_files_lst:
        df=read_excel(os.path.join(folder_yield,file))
        df=map_yield_report(df)
        df['file_name']=file
        df_yield=pd.concat([df,df_yield])


    dict_xl_fields={"ENSAMBLE DE BISAGRA Y TAPAS":"Complete mechanical Assy output",
    "ANGLE/DIMMING TEST":"Angle/Dimming test output",
    #  "DIMMING TEST":"Angle/Dimming test output",
    "ENSAMBLE DEL BRACKET":"Mechanical Assy"}

    for process in df_yield['Process'].drop_duplicates():
        if not process in dict_xl_fields.keys():
            continue
        df_yield.loc[df_yield['Process']==process,'xl_field']=dict_xl_fields[process]


    df_yield=df_yield.groupby(['date_from','xl_field']).sum('Total Tested').reset_index()
    df_yield=df_yield[['date_from','xl_field','Total Tested']]
    df_yield['date_from']=pd.to_datetime(df_yield['date_from']).dt.normalize()
    # Llenar el formato
    search_parms={
        "column":'date_from',
        "row":'xl_field',
        "offset":
                {"Category":
                {
                "scheduled":(0,0),
                "real":(1,0)
                },
                "Shift":
                {"1s":(0,0),
                "2s":(0,2)},
                }
    }

    wb=load_workbook(path_oor_old)
    wb=fill_yield_report(df_yield,wb,sheet_name='Trov Daily Status',search_parms=search_parms)
    wb=fill_yield_report(df_yield,wb,sheet_name='Rise Daily Status',search_parms=search_parms)
    wb.save(path_oor_old)
    os.startfile(path_oor_old)

def gating_parts():
    st.write("Ejecutando función: Gating Parts")
    # Gating parts
    dict_special_formats=get_xl_formatting('special_format')
    format_more_than_one_value=dict_special_formats['more_than_one_value']
    neutral_format=dict_special_formats['neutral_format']
    path_gating=get_path(state,'Gating Parts')
    if path_gating=="Not selected":
        st.error("Seleccionar reporte de Gating Parts")
        st.stop()
    dict_gating=read_excel(path_gating,sheet_name=None)
    path_oor=get_path(state,'OOR')
    close_xl_if_open(path_oor)
    wb_oor=load_workbook(path_oor)

    ws_oor=wb_oor['OOR']
    dict_oor=get_worksheet_df(ws_oor,'Family',data_only=True)
    df_oor=rename_columns(dict_oor['df'],df_col_rel,table_from='OOR Report',sheet_from='OOR')

    # Ficture Gating part and ETA
    df_ready=dict_gating['Ready'].drop_duplicates()
    df_ready['rdy']='CTB'
    df_ready=rename_columns(df_ready,df_col_rel,table_from='Gating Parts',sheet_from='Ready')

    if 'rdy' in df_oor.columns:
        df_oor.drop(columns=['rdy'],inplace=True)
    df_oor=df_oor.merge(df_ready[['po','modelo','rdy']],how='left',on=['po','modelo'])
    df_oor.loc[df_oor['rdy']=='CTB','fixture_gating_part']='CTB'
    df_short_detail=rename_columns(dict_gating['Shorts'],df_col_rel,table_from='Gating Parts',sheet_from='Shorts')
    df_short_detail.sort_values(['po','modelo','gating_due'],inplace=True)

    df_short=df_short_detail.drop_duplicates(['po','modelo'],keep='last').copy()
    df_short=df_short[['po','modelo','component','gating_due']]
    # More than one short gating part for a line, we keep the last line to keep the last date
    df_short=df_short_detail.copy()
    df_short=df_short[['po','modelo','component','gating_due']]
    df_short['plus_one']=False
    df_short.loc[df_short.duplicated(['po','modelo'],keep=False),'plus_one']=True
    df_short=df_short.drop_duplicates(['po','modelo'],keep='last')

    df_oor=df_oor.merge(df_short,how='left',on=['po','modelo'])
    df_oor['fixture_gating_part']=''
    df_oor.loc[~df_oor['component'].isnull(),'fixture_gating_part']=df_oor.loc[~df_oor['component'].isnull(),'component']
    df_oor.loc[~df_oor['component'].isnull(),'fixt_gp_eta']=df_oor.loc[~df_oor['component'].isnull(),'gating_due']

    # Short of fixture parts, not accessories
    df_short_detail.sort_values(['po','modelo','component','short_qty','gating_due','arrival_qty'],inplace=True)
    df_short_detail.reset_index(inplace=True,drop=True)
    create_new_sheet(wb_oor,sheet_name='Short Detail',df=df_short_detail)
    df_idx_short=get_df_idx(df_short_detail,idx_cols=['po','modelo'],idx_name='idx_short')
    df_oor=df_oor.merge(df_idx_short,how='left',on=['po','modelo'])
    df_oor=set_hyperlink(df_oor,sheet_name='Short Detail',col_name='fixture_gating_part',idx_name='idx_short')
    df_arrivals=dict_gating['Arrivals']
    df_arrivals=rename_columns(df_arrivals,df_col_rel,table_from='Gating Parts',sheet_from='Arrivals')
    df_arrivals.sort_values(['modelo','arrival_due'], inplace=True)
    df_arrivals['Cumulative Sum'] = df_arrivals.groupby(['modelo'])['arrival_qty'].cumsum()

    df_acc_shorts=df_oor[(df_oor['accessory_gating_part']=='Acc Short')&
                        (df_oor['oor_status'].str.lower()!='cancelled')&
                        (df_oor['family'].str[0:5].str.lower()=='acces')]
    df_acc_shorts['Cumulative Sum'] = df_acc_shorts.groupby(['modelo'])['quantity'].cumsum()
    df_arrivals['Cumulative Sum']=df_arrivals['Cumulative Sum'].astype(int)
    df_acc_shorts['Cumulative Sum']=df_acc_shorts['Cumulative Sum'].astype(int)

    df_acc_shorts = pd.merge_asof(
        df_acc_shorts.sort_values('Cumulative Sum'),
        df_arrivals.sort_values('Cumulative Sum'),
        left_on='Cumulative Sum',
        right_on='Cumulative Sum',
        by='modelo',
        direction='forward'
    )
    df_acc_shorts.dropna(subset=['po','modelo','LineNumber','arrival_due','quantity','arrival_qty'],inplace=True)
    df_acc_shorts=df_acc_shorts[['po','modelo','LineNumber','arrival_due','quantity','arrival_qty']]
    df_acc_shorts.drop_duplicates(subset=['po','modelo','LineNumber'],inplace=True,keep='last')
    if 'arrival_due' in df_oor.columns:
        df_oor.drop(columns=['arrival_due'])
    df_oor['acc_gp_eta']=''
    df_oor=df_oor.merge(df_acc_shorts[['po','modelo','LineNumber','arrival_due']],how='left',on=['po','modelo','LineNumber'])

    ws_oor=wb_oor['OOR']
    for index,row in df_oor[['fixture_gating_part','fixture_gating_part_cell','fixt_gp_eta','fixt_gp_eta_cell','plus_one','acc_gp_eta_cell','arrival_due']].iterrows():
        cell=ws_oor[row['fixture_gating_part_cell']]
        cell.value=row['fixture_gating_part']
        if row['plus_one']==True:    
            format_cell(cell,format_more_than_one_value)
        else:
            format_cell(cell,neutral_format)
        cell=ws_oor[row['fixt_gp_eta_cell']]
        cell.value=row['fixt_gp_eta']
        cell.number_format = 'mm/dd/yyyy'
        cell=ws_oor[row['acc_gp_eta_cell']]
        cell.value=row['arrival_due']
    # Value A1=1 is just to check later if formulas are evaluated
    ws_oor['A1'].value="=1"
    save_wb(wb_oor,path_oor)
    os.startfile(path_oor)


def actualizar_status():
    st.write("Ejecutando función: Actualizar Status")
    # Actualizar Status

    # Generar grupos de filtros por PO
    # Columnas tipo texto se hace un .join ordenado
    # Columnas tipo fecha se toma la ultima fecha
    # Columnas numericas se suman

    path_oor=get_path(state,'OOR')
    sheet_rel=extract_selected_sheets(path_oor,sheets_to_keep,keep_original=False)
    # df_oor=load_excel_with_header_key(path_oor,sheet_name='OOR',key_text='Family ', dtype=str)
    wb_oor=load_workbook(path_oor,data_only=True)
    ws_oor=wb_oor['OOR']
    if ws_oor['A1'].value is None:
        st.error(f"Favor de Guardar el archivo {path_oor}")
        st.stop()

    dict_oor=get_worksheet_df(ws_oor,'Family ',data_only=True)
    df_oor=dict_oor['df']
    df_oor_stat=rename_columns(df_oor,df_col_rel,table_from='OOR Report',sheet_from='OOR')
    df_oor_stat=df_oor_stat.copy()

    # Tipo de filtros
    df_oor_cols=df_columns[df_columns['sheet']=='OOR']

    all_filters=df_oor_cols.loc[(~df_oor_cols['filter'].isna()),'std_name'].drop_duplicates().to_list()

    # Tipo has_value
    has_val_cols=df_oor_cols.loc[(df_oor_cols['filter']=='has_value'),'std_name'].drop_duplicates().to_list()
    for col in has_val_cols:
        df_oor_stat.loc[(df_oor_stat[col]=='') | (df_oor_stat[col].isna()),col]='blank'
        df_oor_stat.loc[(df_oor_stat[col]!='') & (~df_oor_stat[col].isna()) & (df_oor_stat[col]!='blank'),col]='has_value'


    # Tipo lista:
    list_cols=df_oor_cols.loc[(df_oor_cols['filter']=='list'),'std_name'].drop_duplicates().to_list()
    for col in list_cols:
        df_oor_stat.loc[(df_oor_stat[col]=='') | (df_oor_stat[col].isna()),col]='blank'
        df_oor_stat[col]=df_oor_stat[col].astype(str).str.lower()
    list_cols=list_cols+has_val_cols

    # Tipo fecha
    date_cols=df_oor_cols.loc[(df_oor_cols['filter']=='date'),'std_name'].drop_duplicates().to_list()
    for col in date_cols:
        df_oor_stat[col] = pd.to_datetime(df_oor_stat[col], errors='coerce')
        df_oor_stat[col]=df_oor_stat[col].fillna(pd.Timestamp('2099-12-31'))
    date_cols_first=[f"{x}_first" for x in date_cols]
    date_cols_last=[f"{x}_last" for x in date_cols]

    # Tipo cantidad
    qty_cols=df_oor_cols.loc[(df_oor_cols['filter']=='qty'),'std_name'].drop_duplicates().to_list()


    # Key
    key_cols=['po']

    # Grupos de propiedades
    # 1. Copiar original df_oor_stat
    # 2. Aplicar filtros
    # 3. Agregar fechas first, last
    # 4. Aggregate
    # 4. Merge

    # Agregar valor in plan, existe como columna ficticia
    df_oor_stat['in_plan']='blank'
    df_oor_stat.loc[(df_oor_stat['balance_to_move']==0)|(df_oor_stat['wo']=='has_value'),'in_plan']='yes'
    if not 'in_plan' in list_cols:
        list_cols.append('in_plan')

    # Agregar valor short, existe como columna ficticia
    df_oor_stat['acc_short']='blank'
    df_oor_stat.loc[(df_oor_stat['acc_gp_eta']>pd.to_datetime(date.today()+timedelta(days=15)))&(df_oor_stat['accessory_gating_part']=='acc short'),'acc_short']='yes'
    if not 'acc_short' in list_cols:
        list_cols.append('acc_short')

    ## Propiedades de PO (Status diferente de Shipped o Cancelled)
    df_oor_by_po_active=df_oor_stat.copy()
    df_oor_by_po_active=df_oor_by_po_active[~df_oor_by_po_active['oor_status'].str.lower().str.strip().isin(["shipped","cancelled"])]
    df_oor_by_po_active=apply_grouping(df_oor_by_po_active,prefix='po_active_all_',key_cols=key_cols,list_cols=list_cols,qty_cols=qty_cols,date_cols=date_cols)
    # Propiedades de accesorios
    df_oor_po_acc_active=df_oor_stat.copy()
    df_oor_po_acc_active=df_oor_po_acc_active[(~df_oor_po_acc_active['oor_status'].str.lower().str.strip().isin(["shipped","cancelled"]))
                                & (df_oor_po_acc_active['family'].str[0:5].str.lower()=='acces')]
    df_oor_po_acc_active=apply_grouping(df_oor_po_acc_active,prefix='po_active_acc_',key_cols=key_cols,list_cols=list_cols,qty_cols=qty_cols,date_cols=date_cols)
    # Propiedades de fixturas
    df_oor_po_fix_active=df_oor_stat.copy()
    df_oor_po_fix_active=df_oor_po_fix_active[(~df_oor_po_fix_active['oor_status'].str.lower().str.strip().isin(["shipped","cancelled"]))
                                & (df_oor_po_fix_active['family'].str[0:5].str.lower()!='acces')]
    df_oor_po_fix_active=apply_grouping(df_oor_po_fix_active,prefix='po_active_fix_',key_cols=key_cols,list_cols=list_cols,qty_cols=qty_cols,date_cols=date_cols)

    # Merge
    df_oor_stat=df_oor_stat.merge(df_oor_by_po_active.df,how='left',on=key_cols)
    df_oor_stat=df_oor_stat.merge(df_oor_po_acc_active.df,how='left',on=key_cols)
    df_oor_stat=df_oor_stat.merge(df_oor_po_fix_active.df,how='left',on=key_cols)


    # Actualizar listas en columns and formatting.xlsx
    col_options={}
    for col in list_cols + df_oor_by_po_active.list_cols + df_oor_po_acc_active.list_cols + df_oor_po_fix_active.list_cols:
        col_options[col]=df_oor_stat[col].dropna().drop_duplicates().to_list()
        if 'blank' not in col_options[col]:
            col_options[col]=col_options[col]+['blank']
    # Quantity and date
    for col in qty_cols + df_oor_by_po_active.qty_cols + df_oor_po_acc_active.qty_cols + df_oor_po_fix_active.qty_cols:
        col_options[col]=["(quantity)"]
    for col in date_cols + df_oor_by_po_active.date_cols + df_oor_po_acc_active.date_cols + df_oor_po_fix_active.date_cols:
        col_options[col]=["(date)","today + X days","no date"]

    col_options['columns_list']=sorted(col_options.keys())
    # Agregar opciones
    wb_cols=load_workbook(output_paths['path_xl_format'])
    ws=wb_cols['options']
    dict_opt=get_worksheet_df(ws,data_only=True)
    # Add options
    for col,options in col_options.items():
        if col in dict_opt['df'].columns:
            df=dict_opt['df'][[col]]
            old_opt=df[~df[col].isnull()][col].astype(str).tolist()
        else:
            old_opt=[]
        new_opt=[str(option) for option in options]
        new_opt=sorted(set(old_opt+new_opt))
        df_new_opt=pd.DataFrame(columns=[col],data=new_opt)
        if col in dict_opt['header']:
            dict_opt['df']=pd.concat([dict_opt['df'].drop(columns=[col]),df_new_opt],axis=1)
        else:
            dict_opt['df']=pd.concat([dict_opt['df'],df_new_opt],axis=1)

    ws=update_sheet(dict_opt,ws,apply_formats=False)
    # Create defined names based on options sheet
    dict_opt=get_worksheet_df(ws,data_only=True)
    df=dict_opt['df']
    for col in df.columns:
        if col.endswith('_cell'):
            continue
        coorda=df.loc[~df[col].isnull(),f'{col}_cell'].tolist()[0]
        coordb=df.loc[~df[col].isnull(),f'{col}_cell'].tolist()[-1]
        if col in wb_cols.defined_names:
            del(wb_cols.defined_names[col])
        wb_cols.defined_names.add(DefinedName(col, attr_text=f"options!${coorda}:${coordb}"))


    # Agregar data validation para condiciones 
    ws=wb_cols['conditions']

    dict_condit=get_worksheet_df(ws,key_text="column",data_only=True)
    if 'column_cell' in dict_condit['df'].columns:
        last_cell1=dict_condit['df']['column_cell'].iloc[-1]
        last_cell2=dict_condit['df']['options_cell'].iloc[-1]
        first_cell1=dict_condit['df']['column_cell'].iloc[0]
        first_cell2=dict_condit['df']['options_cell'].iloc[0]
    else:
        last_cell1=ws[find_cell_by_text(ws,"column")].offset(1,0).coordinate
        last_cell2=ws[find_cell_by_text(ws,"options")].offset(1,0).coordinate
        first_cell1=last_cell1
        first_cell2=last_cell2

    range1=ws[f"{ws[first_cell1].coordinate}:{ws[last_cell1].offset(11,0).coordinate}"]
    range2=ws[f"{ws[first_cell2].coordinate}:{ws[last_cell2].offset(11,0).coordinate}"]
    ws.data_validations.dataValidation=[]
    options_str = '"' + ','.join(sorted(df_oor_stat.columns)) + '"'

    dv1=DataValidation(type="list", formula1=f'=INDIRECT("columns_list")', allow_blank=True)
    ws.add_data_validation(dv1)
    dv1.add(f"{range1[0][0].coordinate}:{range1[-1][0].coordinate}")
    for cols in zip(range1,range2):
        dv2=DataValidation(type="list", formula1=f"=INDIRECT({cols[0][0].coordinate})", allow_blank=True)
        ws.add_data_validation(dv2)
        dv2.add(cols[1][0])
    save_wb(wb_cols,output_paths['path_xl_format'])


    # Actualizar status en el OOR
    df_conditions=read_excel(output_paths['path_xl_format'],sheet_name='conditions')
    df_conditions=df_conditions[df_conditions['sheet']=='OOR']
    df_oor_stat['status_new']=''
    for col in date_cols+date_cols_first+date_cols_last:
        if not col in df_oor_stat.columns:
            continue
        df_oor_stat[col] = pd.to_datetime(df_oor_stat[col], errors='coerce')
    df_oor_status_final=apply_dynamic_rules(df=df_oor_stat,rules_df=df_conditions,target_col='status_new')
    df_oor_status_final=df_oor_status_final[['po','modelo','dz','LineNumber','status_new']]
    df_oor_status_final.drop_duplicates(['po','modelo','dz','LineNumber'],inplace=True,keep='last')
    # df_oor_status_final=rename_columns(df_oor_status_final,table_to='OOR Report',sheet_to='OOR',df_col_rel=df_col_rel)
    df_oor=rename_columns(df_oor,df_col_rel,table_from='OOR Report',sheet_from='OOR')
    if 'status_new' in df_oor.columns:
        df_oor.drop(columns=['status_new'],inplace=True)
    df_oor=df_oor.merge(df_oor_status_final,how='left',on=['po','modelo','LineNumber','dz'])
    idx=(~df_oor['oor_status'].str.lower().isin(['cancelled','shipped']))
    df_oor=df_oor[idx]
    df_oor['oor_status']=df_oor['status_new']


    # Actualizar fecha: Last commit date
    df_short_dates=df_oor.copy()
    df_short_dates=df_short_dates[['po','fixt_gp_eta','acc_gp_eta','estimated_move_date_cuu']]
    df_short_dates['fixt_gp_eta_7_days']=pd.to_datetime(df_short_dates['fixt_gp_eta'], errors='coerce')+7*BDay()
    df_short_dates['acc_gp_eta_2_days']=pd.to_datetime(df_short_dates['acc_gp_eta'], errors='coerce')+2*BDay()
    df_short_dates['estimated_move_date_cuu_2_days']=pd.to_datetime(df_short_dates['estimated_move_date_cuu'], errors='coerce')+2*BDay()
    # Max dates for shortage status
    df_short_dates['max_date_short']=df_short_dates[['fixt_gp_eta_7_days','acc_gp_eta_2_days']].max(axis=1)
    # Max datefor In shipping plan Case2
    df_short_dates['max_date_ship_c2']=df_short_dates[['estimated_move_date_cuu_2_days','acc_gp_eta_2_days']].max(axis=1)
    df_short_dates.drop(columns=['fixt_gp_eta','acc_gp_eta','estimated_move_date_cuu'],inplace=True)
    df_short_dates=df_short_dates.groupby('po').max().reset_index()
    df_oor['latest_commit_date']=None
    df_oor=df_oor.merge(df_short_dates,how='left',on=['po'])


    # Update latest commit date
    df_oor.loc[df_oor['oor_status']=='Shortage','latest_commit_date']=df_oor.loc[df_oor['oor_status']=='Shortage','max_date_short']
    df_oor.loc[df_oor['oor_status']=='In shipping plan_c1','latest_commit_date']=df_oor.loc[df_oor['oor_status']=='In shipping plan_c1','estimated_move_date_cuu_2_days']
    df_oor.loc[df_oor['oor_status']=='In shipping plan_c2','latest_commit_date']=df_oor.loc[df_oor['oor_status']=='In shipping plan_c2','max_date_ship_c2']
    df_oor['oor_status']=df_oor['oor_status'].str.replace('_c1','').str.replace('_c2','')

    # Update TAT Category

    df_tat=df_oor_stat[['po','modelo','LineNumber','po_active_all_po_shipped_complete_category']].copy()
    df_tat['tat_cat']=''
    df_tat.loc[df_tat['po_active_all_po_shipped_complete_category']=='pending ship complete','tat_cat']='Open'
    df_tat.loc[df_tat['po_active_all_po_shipped_complete_category']=='po ship complete','tat_cat']='Closed'
    df_tat=df_tat[['po','modelo','LineNumber','tat_cat']].drop_duplicates(subset=['po','modelo','LineNumber'])
    if 'tat_cat' in df_oor.columns:
        df_oor.drop(columns=['tat_cat'],inplace=True)
    df_oor=df_oor.merge(df_tat,how='left',on=['po','modelo','LineNumber'])

    for idx,row in df_oor[['oor_status','Status_cell','latest_commit_date','latest_commit_date_cell','tat_cat','TAT Category_cell']].iterrows():
        ws_oor[row['Status_cell']]=row['oor_status']
        ws_oor[row['latest_commit_date_cell']]=row['latest_commit_date']
        ws_oor[row['TAT Category_cell']]=row['tat_cat']

    save_wb(wb_oor,path_oor)
    os.startfile(path_oor)

# ----------------------------------------------------------------
# Flatten Outlook Folders Helper
# ----------------------------------------------------------------
def flatten_folders(folder, indent=0):
    options = []
    prefix = " " * (indent * 4)
    option_label = f"{prefix}{folder.Name}"
    options.append((option_label, folder))
    for subfolder in folder.Folders:
        options.extend(flatten_folders(subfolder, indent + 1))
    return options

# ----------------------------------------------------------------
# Specific functions
# ----------------------------------------------------------------
def assign_quantities(df_pos, df_to_assign, additional_fields=[]):
    """
    Asigna las cantidades del segundo dataframe al primero, respetando los límites indicados en el campo Quantity del primer dataframe.
    df_pos: requiere las columnas 'PO', 'Modelo', 'Quantity' y 'Assigned'
    df_to_assign: requiere las columnas 'PO', 'Modelo', 'Quantity'
    """
    df_pos = df_pos.copy()
    df_to_assign=df_to_assign.copy()
    df_pos['Assigned'] = 0
    df_pos['quantity']=pd.to_numeric(df_pos['quantity']).astype(float)
    # Create a new DataFrame to track WorkOrder assignments
    df_assignments = []

    # Assign produced quantities respecting the limits
    for ln_index,ln_row in df_to_assign[['po','modelo']].drop_duplicates().iterrows():
        df_assign_subset = df_to_assign[(df_to_assign['po'] == ln_row['po']) & (df_to_assign['modelo'] == ln_row['modelo'])].copy()
        df_assign_subset['Remaining'] = df_assign_subset['quantity']

        for wo_index, subset_row in df_assign_subset.iterrows():
            remaining_produced = subset_row['Remaining']

            for index, row in df_pos[(df_pos['po'] == ln_row['po']) & (df_pos['modelo'] == ln_row['modelo'])].iterrows():
                if remaining_produced <= 0:
                    break
                available_space = row['quantity'] - df_pos.at[index, 'Assigned']
                assignable = min(available_space, remaining_produced)

                if assignable > 0:
                    df_pos.at[index, 'Assigned'] += assignable
                    remaining_produced -= assignable

                    # Track the assignment
                    line_assignment ={
                        # 'WO': subset_row['WO'],
                        'po': row['po'],
                        'modelo': row['modelo'],
                        'AvailQuantity': subset_row['quantity'],
                        'LineNumber': row['LineNumber'],
                        'Assigned_Quantity': assignable
                        }
                    # Other fields if exist
                    if additional_fields:
                        for field in additional_fields:
                            if field in subset_row:
                                line_assignment[field]=subset_row[field]
                    df_assignments.append(line_assignment)

    # Create the assignments DataFrame
    df_assignments = pd.DataFrame(df_assignments)
    assigned={'df_pos':df_pos,'df_assignments':df_assignments}
    return assigned

def merge_additional_fields(df=pd.DataFrame(),df_edi=pd.DataFrame(),fields=[],sort_fields=[],key=[]):
    if len(df)==0:
        df=pd.DataFrame(columns=fields)
    df=df.sort_values(sort_fields)
    df.drop_duplicates(key,keep='last',inplace=True)
    df_edi=df_edi.merge(df[fields],how='left',on=key)
    return df_edi

def map_yield_report(df):
    df_yield=df.copy()
    df_yield.columns=df_yield.iloc[8].fillna('NA').to_list()
    df['Yield Reports'].fillna('',inplace=True)
    date_from=df[df['Yield Reports'].str.contains('From')].iloc[0]['Yield Reports']
    date_to=df[df['Yield Reports'].str.contains('From')].iloc[0]['Unnamed: 5']
    df_yield['date_from']=date_from
    df_yield['date_to']=date_to
    df_yield=df_yield[(~df_yield['Part Number'].isna()) & (df_yield['Part Number']!='Part Number')]
    df_yield['date_from']=df_yield['date_from'].str.replace('From: ','')
    df_yield['date_to']=df_yield['date_to'].str.replace('To: ','')
    df_yield['date_from']=pd.to_datetime(df_yield['date_from'])
    df_yield['date_to']=pd.to_datetime(df_yield['date_to'])
    df_yield.drop('NA',axis=1,inplace=True)
    df_yield.fillna(0,inplace=True)
    # Main date is date_from
    df_yield['month']=df_yield['date_from'].dt.strftime('%Y-%m')
    return df_yield

def fill_yield_report(df,wb,sheet_name='',search_parms={}):
    ws=wb[sheet_name]
    df_data_coord=df[[search_parms['column'],search_parms['row']]].drop_duplicates()
    df_data_coord['row']=''
    df_data_coord['column']=''
    coord_y=get_locations(df_data_coord,ws,search_parms['row'])
    coord_x=get_locations(df_data_coord,ws,search_parms['column'])
    for index,row in df.iterrows():
        if not ((coord_x[row[search_parms['column']]] and coord_y[row[search_parms['row']]])):
            continue  
        cell=ws.cell(ws[coord_y[row[search_parms['row']]]].row,
                    ws[coord_x[row[search_parms['column']]]].column)
        for key in search_parms['offset'].keys():
            if not key in df.columns:
                continue
            offset=search_parms['offset'][key][row[key]]
            cell=cell.offset(offset[0],offset[1])
        cell.value=row['Total Tested']
    return wb

# Status
def apply_grouping(df,key_cols=['po'],list_cols=[],qty_cols=[],date_cols=[],prefix=''):
    """
    Returns a dataframe aggregated and with additional date columns
    """
    df[list_cols]=df[list_cols].astype(str)
    df=df[key_cols+list_cols+date_cols+qty_cols]
    # Add cols for first and last date per group
    result_dict=duplicate_df_cols(df,'_first',date_cols)
    df=result_dict['df']
    date_cols_first=result_dict['cols_suffixed']
    result_dict=duplicate_df_cols(df,'_last',date_cols)
    df=result_dict['df']
    date_cols_last=result_dict['cols_suffixed']
    # Format qty columns
    df[qty_cols]=df[qty_cols].astype(float)
    list_col_agg_dict = {str(x): lambda x: ','.join(x.drop_duplicates().sort_values().astype(str)) for x in list_cols}
    qty_col_agg_dict = {str(x): 'sum' for x in qty_cols}
    date_col_first_agg_dict = {str(x): lambda x: pd.to_datetime(x, errors='coerce').min() for x in date_cols_first}
    date_col_last_agg_dict = {str(x): lambda x: pd.to_datetime(x, errors='coerce').max() for x in date_cols_last}
    df=df.groupby(key_cols).agg(
        {**date_col_first_agg_dict, 
        **date_col_last_agg_dict, 
        **list_col_agg_dict, 
        **qty_col_agg_dict}
    )
    df = df.rename(columns={col: f"{prefix}{col}" for col in df.columns if col not in key_cols})
    df.reset_index(inplace=True)
    ns = SimpleNamespace(**{"df":df,
                            "date_cols":[f"{prefix}{col}" for col in date_cols_first]+[f"{prefix}{col}" for col in date_cols_last],
                            "list_cols":[f"{prefix}{col}" for col in list_cols],
                            "qty_cols":[f"{prefix}{col}" for col in qty_cols]                                        
                            })
    
    return ns

def apply_dynamic_rules(df, rules_df, target_col='status'):
    """
    Apply dynamic rules to a DataFrame based on a rules DataFrame using df.query().

    For each unique 'group' in rules_df, this function:
      1. Constructs a query expression by grouping rules. The grouping heuristic is:
         - Each rule is converted to a query expression.
         - If a rule’s linking condition (in 'condition') is non-empty, its expression is
           paired with the very next rule.
         - Then, if the next rule’s linking condition is the same as the previous one, they are
           grouped together; otherwise, if the current group has only one expression, we still join
           the two rules, but if more than one rule is already grouped, a new group is started.
         - Finally, the groups are joined with "and" to form the final query.
      2. Uses df.query() with the built query and its variable dictionary to filter rows.
      3. Updates df[target_col] with the group's result for the matching rows.

    Parameters
    ----------
    df : pandas.DataFrame
        The DataFrame whose rows are to be updated.
    rules_df : pandas.DataFrame
        A DataFrame of rules with columns:
            - group: Identifier for grouping rules.
            - column: The DataFrame column to evaluate.
            - result: The value to assign to df[target_col] if the rule group matches.
            - test: The comparison operator as a string ('=', '!=', '<', '<=', '>', '>=').
            - options: The value to compare against (e.g., 'today-2 workdays', '0', etc.).
            - condition: Logical operator used to link the rule with the next one ('and' or 'or').
    target_col : str, optional
        The column in df to update when the rules match. Default is 'status'.

    Returns
    -------
    pandas.DataFrame
        The updated DataFrame (modifications are done in-place).

    Raises
    ------
    ValueError
        If a rule group contains multiple distinct 'result' values or if an unsupported
        operator is encountered.
    """
    # Group the rules by the 'group' column.
    df=df.copy()
    rules_df = rules_df.sort_values(['priority', 'group'])
    grouped_rules = rules_df.groupby(['priority','group'])
    
    for group_name, group_data in grouped_rules:
        possible_results = group_data['result'].unique()
        if len(possible_results) > 1:
            raise ValueError(f"Multiple 'result' values in group '{group_name}': {possible_results}")
        group_result = possible_results[0]
        
        # Build the query expression and variable dictionary using the new grouping heuristic.
        query_expr, var_dict = build_group_query(group_data)
        
        # Use df.query() to select rows matching the expression.
        selected = df.query(query_expr, local_dict=var_dict)
        
        # Update the target column for the matching rows.
        df.loc[selected.index, target_col] = group_result
    
    return df

def build_group_query(group_data):
    """
    Construct a query expression string (and a variable dictionary) for use with df.query().

    Each rule in group_data is converted into an expression of the form:
         `column` <operator> @var
    where the operator is mapped (with '=' converted to '==') and @var is a placeholder
    for the parsed value.

    The grouping heuristic is modified to use a new column, 'condition_group', to organize
    rules into separate groups. Within each group, the individual expressions are linked
    using the condition provided in each tuple. Finally, the groups are combined with "and"
    to produce the final query expression.

    Parameters
    ----------
    group_data : pandas.DataFrame
        A DataFrame containing rules for a specific group. Expected columns include:
        [column, test, options, condition, condition_group].

    Returns
    -------
    tuple
        (query_expr, var_dict) where:
          - query_expr is a string suitable for df.query().
          - var_dict is a dictionary mapping variable names (e.g. _v0, _v1, ...) to their parsed values.
    """
    # Map from rule operator to Python operator for query strings.
    op_map = {'=': '==', '!=': '!=', '<': '<', '<=': '<=', '>': '>', '>=': '>='}
    
    var_dict = {}
    var_counter = 0
    # Build a dictionary grouping the rules by 'condition_group'
    rules_dict = {}
    
    for _, row in group_data.iterrows():
        parsed_val = parse_value(row['options'])
        var_name = f"_v{var_counter}"
        var_counter += 1
        var_dict[var_name] = parsed_val

        operator = op_map.get(row['test'])
        if operator is None:
            raise ValueError(f"Unsupported test operator: {row['test']}")
        # Use backticks around column names to safeguard against spaces/reserved words.
        expr_str = f"`{row['column']}` {operator} @{var_name}"
        # if parsed_val=='no date':
        #     expr_str = f"`{row['column']}`.isnull()"
        cond = str(row['condition']).strip().lower() if pd.notnull(row['condition']) else ""
        
        # Group the rules by the new column 'condition_group'
        key = row['condition_group']
        if key in rules_dict:
            rules_dict[key].append((expr_str, cond))
        else:
            rules_dict[key] = [(expr_str, cond)]
    
    # Build the expression for each group by joining the expressions inside the group
    group_expressions = []
    for group_key, rule_list in rules_dict.items():
        # Start with the first expression
        group_expr = rule_list[0][0]
        # For each subsequent rule, join it using the previous rule's linking condition.
        for i in range(1, len(rule_list)):
            # Use the linking operator from the previous tuple; if empty, default to "and"
            link_op = rule_list[i-1][1] if rule_list[i-1][1] else "and"
            group_expr += f" {link_op} " + rule_list[i][0]
        # If more than one expression exists, wrap the group in parentheses.
        if len(rule_list) > 1:
            group_expr = f"({group_expr})"
        group_expressions.append(group_expr)
    
    # Combine all groups with "and"
    full_expr = " and ".join(group_expressions)
    return full_expr, var_dict

def parse_value(value):
    """
    Convert a string representation of a value into an appropriate Python object.

    Special cases handled:
      - 'today-2 workdays': returns the date two business days ago.
      - 'today-2': returns the date two calendar days ago.
    If the string can be converted to a float, it is; otherwise, the original string is returned.

    Parameters
    ----------
    value : any
        The value to parse (typically a string).

    Returns
    -------
    any
        The parsed value (float, datetime, or string).
    """
    if isinstance(value, str):
        val_lower = value.lower().strip()
        if val_lower=='no date':
            return pd.Timestamp('2099-12-31')
        if val_lower.startswith('today +'):
            remainder = val_lower[len('today +'):]
            parts = remainder.split()
            try:
                offset_n = int(parts[0])
            except (ValueError, IndexError):
                raise ValueError(f"Cannot parse numeric offset from '{value}'")
            if len(parts) > 1 and 'workday' in parts[1]:
                return pd.Timestamp.now().normalize() + offset_n * BDay()
            else:
                return pd.Timestamp.now().normalize() + pd.Timedelta(days=offset_n)
        else:
            try:
                return float(value)
            except ValueError:
                return value
    else:
        return value
    
# ----------------------------------------------------------------
# Function: Explorar Outlook (with Excel management and file integration)
# ----------------------------------------------------------------
def explorar_outlook():
    st.write("Ejecutando función: Explorar Outlook")
    # Ensure a folder was selected from Outlook
    if "mail_folder" not in st.session_state or st.session_state.mail_folder is None:
        st.error("No se ha seleccionado un folder de Outlook.")
        return
    # Ensure the attachments folder exists
    if not os.path.exists(output_paths.get("path_attachments", "")):
        os.makedirs(output_paths['path_attachments'])
    # Get the date limit from state (fecha_mail)
    if state.get("fecha_mail") is None:
        st.error("La fecha mail no está definida en el estado.")
        return
    date_limit = datetime.combine(state["fecha_mail"], datetime.min.time())
    folder = st.session_state.mail_folder
    try:
        items = folder.Items
    except Exception as e:
        st.error("Error accediendo a los elementos del folder de Outlook.")
        return
    count = 0
    for message in items:
        try:
            received_time = message.ReceivedTime
            # Convert COM date to Python datetime
            received_time = datetime.fromtimestamp(received_time.timestamp())
        except Exception:
            continue
        if received_time < date_limit:
            continue
        if message.Attachments.Count > 0:
            for i in range(1, message.Attachments.Count + 1):
                try:
                    attachment = message.Attachments.Item(i)
                    if "KORRUS" not in attachment.FileName.upper():
                        continue
                    base, ext = os.path.splitext(attachment.FileName)
                    new_name = f"{base}_{received_time.strftime('%Y-%m-%d-%H%M%S')}{ext}"
                    save_path = os.path.join(output_paths['path_attachments'], new_name)
                    attachment.SaveAsFile(save_path)
                    st.write(f"Attachment saved: {save_path}")
                    count += 1
                except Exception as e:
                    st.write(f"Error saving attachment: {str(e)}")
                    continue
    st.success(f"Exploración completada. {count} attachments guardados.")
    
    # Crear la lista de archivos descargados para manejar su integración.
    if not os.path.exists(output_paths['path_attachments_done']):
        os.makedirs(output_paths['path_attachments_done'])
    
    # Status maneja el proceso a realizar en los archivos: r=reprocesar, d=done, e=error
    if os.path.exists(output_paths['path_korrus_list']):
        df_korrus_list = pd.read_excel(output_paths['path_korrus_list'])
    else:
        df_korrus_list = pd.DataFrame(columns=['file_name', 'received_time', 'status'])
    close_xl_if_open(output_paths['path_korrus_list'])
    
    files = os.listdir(output_paths['path_attachments'])
    df_korrus_list_new = pd.DataFrame(columns=['file_name'], data=files)
    df_korrus_list_new['received_time'] = df_korrus_list_new['file_name'].str.extract(r'(\d{4}-\d{2}-\d{2}-\d{6})')
    df_korrus_list_new = df_korrus_list_new.merge(df_korrus_list, how='outer', on=['file_name', 'received_time'])
    df_korrus_list_new.dropna(subset=['received_time'], inplace=True)
    df_korrus_list_new['status'].fillna('', inplace=True)
    
    # Analizar archivos adjuntos
    df_korrus_data_new = pd.DataFrame(columns=['origin_file'] + mandatory_cols['Korrus'])
    for index, row in df_korrus_list_new[df_korrus_list_new['status'].isin(['', 'r'])].iterrows():
        filepath = os.path.join(output_paths['path_attachments'], row['file_name'])
        if (row['status'] == 'r') or (not os.path.exists(filepath)):
            filepath = os.path.join(output_paths['path_attachments_done'], row['file_name'])
        if not os.path.exists(filepath):
            df_korrus_list_new.loc[index, 'status'] = 'e'
            st.info(f"El archivo {row['file_name']} no existe")
            continue
        ext = os.path.splitext(os.path.basename(filepath))[1].lower()
        if ext == ".xlsx":
            df = read_excel(filepath)
        elif ext == ".csv":
            try:
                df = pd.read_csv(filepath, sep='\t', encoding='utf-16')
            except:
                df = pd.read_csv(filepath, sep=',')
        else:
            df_korrus_list_new.loc[index, 'status'] = 'e'
            st.info(f"El archivo {row['file_name']} no pudo ser procesado")
            continue
        if check_mandatory_cols(df.columns, selector_name='Korrus', raise_error=False):
            df['origin_file'] = row['file_name']
            df_korrus_data_new = pd.concat([df_korrus_data_new, df])
            df_korrus_data_new = df_korrus_data_new[['origin_file'] + mandatory_cols['Korrus']]
            df_korrus_list_new.loc[index, 'status'] = 'n'
        else:
            df_korrus_list_new.loc[index, 'status'] = 'e'
    
    # Consolidar datos en un solo archivo
    if os.path.exists(output_paths['path_korrus_data']):
        df_korrus_data = pd.read_excel(output_paths['path_korrus_data'])
    else:
        df_korrus_data = pd.DataFrame(columns=['origin_file'] + mandatory_cols['Korrus'])
    close_xl_if_open(output_paths['path_korrus_data'])
    df_korrus_data = df_korrus_data[~df_korrus_data['origin_file'].isin(df_korrus_data_new['origin_file'])]
    df_korrus_data_new = pd.concat([df_korrus_data, df_korrus_data_new])
    # Mover archivos a la carpeta de archivos procesados
    for index,row in df_korrus_list_new[df_korrus_list_new['status']=='n'].iterrows():
        df_korrus_list_new.loc[index,'status']='y'
        if os.path.exists(f'{output_paths['path_attachments']}/{row["file_name"]}'):
            shutil.move(f'{output_paths['path_attachments']}/{row["file_name"]}', f'{output_paths['path_attachments_done']}/{row["file_name"]}')
    # Eliminar archivos ya procesados del folder de llegada
    df_remove=df_korrus_list_new[(df_korrus_list_new['status']=='y') & (df_korrus_list_new['file_name'].isin(files))]
    for index,row in df_remove.iterrows():
        if os.path.exists(os.path.join(output_paths['path_attachments'],row['file_name'])):
            os.remove(os.path.join(output_paths['path_attachments'],row['file_name']))
    files=os.listdir(output_paths['path_attachments'])
    # Guardar datos y lista de archivos procesados
    save_df(df_korrus_list_new,filepath=output_paths['path_korrus_list'],sheet_name='Korrus List',index=False)
    save_df(df_korrus_data_new,filepath=output_paths['path_korrus_data'],sheet_name='Korrus Data',index=False)
    st.success("Proceso de integración completado.")

# ----------------------------------------------------------------
# Streamlit UI
# ----------------------------------------------------------------
st.set_page_config(page_title="Seguimiento a Embarques", page_icon=":truck:")
st.title("Seguimiento a Embarques")

# Load state and update if needed
state = load_state_pickle()

# Selección de carpeta de salidas
st.header("Seleccionar carpeta de trabajo")
if st.button("Seleccionar carpeta", key="select_folder"):
    folder = select_directory(initialdir=state.get("folder_output", os.getcwd()))
    if folder:
        state["folder_output"] = folder
        save_state_pickle(state)
        st.rerun()
if state.get("folder_output"):
    st.success(f"Carpeta de trabajo: {state['folder_output']}")
else:
    st.info("No se ha seleccionado carpeta de trabajo.")

if state.get("folder_output"):
    set_paths(state["folder_output"])
    

# Selección de folder de Outlook con gestión de estado
st.header("Outlook folders")
pythoncom.CoInitialize()
outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
inbox = outlook.GetDefaultFolder(6)
folder_options = flatten_folders(inbox)
labels = [opt[0] for opt in folder_options]
default_label = state.get("outlook_folder") if state.get("outlook_folder") in labels else labels[0]
selected_label = st.selectbox("Selecciona una carpeta de Outlook", labels, index=labels.index(default_label))
if selected_label != state.get("outlook_folder"):
    state["outlook_folder"] = selected_label
    save_state_pickle(state)
for label, folder in folder_options:
    if label == selected_label:
        st.session_state.mail_folder = folder
        break
st.write("Carpeta seleccionada:", st.session_state.mail_folder.Name)
# pythoncom.CoUninitialize()

# Selección y estado de las fechas con gestión de estado
st.header("Fechas")
fecha_mail_input = st.date_input("Fecha mail", value=state.get("fecha_mail") or (date.today() - timedelta(days=1)))
if fecha_mail_input != state.get("fecha_mail"):
    state["fecha_mail"] = fecha_mail_input
    save_state_pickle(state)

fecha_shipments_input = st.date_input("ELP Shipments", value=state.get("fecha_shipments_elp") or date.today())
if fecha_shipments_input != state.get("fecha_shipments_elp"):
    state["fecha_shipments_elp"] = fecha_shipments_input
    save_state_pickle(state)

fecha_freeze_input = st.date_input("Fecha Inicial de Actualización", value=state.get("fecha_freeze") or date.today())
if fecha_freeze_input != state.get("fecha_freeze"):
    state["fecha_freeze"] = fecha_freeze_input
    save_state_pickle(state)

# Selección de archivos mediante manage_file_selector
st.header("Seleccionar archivos")
for file_key in ['OOR', 'Tracker', 'ELP Master', 'Shipment transactions', 'InventoryStageBakup', 'OH Max', 'Prices', 'Gating Parts']:
    manage_file_selector(file_key, file_key, state)

# Selección de carpeta de yield reports
st.header("Seleccionar carpeta de Yield reports")
if st.button("Seleccionar carpeta", key="select_yield"):
    folder = select_directory(initialdir=state.get("folder_yield", os.getcwd()))
    if folder:
        state["folder_yield"] = folder
        save_state_pickle(state)
        st.rerun()
if state.get("folder_yield"):
    st.success(f"Carpeta de Yield reports: {state['folder_yield']}")
else:
    st.info("No se ha seleccionado carpeta de Yield reports.")

# # Agregar 5 botones para ejecutar funciones específicas
# st.header("Procesos")
# col1, col2, col3, col4, col5 = st.columns(5)
# with col1:
#     if st.button("Explorar Outlook"):
#         explorar_outlook()
# with col2:
#     if st.button("Generar Reportes"):
#         generar_reportes()
# with col3:
#     if st.button("AAR"):
#         fill_aar()
# with col4:
#     if st.button("Gating Parts"):
#         gating_parts()
# with col5:
#     if st.button("Actualizar Status"):
#         actualizar_status()
        
st.header("Procesos")

if st.button("Explorar Outlook", key="explorar_outlook"):
    explorar_outlook()
    st.session_state.explorar_msg = "Exploración completada."
if "explorar_msg" in st.session_state:
    st.write(st.session_state.explorar_msg)

if st.button("Generar Reportes", key="generar_reportes"):
    generar_reportes()
    st.session_state.reportes_msg = "Reportes generados."
if "reportes_msg" in st.session_state:
    st.write(st.session_state.reportes_msg)

if st.button("AAR", key="aar"):
    fill_aar()
    st.session_state.aar_msg = "AAR finalizado."
if "aar_msg" in st.session_state:
    st.write(st.session_state.aar_msg)

if st.button("Gating Parts", key="gating_parts"):
    gating_parts()
    st.session_state.gating_msg = "Gating Parts completado."
if "gating_msg" in st.session_state:
    st.write(st.session_state.gating_msg)

if st.button("Actualizar Status", key="actualizar_status"):
    actualizar_status()
    st.session_state.status_msg = "Status actualizado."
if "status_msg" in st.session_state:
    st.write(st.session_state.status_msg)
